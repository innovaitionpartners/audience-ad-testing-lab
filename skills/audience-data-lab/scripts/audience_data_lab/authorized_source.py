"""Safe, bounded profiling for authorized audience-source bundles.

This module deliberately reports structural evidence only. It does not map,
join, transform, or retain source rows and it never emits observed values.
"""

from __future__ import annotations

import csv
from datetime import date, datetime
import json
from pathlib import Path
import re
import unicodedata
from typing import Any, Iterable, Mapping, Sequence
import zipfile

from openpyxl import load_workbook

from .common import (
    ContractError,
    exact_object,
    require_enum,
    require_boolean,
    require_identifier,
    require_integer,
    require_number,
    require_string,
    require_string_list,
    require_timestamp,
    sha256_file,
    sha256_json,
)


AUTHORIZED_SOURCE_PROFILE_VERSION = "authorized-audience-source-profile-v1"
MAX_INPUT_FILES = 100
MAX_BUNDLE_BYTES = 250 * 1024 * 1024
MAX_WORKBOOK_SHEETS = 25
MAX_TABLE_ROWS = 1_000_000
MAX_TABLE_COLUMNS = 10_000
MAX_WORKBOOK_CELLS = 5_000_000

_TOP_LEVEL_KEYS = {
    "schema_version", "profile_id", "profile_version", "profiled_at",
    "bundle_sha256", "inputs", "tables", "relationships",
    "candidate_semantics", "privacy_risk", "unresolved", "decision",
}
_DECISION_KEYS = {"status", "allowed_next_route", "reasons"}
_INPUT_KEYS = {"display_name", "format", "bytes", "sha256", "workbook_metadata"}
_WORKBOOK_METADATA_KEYS = {"sheet_count", "has_external_links", "has_vba"}
_TABLE_KEYS = {
    "file", "sheet", "record_path", "shape", "row_count", "column_count",
    "field_names", "observed_scalar_types", "null_rates", "sample_safe_value_classes",
    "candidate_units", "candidate_denominators", "candidate_field_roles",
}
_RELATIONSHIP_KEYS = {"kind", "field", "files"}
_SEMANTIC_KEYS = {"file", "field", "candidate_roles", "requires_confirmation"}
_RISK_KEYS = {"code", "field", "source"}
_ISSUE_KEYS = {"code", "detail"}
_STATUSES = {
    "ready_for_mapping", "needs_clarification",
    "requires_private_aggregation", "rejected",
}
_ROUTES = {"aggregate_transform", "private_aggregation", "none"}
_PERSON_FIELD_RISKS = (
    (re.compile(r"(^|_)(e_?mail|emailaddress)($|_)"), "email"),
    (re.compile(r"(^|_)(phone|mobile|telephone|tel)($|_)"), "phone"),
    (re.compile(r"(^|_)(username|user_name|handle)($|_)"), "username_handle"),
    (re.compile(r"(^|_)(first_name|last_name|full_name|name)($|_)"), "name"),
    (re.compile(r"(^|_)(address|street|postal|zip|postcode)($|_)"), "postal_address"),
    (re.compile(r"(^|_)(device_id|deviceid)($|_)"), "device_id"),
    (re.compile(r"(^|_)(cookie_id|cookieid)($|_)"), "cookie_id"),
    (re.compile(r"(^|_)(ad_id|advertising_id|advertisingid|idfa|aaid)($|_)"), "advertising_id"),
    (re.compile(r"(^|_)(ip|ip_address|ipaddress)($|_)"), "ip_address"),
    (re.compile(r"(^|_)(account_id|accountid)($|_)"), "account_id"),
)
_EVENT_FIELD = re.compile(r"(^|_)(event|transaction|purchase|order)_?(id|identifier)($|_)")
_EMAIL_VALUE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_PHONE_VALUE = re.compile(r"^\+?[0-9][0-9 .()/-]{6,}$")
_IP_VALUE = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$")
_DATE_PART = re.compile(r"^(\d{1,4})([-/])(\d{1,2})\2(\d{1,4})$")
_ISO_DATETIME = re.compile(
    r"^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}"
    r"(?::\d{2}(?:\.\d{1,6})?)?(?:Z|[+-]\d{2}:\d{2})?$"
)
_IDENTIFIER_VALUE = re.compile(r"^[a-z][a-z0-9]*(?:[-_][a-z0-9]+)+$", re.I)


def validate_source_profile(payload: object) -> dict[str, object]:
    """Validate the stable profile envelope and its routing decision."""

    profile = exact_object(payload, _TOP_LEVEL_KEYS, "source profile")
    if profile["schema_version"] != AUTHORIZED_SOURCE_PROFILE_VERSION:
        raise ContractError("source profile.schema_version is not supported")
    require_identifier(profile["profile_id"], "source profile.profile_id")
    require_string(profile["profile_version"], "source profile.profile_version")
    require_timestamp(profile["profiled_at"], "source profile.profiled_at")
    bundle_sha = require_string(profile["bundle_sha256"], "source profile.bundle_sha256")
    if not bundle_sha.startswith("sha256:") or len(bundle_sha) != 71:
        raise ContractError("source profile.bundle_sha256 must be a SHA-256 digest")
    _require_digest(bundle_sha, "source profile.bundle_sha256")
    _validate_inputs(profile["inputs"])
    _validate_tables(profile["tables"])
    _validate_relationships(profile["relationships"])
    _validate_semantics(profile["candidate_semantics"])
    _validate_risks(profile["privacy_risk"])
    _validate_issues(profile["unresolved"])
    decision = exact_object(profile["decision"], _DECISION_KEYS, "source profile.decision")
    status = require_enum(decision["status"], _STATUSES, "source profile.decision.status")
    route = require_enum(decision["allowed_next_route"], _ROUTES, "source profile.decision.allowed_next_route")
    if not isinstance(decision["reasons"], list) or not all(isinstance(item, str) for item in decision["reasons"]):
        raise ContractError("source profile.decision.reasons must be an array of strings")
    expected_routes = {
        "ready_for_mapping": "aggregate_transform",
        "needs_clarification": "aggregate_transform",
        "requires_private_aggregation": "private_aggregation",
        "rejected": "none",
    }
    if route != expected_routes[status]:
        raise ContractError("source profile.decision.allowed_next_route conflicts with status")
    return profile


def _require_digest(value: object, path: str) -> str:
    digest = require_string(value, path)
    if not re.fullmatch(r"sha256:[0-9a-f]{64}", digest):
        raise ContractError(f"{path} must be a SHA-256 digest")
    return digest


def _require_array(value: object, path: str) -> list[object]:
    if not isinstance(value, list):
        raise ContractError(f"{path} must be an array")
    return value


def _validate_inputs(value: object) -> None:
    for index, item in enumerate(_require_array(value, "source profile.inputs")):
        input_item = exact_object(item, _INPUT_KEYS, f"source profile.inputs[{index}]")
        require_string(input_item["display_name"], f"source profile.inputs[{index}].display_name")
        require_enum(input_item["format"], {"csv", "json", "xlsx", "xls", "xlsm", "unknown"}, f"source profile.inputs[{index}].format")
        require_integer(input_item["bytes"], f"source profile.inputs[{index}].bytes", minimum=0)
        _require_digest(input_item["sha256"], f"source profile.inputs[{index}].sha256")
        metadata = input_item["workbook_metadata"]
        if metadata is not None:
            metadata = exact_object(metadata, _WORKBOOK_METADATA_KEYS, f"source profile.inputs[{index}].workbook_metadata")
            require_integer(metadata["sheet_count"], f"source profile.inputs[{index}].workbook_metadata.sheet_count", minimum=0)
            require_boolean(metadata["has_external_links"], f"source profile.inputs[{index}].workbook_metadata.has_external_links")
            require_boolean(metadata["has_vba"], f"source profile.inputs[{index}].workbook_metadata.has_vba")


def _validate_tables(value: object) -> None:
    for index, item in enumerate(_require_array(value, "source profile.tables")):
        table = exact_object(item, _TABLE_KEYS, f"source profile.tables[{index}]")
        base = f"source profile.tables[{index}]"
        require_string(table["file"], f"{base}.file")
        if table["sheet"] is not None:
            require_string(table["sheet"], f"{base}.sheet")
        require_string(table["record_path"], f"{base}.record_path")
        require_enum(table["shape"], {"wide", "long", "nested", "relational", "canonical"}, f"{base}.shape")
        require_integer(table["row_count"], f"{base}.row_count", minimum=0, maximum=MAX_TABLE_ROWS)
        require_integer(table["column_count"], f"{base}.column_count", minimum=0, maximum=MAX_TABLE_COLUMNS)
        fields = require_string_list(table["field_names"], f"{base}.field_names")
        if len(fields) != table["column_count"]:
            raise ContractError(f"{base}.column_count must match field_names")
        for key in ("observed_scalar_types", "null_rates", "sample_safe_value_classes", "candidate_field_roles"):
            mapping = table[key]
            if not isinstance(mapping, Mapping) or set(mapping) != set(fields):
                raise ContractError(f"{base}.{key} must have exactly one entry per field")
        for field in fields:
            require_string_list(table["observed_scalar_types"][field], f"{base}.observed_scalar_types.{field}")
            require_number(table["null_rates"][field], f"{base}.null_rates.{field}", minimum=0, maximum=1)
            require_string_list(table["sample_safe_value_classes"][field], f"{base}.sample_safe_value_classes.{field}")
            require_string_list(table["candidate_field_roles"][field], f"{base}.candidate_field_roles.{field}", nonempty=True)
        require_string_list(table["candidate_units"], f"{base}.candidate_units")
        require_string_list(table["candidate_denominators"], f"{base}.candidate_denominators")


def _validate_relationships(value: object) -> None:
    for index, item in enumerate(_require_array(value, "source profile.relationships")):
        relationship = exact_object(item, _RELATIONSHIP_KEYS, f"source profile.relationships[{index}]")
        require_enum(relationship["kind"], {"candidate_shared_field"}, f"source profile.relationships[{index}].kind")
        require_string(relationship["field"], f"source profile.relationships[{index}].field")
        require_string_list(relationship["files"], f"source profile.relationships[{index}].files", nonempty=True)


def _validate_semantics(value: object) -> None:
    for index, item in enumerate(_require_array(value, "source profile.candidate_semantics")):
        semantic = exact_object(item, _SEMANTIC_KEYS, f"source profile.candidate_semantics[{index}]")
        require_string(semantic["file"], f"source profile.candidate_semantics[{index}].file")
        require_string(semantic["field"], f"source profile.candidate_semantics[{index}].field")
        require_string_list(semantic["candidate_roles"], f"source profile.candidate_semantics[{index}].candidate_roles", nonempty=True)
        require_boolean(semantic["requires_confirmation"], f"source profile.candidate_semantics[{index}].requires_confirmation")


def _validate_risks(value: object) -> None:
    for index, item in enumerate(_require_array(value, "source profile.privacy_risk")):
        risk = exact_object(item, _RISK_KEYS, f"source profile.privacy_risk[{index}]")
        require_string(risk["code"], f"source profile.privacy_risk[{index}].code")
        require_string(risk["field"], f"source profile.privacy_risk[{index}].field", allow_empty=True)
        require_enum(risk["source"], {"field_name", "value_class", "field_combination"}, f"source profile.privacy_risk[{index}].source")


def _validate_issues(value: object) -> None:
    for index, item in enumerate(_require_array(value, "source profile.unresolved")):
        issue = exact_object(item, _ISSUE_KEYS, f"source profile.unresolved[{index}]")
        require_string(issue["code"], f"source profile.unresolved[{index}].code")
        require_string(issue["detail"], f"source profile.unresolved[{index}].detail")


def profile_authorized_bundle(
    input_paths: Sequence[Path], *, profile_id: str, profile_version: str, profiled_at: str
) -> dict[str, object]:
    """Profile authorized local sources without retaining or emitting cell values."""

    require_identifier(profile_id, "profile_id")
    require_string(profile_version, "profile_version")
    require_timestamp(profiled_at, "profiled_at")
    if not input_paths:
        raise ContractError("at least one input file is required")
    if len(input_paths) > MAX_INPUT_FILES:
        raise ContractError("at most 100 input files are allowed")
    sources = _normalize_sources(input_paths)
    if sum(int(item["bytes"]) for item in sources) > MAX_BUNDLE_BYTES:
        raise ContractError("authorized source bundle must not exceed 250 MiB")

    inputs: list[dict[str, object]] = []
    tables: list[dict[str, object]] = []
    privacy_risk: list[dict[str, object]] = []
    unresolved: list[dict[str, str]] = []
    for source in sources:
        input_item, source_tables, source_risks, source_issues = _inspect_source(source)
        inputs.append(input_item)
        tables.extend(source_tables)
        privacy_risk.extend(source_risks)
        unresolved.extend(source_issues)
    relationships = _relationships(tables)
    privacy_risk = _unique_records(privacy_risk)
    unresolved = _unique_records(unresolved)
    if unresolved:
        status, route, reasons = "rejected", "none", [item["code"] for item in unresolved]
    elif privacy_risk:
        status, route, reasons = "requires_private_aggregation", "private_aggregation", [item["code"] for item in privacy_risk]
    elif not tables:
        status, route, reasons = "needs_clarification", "aggregate_transform", ["no_structural_records"]
    else:
        status, route, reasons = "ready_for_mapping", "aggregate_transform", []
    profile: dict[str, object] = {
        "schema_version": AUTHORIZED_SOURCE_PROFILE_VERSION,
        "profile_id": profile_id,
        "profile_version": profile_version,
        "profiled_at": profiled_at,
        "bundle_sha256": sha256_json([
            {"display_name": item["display_name"], "sha256": item["sha256"]} for item in inputs
        ]),
        "inputs": inputs,
        "tables": tables,
        "relationships": relationships,
        "candidate_semantics": _candidate_semantics(tables),
        "privacy_risk": privacy_risk,
        "unresolved": unresolved,
        "decision": {"status": status, "allowed_next_route": route, "reasons": reasons},
    }
    return validate_source_profile(profile)


def _normalize_sources(input_paths: Sequence[Path]) -> list[dict[str, object]]:
    sources: list[dict[str, object]] = []
    display_names: set[str] = set()
    for original in input_paths:
        path = Path(original)
        try:
            stat = path.stat()
        except OSError as exc:
            raise ContractError(f"could not inspect authorized input '{path}': {exc}") from exc
        if not path.is_file():
            raise ContractError(f"authorized input must be a regular file: {path}")
        display_name = unicodedata.normalize("NFC", path.name).replace("\\", "/")
        if display_name in display_names:
            raise ContractError(f"normalized display-name collision: {display_name}")
        display_names.add(display_name)
        sources.append({"path": path, "display_name": display_name, "bytes": stat.st_size, "sha256": sha256_file(path)})
    return sorted(sources, key=lambda item: (str(item["display_name"]).casefold(), str(item["display_name"])))


def _inspect_source(source: dict[str, object]) -> tuple[dict[str, object], list[dict[str, object]], list[dict[str, object]], list[dict[str, str]]]:
    path = source["path"]
    assert isinstance(path, Path)
    suffix = path.suffix.casefold()
    profile_format = suffix.removeprefix(".") if suffix in {".csv", ".json", ".xlsx", ".xls", ".xlsm"} else "unknown"
    input_item: dict[str, object] = {
        "display_name": source["display_name"], "format": profile_format,
        "bytes": source["bytes"], "sha256": source["sha256"], "workbook_metadata": None,
    }
    if suffix == ".csv":
        tables, risks, issues = _inspect_csv(path, str(source["display_name"]))
    elif suffix == ".json":
        tables, risks, issues = _inspect_json(path, str(source["display_name"]))
    elif suffix == ".xlsx":
        tables, risks, issues, metadata = _inspect_xlsx(path, str(source["display_name"]))
        input_item["workbook_metadata"] = metadata
    elif suffix in {".xls", ".xlsm"}:
        tables, risks, issues = [], [], [{"code": "unsupported_format", "detail": "legacy or macro-enabled workbooks are not accepted"}]
    else:
        with path.open("rb") as handle:
            starts_zip = handle.read(4).startswith(b"PK")
        detail = "ZIP signatures require a validated .xlsx input" if starts_zip else "only .csv, .json, and .xlsx are accepted"
        tables, risks, issues = [], [], [{"code": "unsupported_format", "detail": detail}]
    return input_item, tables, risks, issues


def _inspect_csv(path: Path, display_name: str) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, str]]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(handle, dialect)
            header = next(reader, None)
            if not header:
                return [], [], [{"code": "missing_header", "detail": "CSV requires a header row"}]
            fields = [_field_name(value, "CSV header") for value in header]
            if len(fields) != len(set(fields)):
                return [], [], [{"code": "duplicate_field", "detail": "CSV header fields must be unique"}]
            if len(fields) > MAX_TABLE_COLUMNS:
                return [], [], [{"code": "table_column_limit", "detail": "table exceeds 10,000 columns"}]
            shape = "long" if {"metric", "value"}.issubset(set(fields)) else "wide"
            accumulator = _TableAccumulator(display_name, None, "$", shape, fields)
            for row in reader:
                if len(row) != len(fields):
                    accumulator._record_issue("row_width_mismatch", "CSV rows must match header width")
                    break
                accumulator.add_row(row)
                if accumulator.issues():
                    break
            return [accumulator.table()], accumulator.risks(), accumulator.issues()
    except UnicodeDecodeError:
        return [], [], [{"code": "invalid_utf8", "detail": "CSV must use UTF-8 or UTF-8-SIG"}]
    except (csv.Error, OSError):
        return [], [], [{"code": "invalid_csv", "detail": "CSV could not be inspected"}]
    except ContractError:
        return [], [], [{"code": "invalid_header", "detail": "CSV headers must be nonempty strings"}]


def _inspect_json(path: Path, display_name: str) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, str]]]:
    try:
        with path.open("r", encoding="utf-8") as handle:
            return _JsonStructuralParser(handle, display_name).parse()
    except UnicodeDecodeError:
        return [], [], [{"code": "invalid_utf8", "detail": "JSON must use UTF-8"}]
    except (_JsonParseError, OSError):
        return [], [], [{"code": "invalid_json", "detail": "JSON must be one valid object or array"}]


class _JsonParseError(ValueError):
    """Raised for malformed JSON without exposing its contents."""


_JSON_COMPLEX = object()


class _JsonStructuralParser:
    """One-pass stdlib JSON reader that retains only the current record."""

    def __init__(self, handle: object, display_name: str):
        self.handle = handle
        self.display_name = display_name
        self.buffer = ""
        self.index = 0
        self.tables: list[_DynamicTableAccumulator] = []
        self.issues: list[dict[str, str]] = []

    def parse(self) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, str]]]:
        self._skip_whitespace()
        token = self._peek()
        if token == "{":
            root = self._object("$", capture=True)
            if not self.tables and root:
                canonical = _DynamicTableAccumulator(self.display_name, None, "$", "canonical")
                canonical.add_mapping(root)
                self.tables.append(canonical)
        elif token == "[":
            self._array("$")
        else:
            raise _JsonParseError("JSON top level is not object or array")
        self._skip_whitespace()
        if self._peek():
            raise _JsonParseError("JSON contains trailing content")
        tables = [table.table() for table in self.tables]
        risks: list[dict[str, object]] = []
        for table in self.tables:
            risks.extend(table.risks())
            self.issues.extend(table.issues())
        return tables, risks, self.issues

    def _object(self, path: str, *, capture: bool) -> dict[str, object]:
        self._expect("{")
        values: dict[str, object] = {}
        self._skip_whitespace()
        if self._peek() == "}":
            self._consume()
            return values
        while True:
            self._skip_whitespace()
            key = self._string()
            self._skip_whitespace()
            self._expect(":")
            value = self._value(f"{path}.{key}", capture=capture)
            if capture:
                if isinstance(value, dict):
                    for nested_key, nested_value in value.items():
                        self._capture_value(values, f"{key}.{nested_key}", nested_value)
                else:
                    self._capture_value(values, key, value)
            self._skip_whitespace()
            separator = self._consume()
            if separator == "}":
                return values
            if separator != ",":
                raise _JsonParseError("object field separator is invalid")

    def _capture_value(self, values: dict[str, object], key: str, value: object) -> None:
        if key in values:
            self._record_issue("nested_path_collision", "nested JSON scalar paths must not collide with literal keys")
            return
        values[key] = value

    def _array(self, path: str) -> object:
        self._expect("[")
        self._skip_whitespace()
        if self._peek() == "]":
            self._consume()
            return _JSON_COMPLEX
        if self._peek() == "{":
            accumulator = _DynamicTableAccumulator(
                self.display_name, None, path.removeprefix("$.") or "$",
                "nested" if path != "$" else "wide",
            )
            self.tables.append(accumulator)
            while True:
                if self._peek() != "{":
                    self._record_issue("mixed_json_array", "record arrays must contain objects")
                    self._value(f"{path}[]", capture=False)
                else:
                    if accumulator.issues():
                        self._object(f"{path}[]", capture=False)
                    else:
                        accumulator.add_mapping(self._object(f"{path}[]", capture=True))
                self._skip_whitespace()
                separator = self._consume()
                if separator == "]":
                    return _JSON_COMPLEX
                if separator != ",":
                    raise _JsonParseError("array item separator is invalid")
                self._skip_whitespace()
        while True:
            self._value(f"{path}[]", capture=False)
            self._skip_whitespace()
            separator = self._consume()
            if separator == "]":
                return _JSON_COMPLEX
            if separator != ",":
                raise _JsonParseError("array item separator is invalid")
            self._skip_whitespace()

    def _record_issue(self, code: str, detail: str) -> None:
        if not any(item["code"] == code for item in self.issues):
            self.issues.append({"code": code, "detail": detail})

    def _value(self, path: str, *, capture: bool) -> object:
        self._skip_whitespace()
        token = self._peek()
        if token == "{":
            return self._object(path, capture=capture) if capture else self._object(path, capture=False)
        if token == "[":
            return self._array(path)
        if token == '"':
            return self._string()
        if token in "-0123456789":
            return self._number()
        if token == "t":
            self._literal("true")
            return True
        if token == "f":
            self._literal("false")
            return False
        if token == "n":
            self._literal("null")
            return None
        raise _JsonParseError("JSON value is invalid")

    def _string(self) -> str:
        self._expect('"')
        characters: list[str] = []
        while True:
            token = self._consume()
            if not token:
                raise _JsonParseError("unterminated JSON string")
            if token == '"':
                return "".join(characters)
            if token == "\\":
                escaped = self._consume()
                replacements = {"\"": '"', "\\": "\\", "/": "/", "b": "\b", "f": "\f", "n": "\n", "r": "\r", "t": "\t"}
                if escaped == "u":
                    digits = "".join(self._consume() for _ in range(4))
                    if len(digits) != 4 or not re.fullmatch(r"[0-9a-fA-F]{4}", digits):
                        raise _JsonParseError("invalid JSON unicode escape")
                    characters.append(chr(int(digits, 16)))
                elif escaped in replacements:
                    characters.append(replacements[escaped])
                else:
                    raise _JsonParseError("invalid JSON string escape")
            elif ord(token) < 0x20:
                raise _JsonParseError("control character in JSON string")
            else:
                characters.append(token)

    def _number(self) -> int | float:
        characters: list[str] = []
        while self._peek() and self._peek() not in " \t\r\n,]}":
            characters.append(self._consume())
        text = "".join(characters)
        if not re.fullmatch(r"-?(?:0|[1-9]\d*)(?:\.\d+)?(?:[eE][+-]?\d+)?", text):
            raise _JsonParseError("invalid JSON number")
        return float(text) if any(marker in text for marker in ".eE") else int(text)

    def _literal(self, expected: str) -> None:
        if "".join(self._consume() for _ in expected) != expected:
            raise _JsonParseError("invalid JSON literal")

    def _skip_whitespace(self) -> None:
        while (token := self._peek()) and token in " \t\r\n":
            self._consume()

    def _expect(self, expected: str) -> None:
        if self._consume() != expected:
            raise _JsonParseError(f"expected JSON token {expected}")

    def _consume(self) -> str:
        token = self._peek()
        if token:
            self.index += 1
        return token

    def _peek(self) -> str:
        while self.index >= len(self.buffer):
            chunk = self.handle.read(64 * 1024)
            if not chunk:
                return ""
            self.buffer = chunk
            self.index = 0
        return self.buffer[self.index]


def _inspect_xlsx(path: Path, display_name: str) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, str]], dict[str, object]]:
    metadata: dict[str, object] = {"sheet_count": 0, "has_external_links": False, "has_vba": False}
    issues = _xlsx_container_issues(path)
    if issues:
        return [], [], issues, metadata
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except Exception:
        return [], [], [{"code": "invalid_xlsx", "detail": "input is not a validated .xlsx workbook"}], metadata
    try:
        metadata["sheet_count"] = len(workbook.sheetnames)
        if len(workbook.sheetnames) > MAX_WORKBOOK_SHEETS:
            issues.append({"code": "workbook_sheet_limit", "detail": "workbook exceeds 25 sheets"})
        tables: list[dict[str, object]] = []
        risks: list[dict[str, object]] = []
        cell_count = 0
        for sheet_name in workbook.sheetnames[:MAX_WORKBOOK_SHEETS]:
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=False)
            header = next(rows, None)
            if header is None:
                continue
            cell_count += len(header)
            if cell_count > MAX_WORKBOOK_CELLS:
                issues.append({"code": "workbook_cell_limit", "detail": "workbook exceeds 5,000,000 inspected cells"})
                break
            if any(cell.data_type == "f" for cell in header):
                issues.append({"code": "formula_cell", "detail": "formula cells are not accepted"})
                continue
            try:
                fields = [_field_name(cell.value, "XLSX header") for cell in header]
            except ContractError:
                issues.append({"code": "invalid_header", "detail": "XLSX headers must be nonempty strings"})
                continue
            if len(fields) != len(set(fields)):
                issues.append({"code": "duplicate_field", "detail": "XLSX header fields must be unique"})
                continue
            if len(fields) > MAX_TABLE_COLUMNS:
                issues.append({"code": "table_column_limit", "detail": "worksheet exceeds 10,000 columns"})
                continue
            accumulator = _TableAccumulator(display_name, sheet_name, "$", "wide", fields)
            for row in rows:
                cell_count += len(row)
                if cell_count > MAX_WORKBOOK_CELLS:
                    issues.append({"code": "workbook_cell_limit", "detail": "workbook exceeds 5,000,000 inspected cells"})
                    break
                if any(cell.data_type == "f" for cell in row):
                    issues.append({"code": "formula_cell", "detail": "formula cells are not accepted"})
                    break
                accumulator.add_row([cell.value for cell in row])
            tables.append(accumulator.table())
            risks.extend(accumulator.risks())
        return tables, risks, issues, metadata
    finally:
        workbook.close()


def _xlsx_container_issues(path: Path) -> list[dict[str, str]]:
    if not zipfile.is_zipfile(path):
        return [{"code": "invalid_xlsx", "detail": "input is not a validated .xlsx workbook"}]
    try:
        with zipfile.ZipFile(path) as archive:
            if any(info.flag_bits & 0x1 for info in archive.infolist()):
                return [{"code": "encrypted_workbook", "detail": "encrypted workbooks are not accepted"}]
            names = archive.namelist()
            if any(name.casefold().endswith("vbaproject.bin") for name in names):
                return [{"code": "macro_enabled_workbook", "detail": "VBA-bearing workbooks are not accepted"}]
            unsafe_suffixes = (".exe", ".dll", ".com", ".bat", ".cmd", ".ps1", ".js", ".vbs")
            if any(
                name.casefold().startswith(("xl/embeddings/", "xl/activex/"))
                or name.casefold().endswith(unsafe_suffixes)
                for name in names
            ):
                return [{"code": "embedded_executable", "detail": "embedded executable workbook content is not accepted"}]
            content_types = archive.read("[Content_Types].xml").lower() if "[Content_Types].xml" in names else b""
            if any(marker in content_types for marker in (b"macroenabled", b"oleobject", b"activex", b"embeddedpackage")):
                return [{"code": "macro_enabled_workbook", "detail": "unsafe workbook content types are not accepted"}]
            if any("externallinks/" in name.casefold() for name in names):
                return [{"code": "external_links", "detail": "external workbook links are not accepted"}]
            for name in names:
                if name.endswith(".rels") and b'TargetMode="External"' in archive.read(name):
                    return [{"code": "external_links", "detail": "external workbook links are not accepted"}]
    except (OSError, zipfile.BadZipFile):
        return [{"code": "invalid_xlsx", "detail": "input is not a validated .xlsx workbook"}]
    return []


class _TableAccumulator:
    def __init__(self, file_name: str, sheet: str | None, record_path: str, shape: str, fields: list[str]):
        self.file_name, self.sheet, self.record_path, self.shape, self.fields = file_name, sheet, record_path, shape, fields
        self.rows = 0
        self.nulls = {field: 0 for field in fields}
        self.types = {field: set() for field in fields}
        self.classes = {field: set() for field in fields}
        self._risks: list[dict[str, object]] = []
        self._issues: list[dict[str, str]] = []
        for field in fields:
            risk = _risk_for_field(field)
            if risk:
                self._record_risk(risk, field, "field_name")
        if any(_EVENT_FIELD.search(_normalize_field(field)) for field in fields):
            self._record_risk("person_level_event_rows", "", "field_combination")

    def add_row(self, values: Sequence[object]) -> None:
        if self.rows >= MAX_TABLE_ROWS:
            self._record_issue("table_row_limit", "table exceeds 1,000,000 rows")
            return
        self.rows += 1
        for index, field in enumerate(self.fields):
            value = values[index] if index < len(values) else None
            if value is None or (isinstance(value, str) and not value.strip()):
                self.nulls[field] += 1
                continue
            scalar_type, value_class, risk = _classify_value(value)
            self.types[field].add(scalar_type)
            self.classes[field].add(value_class)
            if risk:
                self._record_risk(risk, field, "value_class")

    def _record_risk(self, code: str, field: str, source: str) -> None:
        if not any(item == {"code": code, "field": field, "source": source} for item in self._risks):
            self._risks.append({"code": code, "field": field, "source": source})

    def _record_issue(self, code: str, detail: str) -> None:
        if not any(item["code"] == code for item in self._issues):
            self._issues.append({"code": code, "detail": detail})

    def table(self) -> dict[str, object]:
        denominator = max(self.rows, 1)
        return {
            "file": self.file_name, "sheet": self.sheet, "record_path": self.record_path,
            "shape": self.shape, "row_count": self.rows, "column_count": len(self.fields), "field_names": self.fields,
            "observed_scalar_types": {field: sorted(self.types[field]) for field in self.fields},
            "null_rates": {field: self.nulls[field] / denominator for field in self.fields},
            "sample_safe_value_classes": {field: sorted(self.classes[field]) for field in self.fields},
            "candidate_units": _candidate_units(self.fields), "candidate_denominators": _candidate_denominators(self.fields),
            "candidate_field_roles": {field: _candidate_roles(field) for field in self.fields},
        }

    def risks(self) -> list[dict[str, object]]:
        return self._risks

    def issues(self) -> list[dict[str, str]]:
        return self._issues


class _DynamicTableAccumulator(_TableAccumulator):
    """A table accumulator that learns JSON object fields one record at a time."""

    def __init__(self, file_name: str, sheet: str | None, record_path: str, shape: str):
        super().__init__(file_name, sheet, record_path, shape, [])

    def add_mapping(self, values: dict[str, object]) -> None:
        for field in values:
            self._add_field(field)
        self.add_row([values.get(field) for field in self.fields])

    def _add_field(self, field: str) -> None:
        if field in self.nulls:
            return
        if len(self.fields) >= MAX_TABLE_COLUMNS:
            self._issues.append({"code": "table_column_limit", "detail": "table exceeds 10,000 columns"})
            return
        self.fields.append(field)
        self.nulls[field] = self.rows
        self.types[field] = set()
        self.classes[field] = set()
        risk = _risk_for_field(field)
        if risk:
            self._record_risk(risk, field, "field_name")

    def risks(self) -> list[dict[str, object]]:
        if any(_EVENT_FIELD.search(_normalize_field(field)) for field in self.fields):
            self._record_risk("person_level_event_rows", "", "field_combination")
        return self._risks


def _field_name(value: object, label: str) -> str:
    return require_string(str(value) if value is not None else "", label).strip()


def _normalize_field(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.casefold()).strip("_")


def _risk_for_field(field: str) -> str | None:
    normalized = _normalize_field(field)
    for pattern, risk in _PERSON_FIELD_RISKS:
        if pattern.search(normalized):
            return risk
    return None


def _classify_value(value: object) -> tuple[str, str, str | None]:
    if isinstance(value, bool):
        return "boolean", "boolean", None
    if isinstance(value, int):
        return "integer", "integer", None
    if isinstance(value, float):
        return "number", "number", None
    text = str(value).strip()
    if re.fullmatch(r"[-+]?\d+", text):
        return "integer", "integer", None
    if re.fullmatch(r"[-+]?(?:\d+\.\d*|\d*\.\d+)", text):
        return "number", "number", None
    if _EMAIL_VALUE.fullmatch(text):
        return "string", "email_like", "email"
    if _is_supported_calendar_date(text):
        return "string", "date_or_datetime_like", None
    if _PHONE_VALUE.fullmatch(text):
        return "string", "phone_like", "phone"
    if _IP_VALUE.fullmatch(text):
        return "string", "ip_address_like", "ip_address"
    if _IDENTIFIER_VALUE.fullmatch(text):
        return "string", "identifier_like", None
    return "string", "free_text_like", None


def _is_supported_calendar_date(text: str) -> bool:
    """Recognize only complete, unambiguous calendar dates and datetimes."""

    match = _DATE_PART.fullmatch(text)
    if match:
        first, separator, second, third = match.groups()
        try:
            if len(first) == 4:
                date(int(first), int(second), int(third))
            elif separator == "/":
                date(int(third), int(first), int(second))
            elif int(first) > 12:
                date(int(third), int(second), int(first))
            else:
                return False
        except ValueError:
            return False
        return True

    if not _ISO_DATETIME.fullmatch(text):
        return False
    normalized = f"{text[:-1]}+00:00" if text.endswith("Z") else text
    try:
        datetime.fromisoformat(normalized)
    except ValueError:
        return False
    return True


def _candidate_units(fields: list[str]) -> list[str]:
    values = []
    if any("respondent" in _normalize_field(field) for field in fields):
        values.append("respondent")
    if any("account" in _normalize_field(field) for field in fields):
        values.append("account")
    return values


def _candidate_denominators(fields: list[str]) -> list[str]:
    return [field for field in fields if _normalize_field(field) in {"total", "sample_size", "respondent_count", "count", "n"}]


def _candidate_roles(field: str) -> list[str]:
    normalized = _normalize_field(field)
    if normalized in {"count", "respondent_count", "sample_size", "n"}:
        return ["candidate_count"]
    if normalized in {"share", "rate", "percentage", "pct"}:
        return ["candidate_share_or_rate"]
    if normalized.endswith("_id"):
        return ["candidate_key"]
    return ["unresolved"]


def _candidate_semantics(tables: list[dict[str, object]]) -> list[dict[str, object]]:
    items = []
    for table in tables:
        for field, roles in table["candidate_field_roles"].items():
            if roles != ["unresolved"]:
                items.append({"file": table["file"], "field": field, "candidate_roles": roles, "requires_confirmation": True})
    return sorted(items, key=lambda item: (str(item["file"]), str(item["field"])))


def _relationships(tables: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[str, set[str]] = {}
    for table in tables:
        for field in table["field_names"]:
            if _candidate_roles(field) == ["candidate_key"]:
                grouped.setdefault(field, set()).add(str(table["file"]))
    return [{"kind": "candidate_shared_field", "field": field, "files": sorted(files)} for field, files in sorted(grouped.items()) if len(files) > 1]


def _unique_records(records: list[dict[str, object]] | list[dict[str, str]]) -> list[dict[str, object]]:
    unique: dict[str, dict[str, object]] = {}
    for record in records:
        key = json.dumps(record, sort_keys=True, separators=(",", ":"))
        unique[key] = dict(record)
    return [unique[key] for key in sorted(unique)]
