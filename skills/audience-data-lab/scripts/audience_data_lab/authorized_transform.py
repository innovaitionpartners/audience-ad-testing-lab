"""Deterministic transformation of approved aggregate audience sources."""

from __future__ import annotations

from collections import defaultdict
import csv
import ctypes
from datetime import date, datetime
import errno
import io
import json
import os
from pathlib import Path
import re
import shutil
import sys
import tempfile
from typing import Callable, Mapping

from openpyxl import load_workbook

_SKILLS_ROOT = Path(__file__).resolve().parents[3]
for _dependency_scripts in (
    _SKILLS_ROOT / "audience-ad-testing-lab" / "scripts",
    _SKILLS_ROOT / "audience-panel-builder" / "scripts",
):
    if str(_dependency_scripts) not in sys.path:
        sys.path.insert(0, str(_dependency_scripts))

from audience_lab.audience_research_v3 import (  # noqa: E402
    AudienceResearchV3ValidationError,
    validate_observation_batch,
    validate_outcome_feedback,
)
from audience_panel_builder.common import (  # noqa: E402
    ContractError as PanelBuilderContractError,
)
from audience_panel_builder.evidence import build_evidence_ledger  # noqa: E402

from .authorized_mapping import (
    ALLOWED_OPERATIONS,
    CANONICAL_OUTPUT_REGISTRY,
    mapping_sha256,
    validate_authorized_mapping,
)
from .authorized_source import validate_source_profile
from .common import (
    ContractError,
    canonical_json_bytes,
    exact_object,
    require_boolean,
    require_enum,
    require_identifier,
    require_integer,
    require_string,
    require_string_list,
    sha256_bytes,
    sha256_file,
    write_new_bytes,
)


TRANSFORMATION_REPORT_VERSION = "authorized-audience-transformation-report-v1"
AUTHORIZED_HANDOFF_VERSION = "authorized-audience-handoff-v1"

_REPORT_KEYS = {
    "schema_version",
    "source_profile",
    "mapping",
    "transformer_version",
    "input_hashes",
    "source_reads",
    "operation_log",
    "field_changes",
    "value_changes",
    "route_summary",
    "loss_summary",
    "warnings",
    "blocking_errors",
    "outputs",
    "status",
}
_SOURCE_READ_KEYS = {
    "selection_id", "file", "file_sha256", "sheet", "record_path", "rows_read",
    "fields_read", "unit", "denominator",
}
_OPERATION_LOG_KEYS = {
    "operation_id", "op", "input_rows", "output_rows", "filtered_rows",
    "loss_consequence", "loss_consequences", "details",
}
_FIELD_CHANGE_KEYS = {"renamed", "cast", "combined", "ignored", "dropped"}
_VALUE_CHANGE_KEYS = {
    "category_merges", "unmapped_values", "missing_values",
    "suppressed_values", "rejected_values", "filtered_values",
}
_ROUTE_SUMMARY_KEYS = {
    "path", "route", "schema_version", "row_count", "field_count",
}
_LOSS_SUMMARY_KEYS = {"consequences", "coverage"}
_OUTPUT_KEYS = {
    "path", "sha256", "route", "schema_version", "row_count", "unit",
    "denominator", "field_count",
}
_HANDOFF_KEYS = {
    "schema_version", "status", "source_profile", "mapping",
    "transformation_report", "outputs", "profile_seeds", "privacy_permission",
    "cohort_identity",
}
_COHORT_IDENTITY_KEYS = {
    "cohort_id", "source_profile_sha256", "source_bundle_sha256",
    "structural_outputs",
}
_COHORT_STRUCTURAL_OUTPUT_KEYS = {
    "path", "sha256", "schema_version", "batch_id", "unit",
    "denominator", "row_count",
}
_REPORT_REFERENCE_KEYS = {"path", "sha256"}
_PRIVACY_PERMISSION_KEYS = {
    "permission_confirmed", "aggregate_only", "minimum_cell_size",
}
_DIGEST = re.compile(r"sha256:[0-9a-f]{64}")
_SAFE_OUTPUT = re.compile(
    r"(?:frame-observations|structured-evidence|social-observations|"
    r"profile-seeds|outcome-feedback)-\d{4}\.json"
)
_CANONICAL_DOCUMENT_KEYS = {
    "schema_version", "batch_id", "route", "unit", "denominator", "records",
}
_STRUCTURAL_VALUE_FIELDS = {
    "count", "respondent_count", "sample_size", "n", "population",
}
_PROHIBITED_OUTPUT_FIELDS = {
    "email", "phone", "name", "address", "device_id", "cookie_id",
    "advertising_id", "ip_address", "account_id",
}
_DETAIL_KEYS = {
    "select": {"selected_fields", "dropped_fields"},
    "rename": {"renamed_fields"},
    "cast": {"cast_fields"},
    "flatten": {"flattened_paths"},
    "wide_to_long": {
        "reshape", "id_fields", "value_fields", "name_field", "value_field",
    },
    "pivot": {
        "reshape", "index_fields", "column_field", "value_field", "columns",
    },
    "join": {"join", "unused_right_rows"},
    "category_map": {
        "field", "mapping", "category_merges", "unmapped_values",
        "unmapped_policy",
    },
    "normalize_missing": {"missing_values_normalized", "fields", "values"},
    "normalize_suppression": {
        "suppressed_values_normalized", "field", "values", "status_field",
    },
    "derive_share": {"derived_field", "count_field", "denominator_field"},
    "normalize_weight": {
        "source_field", "normalized_weight_field", "group_by",
    },
    "aggregate": {"aggregate_group_by", "aggregate_metrics"},
    "filter": {"predicate", "field", "value", "filtered_values"},
    "sort": {"sort_fields"},
}

Rows = list[dict[str, object]]
Audit = dict[str, object]


def _stable_key(value: object) -> str:
    return canonical_json_bytes(value).decode("utf-8")


def _copy_rows(rows: Rows) -> Rows:
    return [dict(row) for row in rows]


def _operation_audit(operation: Mapping[str, object], before: int, after: int) -> Audit:
    filtered = before - after if operation["op"] == "filter" else 0
    losses = ["row_filter"] if filtered else []
    return {
        "operation_id": operation["operation_id"],
        "op": operation["op"],
        "input_rows": before,
        "output_rows": after,
        "filtered_rows": filtered,
        "loss_consequence": losses[0] if losses else None,
        "loss_consequences": losses,
        "details": {},
    }


def _select(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    fields = list(operation["fields"])
    result = [{field: row[field] for field in fields} for row in rows]
    input_fields = set().union(*(row.keys() for row in rows)) if rows else set()
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "selected_fields": fields,
        "dropped_fields": sorted(input_fields - set(fields)),
    }
    if audit["details"]["dropped_fields"]:
        audit["loss_consequences"].append("field_drop")
        audit["loss_consequence"] = audit["loss_consequences"][0]
    return result, audit


def _rename(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    names = dict(operation["fields"])
    result = [
        {names.get(field, field): value for field, value in row.items()}
        for row in rows
    ]
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {"renamed_fields": dict(sorted(names.items()))}
    return result, audit


def _cast_value(value: object, cast_type: str, path: str) -> object:
    if value is None:
        return None
    try:
        if cast_type == "string":
            return str(value)
        if cast_type == "integer":
            if isinstance(value, bool):
                raise ValueError
            number = float(value)
            if not number.is_integer():
                raise ValueError
            return int(number)
        if cast_type == "number":
            if isinstance(value, bool):
                raise ValueError
            return float(value)
        if cast_type == "boolean":
            if isinstance(value, bool):
                return value
            normalized = str(value).strip().casefold()
            if normalized in {"true", "1", "yes"}:
                return True
            if normalized in {"false", "0", "no"}:
                return False
            raise ValueError
        if cast_type == "date":
            if isinstance(value, datetime):
                return value.date().isoformat()
            if isinstance(value, date):
                return value.isoformat()
            return date.fromisoformat(str(value).strip()).isoformat()
    except (TypeError, ValueError, OverflowError) as exc:
        raise ContractError(f"{path} cannot be cast to {cast_type}") from exc
    raise ContractError(f"{path} uses an unsupported cast")


def _cast(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    casts = dict(operation["fields"])
    result: Rows = []
    for row_index, row in enumerate(rows):
        item = dict(row)
        for field, cast_type in casts.items():
            item[field] = _cast_value(
                item[field], str(cast_type),
                f"{operation['operation_id']} row {row_index} field {field}",
            )
        result.append(item)
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {"cast_fields": dict(sorted(casts.items()))}
    return result, audit


def _get_nested(row: Mapping[str, object], dotted_path: str) -> object:
    current: object = row
    for part in dotted_path.split("."):
        if not isinstance(current, Mapping) or part not in current:
            raise ContractError(f"flatten path does not resolve exactly: {dotted_path}")
        current = current[part]
    return current


def _flatten(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    fields = dict(operation["fields"])
    result = [
        {output: _get_nested(row, str(source)) for output, source in fields.items()}
        for row in rows
    ]
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {"flattened_paths": dict(sorted(fields.items()))}
    return result, audit


def _wide_to_long(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    ids = list(operation["id_fields"])
    values = list(operation["value_fields"])
    name_field = str(operation["name_field"])
    value_field = str(operation["value_field"])
    result = [
        {
            **{field: row[field] for field in ids},
            name_field: source_field,
            value_field: row[source_field],
        }
        for row in rows
        for source_field in values
    ]
    result.sort(
        key=lambda row: (
            tuple(_stable_key(row[field]) for field in ids),
            values.index(str(row[name_field])),
        )
    )
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "reshape": "wide_to_long",
        "id_fields": ids,
        "value_fields": values,
        "name_field": name_field,
        "value_field": value_field,
    }
    return result, audit


def _pivot(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    indexes = list(operation["index_fields"])
    column_field = str(operation["column_field"])
    value_field = str(operation["value_field"])
    columns = list(operation["columns"])
    allowed = set(columns)
    grouped: dict[tuple[object, ...], dict[str, object]] = {}
    for row in rows:
        key = tuple(row[field] for field in indexes)
        column = str(row[column_field])
        if column not in allowed:
            raise ContractError(f"pivot encountered undeclared category: {column}")
        item = grouped.setdefault(key, {field: row[field] for field in indexes})
        if column in item:
            raise ContractError(f"pivot encountered duplicate key/category: {key!r}/{column}")
        item[column] = row[value_field]
    for key, row in grouped.items():
        missing = [column for column in columns if column not in row]
        if missing:
            raise ContractError(f"pivot key {key!r} is missing declared categories: {', '.join(missing)}")
    result = sorted(grouped.values(), key=_stable_key)
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "reshape": "pivot",
        "index_fields": indexes,
        "column_field": column_field,
        "value_field": value_field,
        "columns": columns,
    }
    return result, audit


def _join_key(row: Mapping[str, object], fields: list[str]) -> tuple[object, ...]:
    return tuple(row[field] for field in fields)


def _join(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    left = tables[str(operation["left"])]
    right = tables[str(operation["right"])]
    on = list(operation["on"])
    right_index: dict[tuple[object, ...], dict[str, object]] = {}
    for row in right:
        key = _join_key(row, on)
        if key in right_index:
            raise ContractError(f"duplicate join key in right dataset: {key!r}")
        right_index[key] = row
    if operation["cardinality"] == "one_to_one":
        seen: set[tuple[object, ...]] = set()
        for row in left:
            key = _join_key(row, on)
            if key in seen:
                raise ContractError(f"duplicate join key in left dataset: {key!r}")
            seen.add(key)
    result: Rows = []
    matched_right_keys: set[tuple[object, ...]] = set()
    for row in left:
        key = _join_key(row, on)
        if key not in right_index:
            raise ContractError(f"join key has no exact match: {key!r}")
        matched = right_index[key]
        matched_right_keys.add(key)
        result.append({**row, **{field: value for field, value in matched.items() if field not in on}})
    result.sort(key=_stable_key)
    audit = _operation_audit(operation, len(left) + len(right), len(result))
    audit["details"] = {
        "join": {
            "left": operation["left"],
            "right": operation["right"],
            "on": on,
            "cardinality": operation["cardinality"],
            "matched_rows": len(result),
        },
        "unused_right_rows": len(set(right_index) - matched_right_keys),
    }
    if audit["details"]["unused_right_rows"]:
        audit["loss_consequences"].append("unmatched_join_rows")
        audit["loss_consequence"] = audit["loss_consequences"][0]
    return result, audit


def _category_map(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    field = str(operation["field"])
    categories = dict(operation["mapping"])
    policy = str(operation["unmapped"])
    result: Rows = []
    unmapped = 0
    merges: dict[str, int] = defaultdict(int)
    for row in rows:
        item = dict(row)
        source = item[field]
        if source in categories:
            target = categories[source]
            item[field] = target
            merges[f"{source}->{target}"] += 1
        else:
            unmapped += 1
            if policy == "error":
                raise ContractError(f"category_map encountered undeclared category: {source!r}")
            if policy == "null":
                item[field] = None
        result.append(item)
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "field": field,
        "mapping": dict(sorted(categories.items())),
        "category_merges": dict(sorted(merges.items())),
        "unmapped_values": unmapped,
        "unmapped_policy": policy,
    }
    observed_targets = {
        categories[source]
        for row in rows
        for source in (row[field],)
        if source in categories
    }
    if any(
        sum(1 for source in categories if categories[source] == target) > 1
        for target in observed_targets
    ):
        audit["loss_consequences"].append("category_merge")
        audit["loss_consequence"] = audit["loss_consequences"][0]
    return result, audit


def _normalize_missing(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    fields = list(operation["fields"])
    values = list(operation["values"])
    result: Rows = []
    changed = 0
    for row in rows:
        item = dict(row)
        for field in fields:
            if item[field] in values:
                item[field] = None
                changed += 1
        result.append(item)
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "missing_values_normalized": changed,
        "fields": fields,
        "values": values,
    }
    return result, audit


def _normalize_suppression(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    field = str(operation["field"])
    values = list(operation["values"])
    status_field = str(operation["status_field"])
    result: Rows = []
    changed = 0
    for row in rows:
        item = dict(row)
        suppressed = item[field] in values
        item[status_field] = suppressed
        if suppressed:
            item[field] = None
            changed += 1
        result.append(item)
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "suppressed_values_normalized": changed,
        "field": field,
        "values": values,
        "status_field": status_field,
    }
    return result, audit


def _derive_share(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    count_field = str(operation["count_field"])
    denominator_field = str(operation["denominator_field"])
    output_field = str(operation["output_field"])
    result: Rows = []
    for row in rows:
        item = dict(row)
        count = item[count_field]
        denominator = item[denominator_field]
        if count is None or denominator is None:
            item[output_field] = None
        elif isinstance(count, bool) or isinstance(denominator, bool) or not isinstance(count, (int, float)) or not isinstance(denominator, (int, float)):
            raise ContractError("derive_share requires numeric count and denominator")
        elif denominator <= 0:
            raise ContractError("derive_share denominator must be positive")
        else:
            item[output_field] = float(count) / float(denominator)
        result.append(item)
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "derived_field": output_field,
        "count_field": count_field,
        "denominator_field": denominator_field,
    }
    return result, audit


def _normalize_weight(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    field = str(operation["field"])
    output_field = str(operation["output_field"])
    groups = list(operation["group_by"])
    totals: dict[tuple[object, ...], float] = defaultdict(float)
    for row in rows:
        value = row[field]
        if isinstance(value, bool) or not isinstance(value, (int, float)) or value < 0:
            raise ContractError("normalize_weight requires nonnegative numeric values")
        totals[_join_key(row, groups)] += float(value)
    if any(total <= 0 for total in totals.values()):
        raise ContractError("normalize_weight group totals must be positive")
    result = [
        {
            **row,
            output_field: float(row[field]) / totals[_join_key(row, groups)],
        }
        for row in rows
    ]
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "source_field": field,
        "normalized_weight_field": output_field,
        "group_by": groups,
    }
    return result, audit


def _aggregate(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    groups = list(operation["group_by"])
    metrics = dict(operation["metrics"])
    grouped: dict[tuple[object, ...], list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        grouped[_join_key(row, groups)].append(row)
    result: Rows = []
    for key, members in grouped.items():
        item: dict[str, object] = dict(zip(groups, key))
        for output_field, raw_metric in metrics.items():
            metric = dict(raw_metric)
            values = [member[str(metric["field"])] for member in members]
            if metric["function"] == "count":
                item[output_field] = sum(value is not None for value in values)
                continue
            numeric = [
                float(value) for value in values
                if value is not None and not isinstance(value, bool) and isinstance(value, (int, float))
            ]
            if len(numeric) != sum(value is not None for value in values):
                raise ContractError("aggregate numeric functions require numeric values")
            function = metric["function"]
            if function == "sum":
                value: object = sum(numeric)
                if all(isinstance(candidate, int) and not isinstance(candidate, bool) for candidate in values):
                    value = int(value)
            elif function == "min":
                value = min(numeric) if numeric else None
            elif function == "max":
                value = max(numeric) if numeric else None
            elif function == "mean":
                value = sum(numeric) / len(numeric) if numeric else None
            else:
                raise ContractError("aggregate function is unsupported")
            item[output_field] = value
        result.append(item)
    result.sort(key=_stable_key)
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "aggregate_group_by": groups,
        "aggregate_metrics": dict(sorted(metrics.items())),
    }
    if len(result) < len(rows):
        audit["loss_consequences"].append("aggregation_granularity")
        audit["loss_consequence"] = audit["loss_consequences"][0]
    return result, audit


def _filter(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = tables[str(operation["input"])]
    field = str(operation["field"])
    predicate = str(operation["predicate"])
    wanted = operation["value"]

    def keep(value: object) -> bool:
        if predicate == "equals":
            return value == wanted
        if predicate == "not_equals":
            return value != wanted
        if predicate == "in":
            return value in wanted
        if predicate == "not_in":
            return value not in wanted
        if predicate == "is_null":
            return value is None
        if predicate == "is_not_null":
            return value is not None
        raise ContractError("filter predicate is unsupported")

    result = [dict(row) for row in rows if keep(row[field])]
    audit = _operation_audit(operation, len(rows), len(result))
    audit["details"] = {
        "predicate": predicate,
        "field": field,
        "value": wanted,
        "filtered_values": len(rows) - len(result),
    }
    return result, audit


def _sort(tables: dict[str, Rows], operation: Mapping[str, object]) -> tuple[Rows, Audit]:
    rows = _copy_rows(tables[str(operation["input"])])
    fields = list(operation["fields"])
    rows.sort(key=lambda row: tuple(_stable_key(row[field]) for field in fields))
    audit = _operation_audit(operation, len(rows), len(rows))
    audit["details"] = {"sort_fields": fields}
    return rows, audit


OPERATION_HANDLERS: dict[str, Callable[[dict[str, Rows], Mapping[str, object]], tuple[Rows, Audit]]] = {
    "select": _select,
    "rename": _rename,
    "cast": _cast,
    "flatten": _flatten,
    "wide_to_long": _wide_to_long,
    "pivot": _pivot,
    "join": _join,
    "category_map": _category_map,
    "normalize_missing": _normalize_missing,
    "normalize_suppression": _normalize_suppression,
    "derive_share": _derive_share,
    "normalize_weight": _normalize_weight,
    "aggregate": _aggregate,
    "filter": _filter,
    "sort": _sort,
}


def apply_authorized_operations(
    tables: Mapping[str, Rows],
    operations: list[dict[str, object]],
) -> tuple[dict[str, Rows], list[Audit]]:
    """Apply only named, statically-dispatched declarative operations."""

    datasets = {name: _copy_rows(rows) for name, rows in tables.items()}
    audit: list[Audit] = []
    for operation in operations:
        op = operation.get("op")
        if op not in OPERATION_HANDLERS:
            raise ContractError(f"unsupported authorized operation: {op}")
        output = str(operation["output"])
        if output in datasets:
            raise ContractError(f"operation output already exists: {output}")
        rows, record = OPERATION_HANDLERS[str(op)](datasets, operation)
        datasets[output] = rows
        audit.append(record)
    return datasets, audit


def _safe_input_path(input_root: Path, display_name: str) -> Path:
    root = input_root.resolve()
    candidate = (root / display_name).resolve()
    if candidate.parent != root or candidate.name != display_name:
        raise ContractError(f"input file must be a direct child of input_root: {display_name}")
    if not candidate.is_file():
        raise ContractError(f"authorized input is missing or not a regular file: {display_name}")
    return candidate


def _snapshot_input(path: Path, expected_digest: object) -> tuple[bytes, str]:
    try:
        with path.open("rb") as handle:
            before = os.fstat(handle.fileno())
            data = handle.read()
            after = os.fstat(handle.fileno())
    except OSError as exc:
        raise ContractError(f"authorized input cannot be read: {path.name}") from exc
    identity_before = (
        before.st_dev, before.st_ino, before.st_size, before.st_mtime_ns
    )
    identity_after = (
        after.st_dev, after.st_ino, after.st_size, after.st_mtime_ns
    )
    if identity_before != identity_after or len(data) != before.st_size:
        raise ContractError(f"authorized input changed while being consumed: {path.name}")
    digest = sha256_bytes(data)
    if digest != expected_digest:
        raise ContractError(f"authorized input hash changed after profiling: {path.name}")
    return data, digest


def _read_csv(data: bytes) -> tuple[list[str], Rows]:
    try:
        with io.TextIOWrapper(
            io.BytesIO(data), encoding="utf-8-sig", newline=""
        ) as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames or any(field is None or not field.strip() for field in reader.fieldnames):
                raise ContractError("CSV requires nonempty headers")
            if len(reader.fieldnames) != len(set(reader.fieldnames)):
                raise ContractError("CSV headers must be unique")
            rows: Rows = []
            for row in reader:
                if None in row:
                    raise ContractError("CSV row width does not match its header")
                rows.append(dict(row))
            return list(reader.fieldnames), rows
    except UnicodeDecodeError as exc:
        raise ContractError("CSV must use UTF-8 or UTF-8-SIG") from exc


def _resolve_record_path(payload: object, record_path: str) -> object:
    if record_path == "$":
        return payload
    current = payload
    for part in record_path.split("."):
        if not isinstance(current, Mapping) or part not in current:
            raise ContractError(f"JSON record path does not resolve exactly: {record_path}")
        current = current[part]
    return current


def _read_json(data: bytes, record_path: str) -> tuple[list[str], Rows]:
    try:
        payload = json.loads(data.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ContractError("JSON input cannot be read as one UTF-8 document") from exc
    records = _resolve_record_path(payload, record_path)
    if isinstance(records, Mapping):
        records = [records]
    if not isinstance(records, list) or not all(isinstance(row, Mapping) for row in records):
        raise ContractError("JSON record path must resolve to an object or array of objects")
    fields: list[str] = []
    for row in records:
        for field in row:
            if not isinstance(field, str) or not field:
                raise ContractError("JSON record fields must be nonempty strings")
            if field not in fields:
                fields.append(field)
    return fields, [dict(row) for row in records]


def _read_xlsx(data: bytes, sheet_name: str | None) -> tuple[list[str], Rows]:
    if sheet_name is None:
        raise ContractError("XLSX selection must name an approved sheet")
    try:
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise ContractError("XLSX input cannot be opened safely") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            raise ContractError(f"XLSX sheet does not resolve exactly: {sheet_name}")
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=False)
        header = next(iterator, None)
        if not header:
            raise ContractError("XLSX sheet requires a header")
        if any(cell.data_type == "f" for cell in header):
            raise ContractError("XLSX formulas are not accepted")
        fields = [require_string(cell.value, "XLSX header") for cell in header]
        if len(fields) != len(set(fields)):
            raise ContractError("XLSX headers must be unique")
        rows: Rows = []
        for raw in iterator:
            if any(cell.data_type == "f" for cell in raw):
                raise ContractError("XLSX formulas are not accepted")
            values = [cell.value for cell in raw]
            if len(values) != len(fields):
                raise ContractError("XLSX row width does not match its header")
            rows.append(dict(zip(fields, values)))
        return fields, rows
    finally:
        workbook.close()


def _read_selection(
    selection: Mapping[str, object],
    input_root: Path,
) -> tuple[Rows, dict[str, object]]:
    display_name = str(selection["file"])
    path = _safe_input_path(input_root, display_name)
    data, digest = _snapshot_input(path, selection["file_sha256"])
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        fields, rows = _read_csv(data)
    elif suffix == ".json":
        fields, rows = _read_json(data, str(selection["record_path"]))
    elif suffix == ".xlsx":
        fields, rows = _read_xlsx(data, selection["sheet"])
    else:
        raise ContractError(f"authorized transformation does not support {suffix}")
    selected = list(selection["fields"])
    if not set(selected) <= set(fields):
        raise ContractError(f"authorized source fields changed after profiling: {display_name}")
    normalized = [{field: row.get(field) for field in selected} for row in rows]
    return normalized, {
        "selection_id": selection["selection_id"],
        "file": display_name,
        "file_sha256": digest,
        "sheet": selection["sheet"],
        "record_path": selection["record_path"],
        "rows_read": len(rows),
        "fields_read": selected,
        "unit": selection["unit"],
        "denominator": selection["denominator"],
    }


def _dataset_ancestors(mapping: Mapping[str, object]) -> dict[str, set[str]]:
    ancestors = {
        str(selection["selection_id"]): {str(selection["selection_id"])}
        for selection in mapping["selections"]
    }
    for operation in mapping["operations"]:
        output = str(operation["output"])
        if operation["op"] == "join":
            ancestors[output] = ancestors[str(operation["left"])] | ancestors[str(operation["right"])]
        else:
            ancestors[output] = set(ancestors[str(operation["input"])])
    return ancestors


def _canonical_document(
    *,
    mapping: Mapping[str, object],
    output: Mapping[str, object],
    rows: Rows,
    units: dict[str, tuple[str, str]],
    ancestors: dict[str, set[str]],
) -> tuple[dict[str, object], str, str]:
    dataset = str(output["dataset"])
    source_units = {units[item] for item in ancestors[dataset]}
    if len(source_units) != 1:
        raise ContractError(f"output {dataset} combines incompatible units or denominators")
    source_unit, source_denominator = next(iter(source_units))
    metadata = dict(output["metadata"])
    filename = str(output["filename"])
    family = filename.removesuffix(".json").rsplit("-", 1)[0]
    records = sorted(_copy_rows(rows), key=_stable_key)
    for index, record in enumerate(records):
        if set(record) & _PROHIBITED_OUTPUT_FIELDS:
            raise ContractError(
                f"canonical source row {index} contains a prohibited person-level field"
            )

    if family == "frame-observations":
        dimension_fields = dict(metadata["dimension_fields"])
        cell_key_field = str(metadata["cell_key_field"])
        estimate_field = str(metadata["estimate_field"])
        cell_metadata = dict(metadata["cell_metadata"])
        observed_keys = {str(record[cell_key_field]) for record in records}
        if observed_keys != set(cell_metadata):
            raise ContractError(
                "frame output cell_metadata must exactly cover transformed rows"
            )
        cells = []
        for record in records:
            source_key = str(record[cell_key_field])
            cell = dict(cell_metadata[source_key])
            uncertainty = dict(cell["uncertainty"])
            cells.append(
                {
                    "cell_id": cell["cell_id"],
                    "dimension_values": {
                        dimension: str(record[field])
                        for dimension, field in dimension_fields.items()
                    },
                    "estimate": (
                        None if cell["suppressed"] else record[estimate_field]
                    ),
                    "uncertainty": {
                        "lower": (
                            None
                            if cell["suppressed"]
                            else record[str(uncertainty["lower_field"])]
                        ),
                        "upper": (
                            None
                            if cell["suppressed"]
                            else record[str(uncertainty["upper_field"])]
                        ),
                        "method": uncertainty["method"],
                    },
                    "suppressed": cell["suppressed"],
                    "status": cell["status"],
                    "relationship": cell["relationship"],
                    "source_location": cell["source_location"],
                }
            )
        document = {
            "schema_version": output["schema_version"],
            "batch_id": metadata["batch_id"],
            "frame_request_id": metadata["frame_request_id"],
            "adapter_id": metadata["adapter_id"],
            "source_family": metadata["source_family"],
            "source": metadata["source"],
            "raw_snapshot_sha256": metadata["raw_snapshot_sha256"],
            "normalized_batch_sha256": "",
            "access": metadata["access"],
            "geography": metadata["geography"],
            "unit": metadata["unit"],
            "denominator": metadata["denominator"],
            "dimensions": list(dimension_fields),
            "cells": cells,
            "selection_notes": metadata["selection_notes"],
            "coverage_notes": metadata["coverage_notes"],
            "citations": metadata["citations"],
        }
        hash_input = dict(document)
        hash_input.pop("normalized_batch_sha256")
        document["normalized_batch_sha256"] = sha256_bytes(
            canonical_json_bytes(hash_input)
        )
    elif family == "structured-evidence":
        item_id_field = str(metadata["item_id_field"])
        summary_field = str(metadata["content_summary_field"])
        item_metadata = dict(metadata["item_metadata"])
        observed_ids = {str(record[item_id_field]) for record in records}
        if observed_ids != set(item_metadata):
            raise ContractError(
                "structured output item_metadata must exactly cover transformed rows"
            )
        items = []
        for record in records:
            item_id = str(record[item_id_field])
            item = dict(item_metadata[item_id])
            items.append(
                {
                    "evidence_item_id": item_id,
                    "source_url": item["source_url"],
                    "item_type": item["item_type"],
                    "content_summary": record[summary_field],
                    "text_fidelity": item["text_fidelity"],
                    "content_sha256": item["content_sha256"],
                    "source_pointer": item["source_pointer"],
                    "upstream_source_ids": item["upstream_source_ids"],
                    "use_constraints": item["use_constraints"],
                    "quality_flags": item["quality_flags"],
                }
            )
        document = {
            "schema_version": output["schema_version"],
            "batch_id": metadata["batch_id"],
            "created_at": metadata["created_at"],
            "source_adapter": metadata["source_adapter"],
            "source_schema_version": metadata["source_schema_version"],
            "input_sha256": metadata["input_sha256"],
            "permission": metadata["permission"],
            "source_status": metadata["source_status"],
            "items": items,
        }
    elif family == "social-observations":
        observation_id_field = str(metadata["observation_id_field"])
        text_excerpt_field = str(metadata["text_excerpt_field"])
        observation_metadata = dict(metadata["observation_metadata"])
        observed_ids = {str(record[observation_id_field]) for record in records}
        if observed_ids != set(observation_metadata):
            raise ContractError(
                "social output observation_metadata must exactly cover "
                "transformed rows"
            )
        if metadata["collection"]["item_limit"] != len(records):
            raise ContractError(
                "social output collection.item_limit must equal transformed rows"
            )
        observations = []
        for record in records:
            observation_id = str(record[observation_id_field])
            observation = dict(observation_metadata[observation_id])
            observations.append(
                {
                    "observation_id": observation_id,
                    **observation,
                    "text_excerpt": record[text_excerpt_field],
                }
            )
        document = {
            "schema_version": output["schema_version"],
            **{
                key: value
                for key, value in metadata.items()
                if key
                not in {
                    "observation_id_field",
                    "text_excerpt_field",
                    "observation_metadata",
                }
            },
            "observations": observations,
        }
    elif family == "outcome-feedback":
        record_match = dict(metadata["record_match"])
        selected = [
            record
            for record in records
            if all(record.get(field) == value for field, value in record_match.items())
        ]
        if len(selected) != 1:
            raise ContractError(
                "outcome output record_match must resolve exactly one transformed row"
            )
        record = selected[0]
        aggregate_fields = dict(metadata["aggregate_fields"])
        aggregate = {
            key: (
                None if field is None else record[str(field)]
            )
            for key, field in aggregate_fields.items()
        }
        document = {
            "schema_version": output["schema_version"],
            "feedback_id": metadata["feedback_id"],
            "panel_id": metadata["panel_id"],
            "study_id": metadata["study_id"],
            "variant_id": metadata["variant_id"],
            "cohort_id": metadata["cohort_id"],
            "metric": metadata["metric"],
            "metric_direction": metadata["metric_direction"],
            "units": metadata["units"],
            "windows": metadata["windows"],
            "aggregate": aggregate,
            "design": metadata["design"],
            "source": metadata["source"],
            "holdout": metadata["holdout"],
            "missingness": metadata["missingness"],
            "limitations": metadata["limitations"],
            "source_sha256": metadata["source_sha256"],
        }
    else:
        route = str(output["route"])
        sequence = Path(filename).stem.rsplit("-", 1)[-1]
        document = {
            "schema_version": output["schema_version"],
            "batch_id": f"{mapping['mapping_id']}-{route.replace('_', '-')}-{sequence}",
            "route": route,
            "unit": source_unit,
            "denominator": source_denominator,
            "records": records,
        }
    return document, source_unit, source_denominator


def _validate_canonical_document(
    payload: object,
    *,
    expected_output: Mapping[str, object],
    minimum_cell_size: int,
) -> dict[str, object]:
    filename = str(expected_output["filename"])
    family = filename.removesuffix(".json").rsplit("-", 1)[0]
    expected_route, expected_schema = CANONICAL_OUTPUT_REGISTRY[family]
    if not isinstance(payload, Mapping):
        raise ContractError("canonical output document must be an object")
    document = dict(payload)
    if (
        document.get("schema_version") != expected_schema
        or expected_output["route"] != expected_route
        or expected_output["schema_version"] != expected_schema
    ):
        raise ContractError("canonical output violates the route/schema registry")
    try:
        if family == "frame-observations":
            validated = validate_observation_batch(document)
            if validated != document:
                raise ContractError(
                    "canonical frame document differs from authoritative validation"
                )
            estimate_field = str(expected_output["metadata"]["estimate_field"])
            if estimate_field in _STRUCTURAL_VALUE_FIELDS:
                for index, cell in enumerate(document["cells"]):
                    estimate = cell["estimate"]
                    if (
                        isinstance(estimate, (int, float))
                        and not isinstance(estimate, bool)
                        and 0 < estimate < minimum_cell_size
                        and not cell["suppressed"]
                    ):
                        raise ContractError(
                            f"canonical frame cell {index} is below the approved "
                            "minimum cell size"
                        )
        elif family == "outcome-feedback":
            validated = validate_outcome_feedback(document)["canonical_copy"]
            if validated != document:
                raise ContractError(
                    "canonical outcome document differs from authoritative validation"
                )
        elif family in {"structured-evidence", "social-observations"}:
            ledger = build_evidence_ledger(
                "authorized-output-validation",
                [document],
                created_at=str(expected_output["metadata"]["created_at"]),
            )
            item_key = "items" if family == "structured-evidence" else "observations"
            if ledger["summary"]["accepted_items"] != len(document[item_key]):
                raise ContractError(
                    "canonical evidence document was not fully accepted"
                )
        else:
            document = exact_object(
                document,
                _CANONICAL_DOCUMENT_KEYS,
                "canonical output document",
            )
    except (
        AudienceResearchV3ValidationError,
        PanelBuilderContractError,
    ) as exc:
        raise ContractError(
            f"canonical {family} document failed authoritative validation: {exc}"
        ) from exc
    return document


def _document_items(
    output: Mapping[str, object],
    document: Mapping[str, object],
) -> list[Mapping[str, object]]:
    family = str(output["filename"]).removesuffix(".json").rsplit("-", 1)[0]
    if family == "frame-observations":
        return list(document["cells"])
    if family == "structured-evidence":
        return list(document["items"])
    if family == "social-observations":
        return list(document["observations"])
    if family == "outcome-feedback":
        return [document]
    return list(document["records"])


def _document_field_count(items: list[Mapping[str, object]]) -> int:
    return len(set().union(*(item.keys() for item in items))) if items else 0


def _report_from_run(
    *,
    profile: Mapping[str, object],
    mapping: Mapping[str, object],
    transformer_version: str,
    source_reads: list[dict[str, object]],
    operation_log: list[Audit],
    documents: list[
        tuple[Mapping[str, object], dict[str, object], bytes, str, str]
    ],
) -> dict[str, object]:
    renamed: list[dict[str, object]] = []
    casts: list[dict[str, object]] = []
    combined: list[dict[str, object]] = []
    dropped: list[dict[str, object]] = []
    category_merges: list[dict[str, object]] = []
    missing_values = suppressed_values = filtered_values = unmapped_values = 0
    consequences: list[str] = []
    for record in operation_log:
        details = record["details"]
        if "renamed_fields" in details:
            renamed.extend(
                {
                    "source": record["operation_id"],
                    "field": f"{source}->{target}",
                    "reason": "declared rename",
                }
                for source, target in details["renamed_fields"].items()
            )
        if "cast_fields" in details:
            casts.extend(
                {
                    "source": record["operation_id"],
                    "field": f"{field}:{cast_type}",
                    "reason": "declared cast",
                }
                for field, cast_type in details["cast_fields"].items()
            )
        if details.get("dropped_fields"):
            dropped.extend(
                {
                    "source": record["operation_id"],
                    "field": field,
                    "reason": "declared select drop",
                }
                for field in details["dropped_fields"]
            )
        if "join" in details or "reshape" in details or "aggregate_group_by" in details:
            combined.append(
                {
                    "source": record["operation_id"],
                    "field": "*",
                    "reason": str(record["op"]),
                }
            )
        if details.get("category_merges"):
            category_merges.append({"operation_id": record["operation_id"], "values": details["category_merges"]})
        missing_values += int(details.get("missing_values_normalized", 0))
        suppressed_values += int(details.get("suppressed_values_normalized", 0))
        filtered_values += int(details.get("filtered_values", 0))
        unmapped_values += int(details.get("unmapped_values", 0))
        for consequence in record["loss_consequences"]:
            if consequence not in consequences:
                consequences.append(str(consequence))
    ignored = [
        {
            "source": "#".join(
                (
                    str(item["file"]),
                    str(item["sheet"] or ""),
                    str(item["record_path"]),
                )
            ),
            "field": item["field"],
            "reason": item["reason"],
        }
        for item in mapping["ignored_fields"]
    ]
    if ignored and "omitted_source_coverage" not in consequences:
        consequences.append("omitted_source_coverage")
    if missing_values and "missing_normalization" not in consequences:
        consequences.append("missing_normalization")
    if suppressed_values and "suppression" not in consequences:
        consequences.append("suppression")
    if unmapped_values and "unmapped_category" not in consequences:
        consequences.append("unmapped_category")
    output_records = [
        {
            "path": output["filename"],
            "sha256": sha256_bytes(data),
            "route": output["route"],
            "schema_version": output["schema_version"],
            "row_count": len(_document_items(output, document)),
            "field_count": _document_field_count(
                _document_items(output, document)
            ),
            "unit": source_unit,
            "denominator": source_denominator,
        }
        for output, document, data, source_unit, source_denominator in documents
    ]
    status = "complete_with_loss" if consequences or unmapped_values or missing_values or suppressed_values else "complete"
    report = {
        "schema_version": TRANSFORMATION_REPORT_VERSION,
        "source_profile": {
            "path": "approved-source-profile.json",
            "sha256": mapping_sha256(profile),
        },
        "mapping": {
            "path": "approved-mapping.json",
            "sha256": sha256_bytes(canonical_json_bytes(mapping)),
        },
        "transformer_version": transformer_version,
        "input_hashes": dict(sorted(mapping["input_hashes"].items())),
        "source_reads": source_reads,
        "operation_log": operation_log,
        "field_changes": {
            "renamed": renamed,
            "cast": casts,
            "combined": combined,
            "ignored": ignored,
            "dropped": ignored + dropped,
        },
        "value_changes": {
            "category_merges": category_merges,
            "unmapped_values": unmapped_values,
            "missing_values": missing_values,
            "suppressed_values": suppressed_values,
            "rejected_values": 0,
            "filtered_values": filtered_values,
        },
        "route_summary": [
            {
                "route": output["route"],
                "path": output["filename"],
                "schema_version": output["schema_version"],
                "row_count": len(_document_items(output, document)),
                "field_count": _document_field_count(
                    _document_items(output, document)
                ),
            }
            for output, document, _, _, _ in documents
        ],
        "loss_summary": {
            "consequences": consequences,
            "coverage": "all_loss_and_coverage_consequences_reported" if status == "complete_with_loss" else "all_selected_rows_and_fields_preserved",
        },
        "warnings": [],
        "blocking_errors": [],
        "outputs": output_records,
        "status": status,
    }
    return validate_transformation_report(report)


def _require_digest(value: object, path: str) -> str:
    digest = require_string(value, path)
    if not _DIGEST.fullmatch(digest):
        raise ContractError(f"{path} must be a SHA-256 digest")
    return digest


def _array(value: object, path: str) -> list[object]:
    if not isinstance(value, list):
        raise ContractError(f"{path} must be an array")
    return value


def _require_string_mapping(value: object, path: str) -> dict[str, str]:
    if not isinstance(value, Mapping):
        raise ContractError(f"{path} must be an object")
    result: dict[str, str] = {}
    for key, item in value.items():
        result[require_string(key, f"{path} key")] = require_string(
            item, f"{path}.{key}"
        )
    return result


def _require_report_literal(value: object, path: str) -> object:
    if isinstance(value, list):
        for index, item in enumerate(value):
            _require_report_literal(item, f"{path}[{index}]")
        return value
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    raise ContractError(f"{path} must be a scalar or array of scalars")


def _validate_operation_detail_values(
    op: str,
    details: Mapping[str, object],
    path: str,
) -> None:
    if op == "select":
        require_string_list(details["selected_fields"], f"{path}.selected_fields", nonempty=True)
        require_string_list(details["dropped_fields"], f"{path}.dropped_fields")
    elif op in {"rename", "flatten"}:
        key = "renamed_fields" if op == "rename" else "flattened_paths"
        _require_string_mapping(details[key], f"{path}.{key}")
    elif op == "cast":
        casts = _require_string_mapping(details["cast_fields"], f"{path}.cast_fields")
        for field, cast_type in casts.items():
            require_enum(
                cast_type,
                {"string", "integer", "number", "boolean", "date"},
                f"{path}.cast_fields.{field}",
            )
    elif op == "wide_to_long":
        require_enum(details["reshape"], {"wide_to_long"}, f"{path}.reshape")
        require_string_list(details["id_fields"], f"{path}.id_fields")
        require_string_list(details["value_fields"], f"{path}.value_fields", nonempty=True)
        require_string(details["name_field"], f"{path}.name_field")
        require_string(details["value_field"], f"{path}.value_field")
    elif op == "pivot":
        require_enum(details["reshape"], {"pivot"}, f"{path}.reshape")
        require_string_list(details["index_fields"], f"{path}.index_fields", nonempty=True)
        require_string(details["column_field"], f"{path}.column_field")
        require_string(details["value_field"], f"{path}.value_field")
        require_string_list(details["columns"], f"{path}.columns", nonempty=True)
    elif op == "join":
        join = exact_object(
            details["join"],
            {"left", "right", "on", "cardinality", "matched_rows"},
            f"{path}.join",
        )
        require_identifier(join["left"], f"{path}.join.left")
        require_identifier(join["right"], f"{path}.join.right")
        require_string_list(join["on"], f"{path}.join.on", nonempty=True)
        require_enum(
            join["cardinality"],
            {"one_to_one", "many_to_one"},
            f"{path}.join.cardinality",
        )
        require_integer(join["matched_rows"], f"{path}.join.matched_rows", minimum=0)
        require_integer(details["unused_right_rows"], f"{path}.unused_right_rows", minimum=0)
    elif op == "category_map":
        require_string(details["field"], f"{path}.field")
        _require_string_mapping(details["mapping"], f"{path}.mapping")
        merges = details["category_merges"]
        if not isinstance(merges, Mapping) or not all(
            isinstance(key, str)
            and isinstance(value, int)
            and not isinstance(value, bool)
            and value >= 0
            for key, value in merges.items()
        ):
            raise ContractError(f"{path}.category_merges must map strings to counts")
        require_integer(details["unmapped_values"], f"{path}.unmapped_values", minimum=0)
        require_enum(
            details["unmapped_policy"],
            {"error", "keep", "null"},
            f"{path}.unmapped_policy",
        )
    elif op == "normalize_missing":
        require_integer(
            details["missing_values_normalized"],
            f"{path}.missing_values_normalized",
            minimum=0,
        )
        require_string_list(details["fields"], f"{path}.fields", nonempty=True)
        values = _array(details["values"], f"{path}.values")
        if not values:
            raise ContractError(f"{path}.values must not be empty")
        for index, value in enumerate(values):
            _require_report_literal(value, f"{path}.values[{index}]")
    elif op == "normalize_suppression":
        require_integer(
            details["suppressed_values_normalized"],
            f"{path}.suppressed_values_normalized",
            minimum=0,
        )
        require_string(details["field"], f"{path}.field")
        values = _array(details["values"], f"{path}.values")
        if not values:
            raise ContractError(f"{path}.values must not be empty")
        for index, value in enumerate(values):
            _require_report_literal(value, f"{path}.values[{index}]")
        require_string(details["status_field"], f"{path}.status_field")
    elif op == "derive_share":
        for key in ("derived_field", "count_field", "denominator_field"):
            require_string(details[key], f"{path}.{key}")
    elif op == "normalize_weight":
        require_string(details["source_field"], f"{path}.source_field")
        require_string(
            details["normalized_weight_field"],
            f"{path}.normalized_weight_field",
        )
        require_string_list(details["group_by"], f"{path}.group_by")
    elif op == "aggregate":
        require_string_list(details["aggregate_group_by"], f"{path}.aggregate_group_by")
        metrics = details["aggregate_metrics"]
        if not isinstance(metrics, Mapping) or not metrics:
            raise ContractError(f"{path}.aggregate_metrics must be a nonempty object")
        for output_field, raw_metric in metrics.items():
            require_string(output_field, f"{path}.aggregate_metrics key")
            metric = exact_object(
                raw_metric,
                {"field", "function"},
                f"{path}.aggregate_metrics.{output_field}",
            )
            require_string(metric["field"], f"{path}.aggregate_metrics.{output_field}.field")
            require_enum(
                metric["function"],
                {"sum", "count", "min", "max", "mean"},
                f"{path}.aggregate_metrics.{output_field}.function",
            )
    elif op == "filter":
        predicate = require_enum(
            details["predicate"],
            {"equals", "not_equals", "in", "not_in", "is_null", "is_not_null"},
            f"{path}.predicate",
        )
        require_string(details["field"], f"{path}.field")
        value = _require_report_literal(details["value"], f"{path}.value")
        if predicate in {"in", "not_in"} and not isinstance(value, list):
            raise ContractError(f"{path}.value must be an array for {predicate}")
        if predicate in {"is_null", "is_not_null"} and value is not None:
            raise ContractError(f"{path}.value must be null for {predicate}")
        require_integer(details["filtered_values"], f"{path}.filtered_values", minimum=0)
    elif op == "sort":
        require_string_list(details["sort_fields"], f"{path}.sort_fields", nonempty=True)


def _declared_operation_details(
    operation: Mapping[str, object],
    details: Mapping[str, object],
) -> dict[str, object]:
    op = str(operation["op"])
    if op == "select":
        return {"selected_fields": details["selected_fields"]}
    if op == "rename":
        return {"renamed_fields": details["renamed_fields"]}
    if op == "cast":
        return {"cast_fields": details["cast_fields"]}
    if op == "flatten":
        return {"flattened_paths": details["flattened_paths"]}
    if op == "wide_to_long":
        return {
            key: details[key]
            for key in ("id_fields", "value_fields", "name_field", "value_field")
        }
    if op == "pivot":
        return {
            key: details[key]
            for key in ("index_fields", "column_field", "value_field", "columns")
        }
    if op == "join":
        return {
            key: details["join"][key]
            for key in ("left", "right", "on", "cardinality")
        }
    if op == "category_map":
        return {
            "field": details["field"],
            "mapping": details["mapping"],
            "unmapped": details["unmapped_policy"],
        }
    if op == "normalize_missing":
        return {"fields": details["fields"], "values": details["values"]}
    if op == "normalize_suppression":
        return {
            key: details[key]
            for key in ("field", "values", "status_field")
        }
    if op == "derive_share":
        return {
            "output_field": details["derived_field"],
            "count_field": details["count_field"],
            "denominator_field": details["denominator_field"],
        }
    if op == "normalize_weight":
        return {
            "field": details["source_field"],
            "output_field": details["normalized_weight_field"],
            "group_by": details["group_by"],
        }
    if op == "aggregate":
        return {
            "group_by": details["aggregate_group_by"],
            "metrics": details["aggregate_metrics"],
        }
    if op == "filter":
        return {
            key: details[key]
            for key in ("field", "predicate", "value")
        }
    if op == "sort":
        return {"fields": details["sort_fields"]}
    raise ContractError(f"unsupported operation details: {op}")


def _approved_operation_details(
    operation: Mapping[str, object],
) -> dict[str, object]:
    op = str(operation["op"])
    if op == "select":
        return {"selected_fields": operation["fields"]}
    if op == "rename":
        return {"renamed_fields": operation["fields"]}
    if op == "cast":
        return {"cast_fields": operation["fields"]}
    if op == "flatten":
        return {"flattened_paths": operation["fields"]}
    if op == "wide_to_long":
        return {
            key: operation[key]
            for key in ("id_fields", "value_fields", "name_field", "value_field")
        }
    if op == "pivot":
        return {
            key: operation[key]
            for key in ("index_fields", "column_field", "value_field", "columns")
        }
    if op == "join":
        return {
            key: operation[key]
            for key in ("left", "right", "on", "cardinality")
        }
    if op == "category_map":
        return {
            key: operation[key]
            for key in ("field", "mapping", "unmapped")
        }
    if op == "normalize_missing":
        return {
            key: operation[key]
            for key in ("fields", "values")
        }
    if op == "normalize_suppression":
        return {
            key: operation[key]
            for key in ("field", "values", "status_field")
        }
    if op == "derive_share":
        return {
            key: operation[key]
            for key in ("output_field", "count_field", "denominator_field")
        }
    if op == "normalize_weight":
        return {
            key: operation[key]
            for key in ("field", "output_field", "group_by")
        }
    if op == "aggregate":
        return {
            key: operation[key]
            for key in ("group_by", "metrics")
        }
    if op == "filter":
        return {
            key: operation[key]
            for key in ("field", "predicate", "value")
        }
    if op == "sort":
        return {"fields": operation["fields"]}
    raise ContractError(f"unsupported approved operation: {op}")


def validate_transformation_report(payload: object) -> dict[str, object]:
    """Validate the complete deterministic loss and coverage report."""

    report = exact_object(payload, _REPORT_KEYS, "transformation report")
    if report["schema_version"] != TRANSFORMATION_REPORT_VERSION:
        raise ContractError("transformation report.schema_version is not supported")
    source_profile_ref = exact_object(
        report["source_profile"],
        _REPORT_REFERENCE_KEYS,
        "transformation report.source_profile",
    )
    mapping_ref = exact_object(
        report["mapping"],
        _REPORT_REFERENCE_KEYS,
        "transformation report.mapping",
    )
    if source_profile_ref["path"] != "approved-source-profile.json":
        raise ContractError("transformation report source profile path is invalid")
    if mapping_ref["path"] != "approved-mapping.json":
        raise ContractError("transformation report mapping path is invalid")
    _require_digest(source_profile_ref["sha256"], "transformation report.source_profile.sha256")
    _require_digest(mapping_ref["sha256"], "transformation report.mapping.sha256")
    require_string(report["transformer_version"], "transformation report.transformer_version")
    if not isinstance(report["input_hashes"], Mapping):
        raise ContractError("transformation report.input_hashes must be an object")
    for key, value in report["input_hashes"].items():
        require_string(key, "transformation report.input_hashes key")
        _require_digest(value, f"transformation report.input_hashes.{key}")
    selection_ids: set[str] = set()
    for index, raw in enumerate(_array(report["source_reads"], "transformation report.source_reads")):
        path = f"transformation report.source_reads[{index}]"
        item = exact_object(raw, _SOURCE_READ_KEYS, path)
        selection_id = require_identifier(item["selection_id"], f"{path}.selection_id")
        if selection_id in selection_ids:
            raise ContractError("transformation report source selection IDs must be unique")
        selection_ids.add(selection_id)
        file_name = require_string(item["file"], f"{path}.file")
        digest = _require_digest(item["file_sha256"], f"{path}.file_sha256")
        if report["input_hashes"].get(file_name) != digest:
            raise ContractError("transformation report source reads conflict with input hashes")
        if item["sheet"] is not None:
            require_string(item["sheet"], f"{path}.sheet")
        for key in ("record_path", "unit", "denominator"):
            require_string(item[key], f"{path}.{key}")
        require_integer(item["rows_read"], f"{path}.rows_read", minimum=0)
        require_string_list(item["fields_read"], f"{path}.fields_read", nonempty=True)
    operation_ids: set[str] = set()
    expected_consequences: list[str] = []
    computed_missing = computed_suppressed = computed_filtered = computed_unmapped = 0
    for index, raw in enumerate(_array(report["operation_log"], "transformation report.operation_log")):
        path = f"transformation report.operation_log[{index}]"
        item = exact_object(raw, _OPERATION_LOG_KEYS, path)
        operation_id = require_identifier(item["operation_id"], f"{path}.operation_id")
        if operation_id in operation_ids:
            raise ContractError("transformation report operation IDs must be unique")
        operation_ids.add(operation_id)
        op = require_enum(item["op"], set(ALLOWED_OPERATIONS), f"{path}.op")
        for key in ("input_rows", "output_rows", "filtered_rows"):
            require_integer(item[key], f"{path}.{key}", minimum=0)
        losses = require_string_list(
            item["loss_consequences"], f"{path}.loss_consequences"
        )
        if item["loss_consequence"] != (losses[0] if losses else None):
            raise ContractError(f"{path}.loss_consequence conflicts with loss_consequences")
        for loss in losses:
            if loss not in expected_consequences:
                expected_consequences.append(loss)
        details = exact_object(item["details"], _DETAIL_KEYS[op], f"{path}.details")
        _validate_operation_detail_values(op, details, f"{path}.details")
        if op == "category_map":
            computed_unmapped += require_integer(details["unmapped_values"], f"{path}.details.unmapped_values", minimum=0)
        if op == "normalize_missing":
            computed_missing += require_integer(details["missing_values_normalized"], f"{path}.details.missing_values_normalized", minimum=0)
        if op == "normalize_suppression":
            computed_suppressed += require_integer(details["suppressed_values_normalized"], f"{path}.details.suppressed_values_normalized", minimum=0)
        if op == "filter":
            computed_filtered += require_integer(details["filtered_values"], f"{path}.details.filtered_values", minimum=0)
    changes = exact_object(report["field_changes"], _FIELD_CHANGE_KEYS, "transformation report.field_changes")
    for key in _FIELD_CHANGE_KEYS:
        for index, raw in enumerate(_array(changes[key], f"transformation report.field_changes.{key}")):
            path = f"transformation report.field_changes.{key}[{index}]"
            item = exact_object(raw, {"source", "field", "reason"}, path)
            require_string(item["source"], f"{path}.source")
            require_string(item["field"], f"{path}.field")
            require_string(item["reason"], f"{path}.reason")
    if changes["ignored"] and "omitted_source_coverage" not in expected_consequences:
        expected_consequences.append("omitted_source_coverage")
    values = exact_object(report["value_changes"], _VALUE_CHANGE_KEYS, "transformation report.value_changes")
    for index, raw in enumerate(_array(values["category_merges"], "transformation report.value_changes.category_merges")):
        path = f"transformation report.value_changes.category_merges[{index}]"
        item = exact_object(raw, {"operation_id", "values"}, path)
        if require_identifier(item["operation_id"], f"{path}.operation_id") not in operation_ids:
            raise ContractError(f"{path}.operation_id does not resolve")
        if not isinstance(item["values"], Mapping) or not all(
            isinstance(key, str) and isinstance(value, int) and value >= 0
            for key, value in item["values"].items()
        ):
            raise ContractError(f"{path}.values must map strings to counts")
    for key in _VALUE_CHANGE_KEYS - {"category_merges"}:
        require_integer(values[key], f"transformation report.value_changes.{key}", minimum=0)
    if (
        values["missing_values"] != computed_missing
        or values["suppressed_values"] != computed_suppressed
        or values["filtered_values"] != computed_filtered
        or values["unmapped_values"] != computed_unmapped
    ):
        raise ContractError("transformation report value summaries lack coherence")
    if computed_missing and "missing_normalization" not in expected_consequences:
        expected_consequences.append("missing_normalization")
    if computed_suppressed and "suppression" not in expected_consequences:
        expected_consequences.append("suppression")
    if computed_unmapped and "unmapped_category" not in expected_consequences:
        expected_consequences.append("unmapped_category")
    route_summary: list[dict[str, object]] = []
    for index, raw in enumerate(_array(report["route_summary"], "transformation report.route_summary")):
        path = f"transformation report.route_summary[{index}]"
        item = exact_object(raw, _ROUTE_SUMMARY_KEYS, path)
        require_string(item["path"], f"{path}.path")
        require_enum(item["route"], {"structural_frame", "overlay_evidence", "profile_seed", "outcome_feedback"}, f"{path}.route")
        require_string(item["schema_version"], f"{path}.schema_version")
        require_integer(item["row_count"], f"{path}.row_count", minimum=0)
        require_integer(item["field_count"], f"{path}.field_count", minimum=0)
        route_summary.append(item)
    loss = exact_object(report["loss_summary"], _LOSS_SUMMARY_KEYS, "transformation report.loss_summary")
    consequences = require_string_list(loss["consequences"], "transformation report.loss_summary.consequences")
    if consequences != expected_consequences:
        raise ContractError("transformation report loss consequences lack coherence")
    coverage = require_enum(
        loss["coverage"],
        {
            "all_loss_and_coverage_consequences_reported",
            "all_selected_rows_and_fields_preserved",
        },
        "transformation report.loss_summary.coverage",
    )
    expected_coverage = (
        "all_loss_and_coverage_consequences_reported"
        if consequences
        else "all_selected_rows_and_fields_preserved"
    )
    if coverage != expected_coverage:
        raise ContractError("transformation report coverage lacks coherence")
    require_string_list(report["warnings"], "transformation report.warnings")
    require_string_list(report["blocking_errors"], "transformation report.blocking_errors")
    outputs: list[dict[str, object]] = []
    output_paths: set[str] = set()
    for index, raw in enumerate(_array(report["outputs"], "transformation report.outputs")):
        item = _validate_output_reference(raw, f"transformation report.outputs[{index}]")
        if item["path"] in output_paths:
            raise ContractError("transformation report output paths must be unique")
        output_paths.add(str(item["path"]))
        outputs.append(item)
    expected_summary = [
        {
            "path": item["path"],
            "route": item["route"],
            "schema_version": item["schema_version"],
            "row_count": item["row_count"],
            "field_count": item["field_count"],
        }
        for item in outputs
    ]
    if route_summary != expected_summary:
        raise ContractError("transformation report route summary lacks coherence")
    status = require_enum(
        report["status"],
        {"complete", "complete_with_loss", "blocked"},
        "transformation report.status",
    )
    if status == "blocked" and not report["blocking_errors"]:
        raise ContractError("blocked transformation report requires blocking_errors")
    if status != "blocked" and report["blocking_errors"]:
        raise ContractError("completed transformation report cannot contain blocking_errors")
    expected_status = "complete_with_loss" if consequences else "complete"
    if status != "blocked" and status != expected_status:
        raise ContractError("transformation report status lacks coherence")
    return report


def _validate_output_reference(raw: object, path: str) -> dict[str, object]:
    item = exact_object(raw, _OUTPUT_KEYS, path)
    output_path = require_string(item["path"], f"{path}.path")
    if not _SAFE_OUTPUT.fullmatch(output_path):
        raise ContractError(f"{path}.path must be one canonical relative output path")
    _require_digest(item["sha256"], f"{path}.sha256")
    route = require_enum(
        item["route"],
        {"structural_frame", "overlay_evidence", "profile_seed", "outcome_feedback"},
        f"{path}.route",
    )
    schema_version = require_string(item["schema_version"], f"{path}.schema_version")
    family = output_path.removesuffix(".json").rsplit("-", 1)[0]
    if (route, schema_version) != CANONICAL_OUTPUT_REGISTRY[family]:
        raise ContractError(f"{path} has an invalid route/schema registry pair")
    require_string(item["unit"], f"{path}.unit")
    require_string(item["denominator"], f"{path}.denominator")
    require_integer(item["row_count"], f"{path}.row_count", minimum=0)
    require_integer(item["field_count"], f"{path}.field_count", minimum=0)
    return item


def validate_authorized_handoff(
    payload: object,
    *,
    output_root: Path,
) -> dict[str, object]:
    """Validate handoff references against exact paths and on-disk hashes."""

    handoff = exact_object(payload, _HANDOFF_KEYS, "authorized handoff")
    if handoff["schema_version"] != AUTHORIZED_HANDOFF_VERSION:
        raise ContractError("authorized handoff.schema_version is not supported")
    require_enum(handoff["status"], {"complete", "complete_with_loss"}, "authorized handoff.status")
    source_profile_ref = exact_object(
        handoff["source_profile"],
        _REPORT_REFERENCE_KEYS,
        "authorized handoff.source_profile",
    )
    mapping_ref = exact_object(
        handoff["mapping"],
        _REPORT_REFERENCE_KEYS,
        "authorized handoff.mapping",
    )
    if source_profile_ref["path"] != "approved-source-profile.json":
        raise ContractError("authorized handoff source profile path is invalid")
    if mapping_ref["path"] != "approved-mapping.json":
        raise ContractError("authorized handoff mapping path is invalid")
    _require_digest(source_profile_ref["sha256"], "authorized handoff.source_profile.sha256")
    _require_digest(mapping_ref["sha256"], "authorized handoff.mapping.sha256")
    report_ref = exact_object(
        handoff["transformation_report"],
        _REPORT_REFERENCE_KEYS,
        "authorized handoff.transformation_report",
    )
    if report_ref["path"] != "transformation-report.json":
        raise ContractError("authorized handoff transformation report path is invalid")
    _require_digest(report_ref["sha256"], "authorized handoff.transformation_report.sha256")
    references = [
        _validate_output_reference(raw, f"authorized handoff.outputs[{index}]")
        for index, raw in enumerate(_array(handoff["outputs"], "authorized handoff.outputs"))
    ]
    paths = [str(item["path"]) for item in references]
    if len(paths) != len(set(paths)):
        raise ContractError("authorized handoff output paths must be unique")
    cohort_identity = exact_object(
        handoff["cohort_identity"],
        _COHORT_IDENTITY_KEYS,
        "authorized handoff.cohort_identity",
    )
    require_string(
        cohort_identity["cohort_id"],
        "authorized handoff.cohort_identity.cohort_id",
    )
    _require_digest(
        cohort_identity["source_profile_sha256"],
        "authorized handoff.cohort_identity.source_profile_sha256",
    )
    _require_digest(
        cohort_identity["source_bundle_sha256"],
        "authorized handoff.cohort_identity.source_bundle_sha256",
    )
    structural_identity = [
        exact_object(
            raw,
            _COHORT_STRUCTURAL_OUTPUT_KEYS,
            f"authorized handoff.cohort_identity.structural_outputs[{index}]",
        )
        for index, raw in enumerate(
            _array(
                cohort_identity["structural_outputs"],
                "authorized handoff.cohort_identity.structural_outputs",
            )
        )
    ]
    if not structural_identity:
        raise ContractError(
            "authorized handoff cohort identity requires structural outputs"
        )
    root = output_root.resolve()
    for path_text, expected_hash in [
        (str(source_profile_ref["path"]), source_profile_ref["sha256"]),
        (str(mapping_ref["path"]), mapping_ref["sha256"]),
        ("transformation-report.json", report_ref["sha256"]),
        *((str(item["path"]), item["sha256"]) for item in references),
    ]:
        path = Path(path_text)
        if path.is_absolute() or ".." in path.parts or len(path.parts) != 1:
            raise ContractError("authorized handoff requires a canonical relative output path")
        resolved = (root / path).resolve()
        if resolved.parent != root:
            raise ContractError("authorized handoff requires a canonical relative output path")
        if not resolved.is_file():
            raise ContractError(f"authorized handoff output is missing: {path_text}")
        if sha256_file(str(resolved)) != expected_hash:
            raise ContractError(f"authorized handoff output hash mismatch: {path_text}")
    try:
        report_payload = json.loads(
            (root / "transformation-report.json").read_text(encoding="utf-8")
        )
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ContractError("authorized handoff transformation report is unreadable") from exc
    report = validate_transformation_report(report_payload)
    try:
        profile_payload = json.loads(
            (root / "approved-source-profile.json").read_text(encoding="utf-8")
        )
        mapping_payload = json.loads(
            (root / "approved-mapping.json").read_text(encoding="utf-8")
        )
    except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ContractError("authorized handoff source profile or mapping is unreadable") from exc
    profile = validate_source_profile(profile_payload)
    approved_mapping = validate_authorized_mapping(mapping_payload, source_profile=profile)
    expected_reads = [
        {
            "selection_id": item["selection_id"],
            "file": item["file"],
            "file_sha256": item["file_sha256"],
            "sheet": item["sheet"],
            "record_path": item["record_path"],
            "fields_read": item["fields"],
            "unit": item["unit"],
            "denominator": item["denominator"],
        }
        for item in approved_mapping["selections"]
    ]
    observed_reads = [
        {
            key: item[key]
            for key in (
                "selection_id",
                "file",
                "file_sha256",
                "sheet",
                "record_path",
                "fields_read",
                "unit",
                "denominator",
            )
        }
        for item in report["source_reads"]
    ]
    expected_operations = [
        (
            item["operation_id"],
            item["op"],
            _approved_operation_details(item),
        )
        for item in approved_mapping["operations"]
    ]
    observed_operations = [
        (
            item["operation_id"],
            item["op"],
            _declared_operation_details(item, item["details"]),
        )
        for item in report["operation_log"]
    ]
    expected_outputs = [
        (item["filename"], item["route"], item["schema_version"])
        for item in approved_mapping["expected_outputs"]
    ]
    observed_outputs = [
        (item["path"], item["route"], item["schema_version"])
        for item in report["outputs"]
    ]
    if (
        report["source_profile"]["sha256"] != mapping_sha256(profile)
        or report["mapping"]["sha256"]
        != sha256_bytes(canonical_json_bytes(approved_mapping))
        or report["input_hashes"] != approved_mapping["input_hashes"]
        or observed_reads != expected_reads
        or observed_operations != expected_operations
        or observed_outputs != expected_outputs
    ):
        raise ContractError(
            "authorized handoff report does not cohere with its approved mapping"
        )
    if (
        handoff["status"] != report["status"]
        or handoff["source_profile"] != report["source_profile"]
        or handoff["mapping"] != report["mapping"]
        or references != report["outputs"]
    ):
        raise ContractError(
            "authorized handoff must exactly match its transformation report"
        )
    profile_seeds = require_string_list(handoff["profile_seeds"], "authorized handoff.profile_seeds")
    expected_seeds = sorted(item["path"] for item in references if item["route"] == "profile_seed")
    if profile_seeds != expected_seeds:
        raise ContractError("authorized handoff.profile_seeds must exactly reference profile-seed outputs")
    privacy = exact_object(
        handoff["privacy_permission"],
        _PRIVACY_PERMISSION_KEYS,
        "authorized handoff.privacy_permission",
    )
    if not require_boolean(privacy["permission_confirmed"], "authorized handoff.privacy_permission.permission_confirmed"):
        raise ContractError("authorized handoff permission must be confirmed")
    if not require_boolean(privacy["aggregate_only"], "authorized handoff.privacy_permission.aggregate_only"):
        raise ContractError("authorized handoff must remain aggregate-only")
    require_integer(privacy["minimum_cell_size"], "authorized handoff.privacy_permission.minimum_cell_size", minimum=1)
    mapping_privacy = approved_mapping["privacy_requirements"]
    if privacy != {
        "permission_confirmed": mapping_privacy["permission_confirmed"],
        "aggregate_only": mapping_privacy["aggregate_only"],
        "minimum_cell_size": mapping_privacy["minimum_cell_size"],
    }:
        raise ContractError(
            "authorized handoff privacy decision does not match its approved mapping"
        )
    minimum_cell_size = int(mapping_privacy["minimum_cell_size"])
    validated_documents: dict[str, dict[str, object]] = {}
    for expected_output in approved_mapping["expected_outputs"]:
        filename = str(expected_output["filename"])
        path = root / filename
        try:
            raw_bytes = path.read_bytes()
            output_payload = json.loads(raw_bytes.decode("utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ContractError(
                f"authorized handoff canonical output is unreadable: {filename}"
            ) from exc
        validated = _validate_canonical_document(
            output_payload,
            expected_output=expected_output,
            minimum_cell_size=minimum_cell_size,
        )
        validated_documents[filename] = validated
        if canonical_json_bytes(validated) != raw_bytes:
            raise ContractError(
                f"authorized handoff output is not canonical JSON: {filename}"
            )
    expected_structural_identity = [
        {
            "path": item["path"],
            "sha256": item["sha256"],
            "schema_version": item["schema_version"],
            "batch_id": validated_documents[str(item["path"])]["batch_id"],
            "unit": validated_documents[str(item["path"])]["unit"],
            "denominator": validated_documents[str(item["path"])][
                "denominator"
            ],
            "row_count": item["row_count"],
        }
        for item in references
        if item["route"] == "structural_frame"
    ]
    if cohort_identity != {
        "cohort_id": profile["profile_id"],
        "source_profile_sha256": source_profile_ref["sha256"],
        "source_bundle_sha256": profile["bundle_sha256"],
        "structural_outputs": expected_structural_identity,
    }:
        raise ContractError(
            "authorized handoff cohort identity does not match the exact "
            "source profile and structural outputs"
        )
    return handoff


def _atomic_publish_directory(
    staging_root: Path,
    output_root: Path,
) -> None:
    """Atomically publish a directory without replacing an existing path."""

    if os.name != "posix":
        raise ContractError(
            "atomic no-replace publication is unavailable on this platform"
        )
    libc = ctypes.CDLL(None, use_errno=True)
    source = os.fsencode(staging_root)
    destination = os.fsencode(output_root)
    if sys.platform == "darwin":
        try:
            rename_exclusive = libc.renamex_np
        except AttributeError as exc:
            raise ContractError(
                "atomic no-replace publication is unavailable on this POSIX runtime"
            ) from exc
        rename_exclusive.argtypes = [
            ctypes.c_char_p,
            ctypes.c_char_p,
            ctypes.c_uint,
        ]
        rename_exclusive.restype = ctypes.c_int
        result = rename_exclusive(source, destination, 0x00000004)
    elif sys.platform.startswith("linux"):
        try:
            rename_exclusive = libc.renameat2
        except AttributeError as exc:
            raise ContractError(
                "atomic no-replace publication is unavailable on this POSIX runtime"
            ) from exc
        rename_exclusive.argtypes = [
            ctypes.c_int,
            ctypes.c_char_p,
            ctypes.c_int,
            ctypes.c_char_p,
            ctypes.c_uint,
        ]
        rename_exclusive.restype = ctypes.c_int
        result = rename_exclusive(-100, source, -100, destination, 0x00000001)
    else:
        raise ContractError(
            "atomic no-replace publication is unavailable on this POSIX runtime"
        )
    if result == 0:
        return
    error_number = ctypes.get_errno()
    if error_number in {errno.EEXIST, errno.ENOTEMPTY}:
        raise ContractError(
            f"authorized audience output directory already exists: {output_root}"
        )
    raise OSError(
        error_number,
        os.strerror(error_number),
        str(output_root),
    )


def transform_authorized_bundle(
    *,
    source_profile: dict[str, object],
    mapping: dict[str, object],
    input_root: Path,
    output_dir: Path,
    transformer_version: str,
) -> dict[str, object]:
    """Validate first, transform deterministically, and create a no-clobber handoff."""

    profile = validate_source_profile(source_profile)
    approved_mapping = validate_authorized_mapping(mapping, source_profile=profile)
    require_string(transformer_version, "transformer_version")

    tables: dict[str, Rows] = {}
    source_reads: list[dict[str, object]] = []
    units: dict[str, tuple[str, str]] = {}
    for selection in approved_mapping["selections"]:
        rows, read_record = _read_selection(selection, Path(input_root))
        selection_id = str(selection["selection_id"])
        tables[selection_id] = rows
        source_reads.append(read_record)
        units[selection_id] = (str(selection["unit"]), str(selection["denominator"]))
    datasets, operation_log = apply_authorized_operations(tables, approved_mapping["operations"])
    ancestors = _dataset_ancestors(approved_mapping)

    documents: list[
        tuple[Mapping[str, object], dict[str, object], bytes, str, str]
    ] = []
    minimum_cell_size = int(approved_mapping["privacy_requirements"]["minimum_cell_size"])
    for expected_output in approved_mapping["expected_outputs"]:
        dataset = str(expected_output["dataset"])
        document, source_unit, source_denominator = _canonical_document(
            mapping=approved_mapping,
            output=expected_output,
            rows=datasets[dataset],
            units=units,
            ancestors=ancestors,
        )
        document = _validate_canonical_document(
            document,
            expected_output=expected_output,
            minimum_cell_size=minimum_cell_size,
        )
        documents.append(
            (
                expected_output,
                document,
                canonical_json_bytes(document),
                source_unit,
                source_denominator,
            )
        )
    report = _report_from_run(
        profile=profile,
        mapping=approved_mapping,
        transformer_version=transformer_version,
        source_reads=source_reads,
        operation_log=operation_log,
        documents=documents,
    )
    report_bytes = canonical_json_bytes(report)
    privacy = approved_mapping["privacy_requirements"]
    profile_bytes = canonical_json_bytes(profile)
    mapping_bytes = canonical_json_bytes(approved_mapping)
    handoff = {
        "schema_version": AUTHORIZED_HANDOFF_VERSION,
        "status": report["status"],
        "source_profile": {
            "path": "approved-source-profile.json",
            "sha256": sha256_bytes(profile_bytes),
        },
        "mapping": {
            "path": "approved-mapping.json",
            "sha256": sha256_bytes(mapping_bytes),
        },
        "transformation_report": {
            "path": "transformation-report.json",
            "sha256": sha256_bytes(report_bytes),
        },
        "outputs": report["outputs"],
        "profile_seeds": sorted(
            item["path"] for item in report["outputs"] if item["route"] == "profile_seed"
        ),
        "privacy_permission": {
            "permission_confirmed": privacy["permission_confirmed"],
            "aggregate_only": privacy["aggregate_only"],
            "minimum_cell_size": privacy["minimum_cell_size"],
        },
        "cohort_identity": {
            "cohort_id": profile["profile_id"],
            "source_profile_sha256": sha256_bytes(profile_bytes),
            "source_bundle_sha256": profile["bundle_sha256"],
            "structural_outputs": [
                {
                    "path": expected_output["filename"],
                    "sha256": sha256_bytes(document_bytes),
                    "schema_version": document["schema_version"],
                    "batch_id": document["batch_id"],
                    "unit": document["unit"],
                    "denominator": document["denominator"],
                    "row_count": len(
                        document.get("records", document.get("cells", []))
                    ),
                }
                for (
                    expected_output,
                    document,
                    document_bytes,
                    _source_unit,
                    _source_denominator,
                ) in documents
                if expected_output["route"] == "structural_frame"
            ],
        },
    }
    exact_object(handoff, _HANDOFF_KEYS, "authorized handoff")
    output_root = Path(output_dir)
    if output_root.exists():
        raise ContractError(
            f"authorized audience output directory already exists: {output_root}. "
            "Choose a new path; existing outputs are never overwritten."
        )
    try:
        output_root.parent.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise ContractError(
            f"could not prepare authorized audience output parent: {output_root.parent}"
        ) from exc
    temporary_root = Path(
        tempfile.mkdtemp(
            prefix=f".{output_root.name}.publishing-",
            dir=output_root.parent,
        )
    )
    try:
        write_new_bytes(
            temporary_root / "approved-source-profile.json",
            profile_bytes,
            "approved source profile copy",
        )
        write_new_bytes(
            temporary_root / "approved-mapping.json",
            mapping_bytes,
            "approved mapping copy",
        )
        for expected_output, _, data, _, _ in documents:
            write_new_bytes(
                temporary_root / str(expected_output["filename"]),
                data,
                "canonical authorized audience output",
            )
        write_new_bytes(
            temporary_root / "transformation-report.json",
            report_bytes,
            "transformation report",
        )
        write_new_bytes(
            temporary_root / "authorized-audience-handoff.json",
            canonical_json_bytes(handoff),
            "authorized audience handoff",
        )
        validated = validate_authorized_handoff(
            handoff,
            output_root=temporary_root,
        )
        if output_root.exists():
            raise ContractError(
                f"authorized audience output directory already exists: {output_root}"
            )
        _atomic_publish_directory(temporary_root, output_root)
        return validated
    except Exception:
        shutil.rmtree(temporary_root, ignore_errors=True)
        raise
