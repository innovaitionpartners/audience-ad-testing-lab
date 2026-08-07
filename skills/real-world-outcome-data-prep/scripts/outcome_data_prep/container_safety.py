from __future__ import annotations

import csv
from dataclasses import dataclass
import hashlib
from io import BytesIO, StringIO
import json
import os
from pathlib import Path, PurePosixPath
import posixpath
import re
import stat
import struct
import time
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
import zlib

from openpyxl import load_workbook

from .source_snapshot import SourceSnapshot


class ContainerSafetyError(ValueError):
    pass


@dataclass(frozen=True)
class ContainerLimits:
    compressed_bytes: int = 50_000_000
    uncompressed_bytes: int = 250_000_000
    member_count: int = 128
    recursion_depth: int = 1
    expansion_ratio: float = 20.0
    row_count: int = 2_000_000
    metadata_count: int = 100_000
    metadata_value_chars: int = 4_096
    metadata_total_chars: int = 2_000_000
    raw_value_count: int = 2_000_000
    raw_value_chars: int = 1_000_000
    raw_value_total_chars: int = 250_000_000
    logical_value_count: int = 1_000_000
    logical_value_chars: int = 1_000_000
    logical_value_total_chars: int = 250_000_000
    xml_bytes: int = 250_000_000
    processing_seconds: float = 30.0


@dataclass(frozen=True)
class InventoryCell:
    table: str
    row_number: int
    column_name: str
    value: str


@dataclass(frozen=True)
class InventoryMetadata:
    source: str
    name: str
    value: str


@dataclass(frozen=True)
class ContainerInventory:
    media_type: str
    tables: tuple[str, ...]
    headers: tuple[tuple[str, ...], ...]
    cells: tuple[InventoryCell, ...]
    row_count: int
    metadata: tuple[InventoryMetadata, ...] = ()
    raw_values: tuple[InventoryMetadata, ...] = ()
    logical_values: tuple[InventoryMetadata, ...] = ()


@dataclass
class _ArchiveBudget:
    member_count: int = 0
    compressed_bytes: int = 0
    uncompressed_bytes: int = 0

    def consume(
        self,
        members: list[zipfile.ZipInfo],
        limits: ContainerLimits,
    ) -> None:
        self.member_count += len(members)
        self.compressed_bytes += sum(
            member.compress_size for member in members
        )
        self.uncompressed_bytes += sum(member.file_size for member in members)
        if self.member_count > limits.member_count:
            raise ContainerSafetyError("archive member count limit exceeded")
        if self.compressed_bytes > limits.compressed_bytes:
            raise ContainerSafetyError("compressed byte limit exceeded")
        if self.uncompressed_bytes > limits.uncompressed_bytes:
            raise ContainerSafetyError("uncompressed byte limit exceeded")
        if (
            self.uncompressed_bytes / max(self.compressed_bytes, 1)
            > limits.expansion_ratio
        ):
            raise ContainerSafetyError("archive expansion ratio limit exceeded")


@dataclass
class _MetadataBudget:
    count: int = 0
    total_chars: int = 0
    xml_bytes: int = 0

    def consume_xml(self, byte_length: int, limits: ContainerLimits) -> None:
        self.xml_bytes += byte_length
        if self.xml_bytes > limits.xml_bytes:
            raise ContainerSafetyError("XLSX XML byte limit exceeded")

    def add(
        self,
        *,
        source: str,
        name: str,
        value: str,
        limits: ContainerLimits,
    ) -> InventoryMetadata:
        if not all(isinstance(item, str) for item in (source, name, value)):
            raise ContainerSafetyError("metadata value is invalid")
        if any("\x00" in item for item in (source, name, value)):
            raise ContainerSafetyError("NUL bytes are prohibited in metadata")
        if any(
            len(item) > limits.metadata_value_chars
            for item in (source, name, value)
        ):
            raise ContainerSafetyError("metadata value length limit exceeded")
        self.count += 1
        self.total_chars += len(source) + len(name) + len(value)
        if self.count > limits.metadata_count:
            raise ContainerSafetyError("metadata count limit exceeded")
        if self.total_chars > limits.metadata_total_chars:
            raise ContainerSafetyError("metadata character limit exceeded")
        return InventoryMetadata(source=source, name=name, value=value)


@dataclass
class _RawValueBudget:
    count: int = 0
    total_chars: int = 0

    def add(
        self,
        *,
        source: str,
        name: str,
        value: str,
        limits: ContainerLimits,
    ) -> InventoryMetadata:
        if not all(isinstance(item, str) for item in (source, name, value)):
            raise ContainerSafetyError("raw worksheet value is invalid")
        if any("\x00" in item for item in (source, name, value)):
            raise ContainerSafetyError(
                "NUL bytes are prohibited in raw worksheet values"
            )
        if len(value) > limits.raw_value_chars:
            raise ContainerSafetyError("raw worksheet value length limit exceeded")
        self.count += 1
        self.total_chars += len(value)
        if self.count > limits.raw_value_count:
            raise ContainerSafetyError("raw worksheet value count limit exceeded")
        if self.total_chars > limits.raw_value_total_chars:
            raise ContainerSafetyError(
                "raw worksheet value character limit exceeded"
            )
        return InventoryMetadata(source=source, name=name, value=value)


@dataclass
class _LogicalValueBudget:
    count: int = 0
    total_chars: int = 0

    def add(
        self,
        *,
        source: str,
        name: str,
        value: str,
        limits: ContainerLimits,
    ) -> InventoryMetadata:
        if not all(isinstance(item, str) for item in (source, name, value)):
            raise ContainerSafetyError("logical shared string is invalid")
        if any("\x00" in item for item in (source, name, value)):
            raise ContainerSafetyError(
                "NUL bytes are prohibited in logical shared strings"
            )
        if len(value) > limits.logical_value_chars:
            raise ContainerSafetyError("logical shared string length limit exceeded")
        self.count += 1
        self.total_chars += len(value)
        if self.count > limits.logical_value_count:
            raise ContainerSafetyError("logical shared string count limit exceeded")
        if self.total_chars > limits.logical_value_total_chars:
            raise ContainerSafetyError(
                "logical shared string character limit exceeded"
            )
        return InventoryMetadata(source=source, name=name, value=value)


_OOXML_WORKBOOK_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
}
_EXECUTABLE_SUFFIXES = {
    ".app",
    ".bat",
    ".bin",
    ".cmd",
    ".com",
    ".dll",
    ".dylib",
    ".exe",
    ".jar",
    ".js",
    ".msi",
    ".ps1",
    ".scr",
    ".sh",
    ".so",
    ".vbs",
}
_ALLOWED_XLSX_MEMBERS = (
    re.compile(r"^\[content_types\]\.xml$"),
    re.compile(r"^_rels/\.rels$"),
    re.compile(r"^docprops/(?:app|core|custom)\.xml$"),
    re.compile(r"^xl/workbook\.xml$"),
    re.compile(r"^xl/_rels/workbook\.xml\.rels$"),
    re.compile(r"^xl/styles\.xml$"),
    re.compile(r"^xl/sharedstrings\.xml$"),
    re.compile(r"^xl/theme/theme\d+\.xml$"),
    re.compile(r"^xl/worksheets/sheet\d+\.xml$"),
    re.compile(r"^xl/worksheets/_rels/sheet\d+\.xml\.rels$"),
)
_TRUTHY_XML = {"1", "true", "yes", "on"}
_JSON_MAX_NESTING = 64
_ZIP_LOCAL_HEADER = struct.Struct("<4s5H3L2H")
_ZIP_CENTRAL_HEADER = struct.Struct("<4s6H3L5H2L")
_ZIP_EOCD = struct.Struct("<4s4H2LH")
_ZIP_DATA_DESCRIPTOR = struct.Struct("<4L")
_ZIP_DATA_DESCRIPTOR_NO_SIGNATURE = struct.Struct("<3L")
_ZIP_UNICODE_PATH_EXTRA = 0x7075
_ZIP_UNICODE_COMMENT_EXTRA = 0x6375
_SPREADSHEET_NAMESPACE = (
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
_SHEET_DATA_CHILDREN = {
    "sheetData": frozenset({"row"}),
    "row": frozenset({"c"}),
    "c": frozenset({"f", "is", "v"}),
    "is": frozenset({"t"}),
    "f": frozenset(),
    "t": frozenset(),
    "v": frozenset(),
}
_CELL_REFERENCE = re.compile(r"^([A-Z]{1,3})([1-9][0-9]{0,6})$")
_MAX_XLSX_COLUMN = 16_384
_MAX_XLSX_ROW = 1_048_576


@dataclass(frozen=True)
class _WorksheetBounds:
    max_row: int
    max_column: int
    shared_string_indices: tuple[int, ...]


@dataclass(frozen=True)
class _SharedStringTable:
    count: int
    unique_count: int
    values: tuple[str, ...]


_XML_SPACE_ATTRIBUTE = "{http://www.w3.org/XML/1998/namespace}space"
_RICH_TEXT_PROPERTY_ATTRIBUTES = {
    "rFont": frozenset({"val"}),
    "charset": frozenset({"val"}),
    "family": frozenset({"val"}),
    "b": frozenset({"val"}),
    "i": frozenset({"val"}),
    "strike": frozenset({"val"}),
    "outline": frozenset({"val"}),
    "shadow": frozenset({"val"}),
    "condense": frozenset({"val"}),
    "extend": frozenset({"val"}),
    "color": frozenset({"rgb", "indexed", "auto", "theme", "tint"}),
    "sz": frozenset({"val"}),
    "u": frozenset({"val"}),
    "vertAlign": frozenset({"val"}),
    "scheme": frozenset({"val"}),
}


def _looks_executable(value: bytes) -> bool:
    return value.startswith(
        (
            b"MZ",
            b"\x7fELF",
            b"\xcf\xfa\xed\xfe",
            b"\xce\xfa\xed\xfe",
            b"\xfe\xed\xfa\xcf",
            b"\xfe\xed\xfa\xce",
            b"#!",
        )
    )


def _validate_limits(limits: ContainerLimits) -> None:
    integer_fields = (
        limits.compressed_bytes,
        limits.uncompressed_bytes,
        limits.member_count,
        limits.recursion_depth,
        limits.row_count,
        limits.metadata_count,
        limits.metadata_value_chars,
        limits.metadata_total_chars,
        limits.raw_value_count,
        limits.raw_value_chars,
        limits.raw_value_total_chars,
        limits.logical_value_count,
        limits.logical_value_chars,
        limits.logical_value_total_chars,
        limits.xml_bytes,
    )
    if any(
        isinstance(value, bool) or not isinstance(value, int) or value < 0
        for value in integer_fields
    ):
        raise ContainerSafetyError("container limits must be non-negative")
    if (
        isinstance(limits.expansion_ratio, bool)
        or not isinstance(limits.expansion_ratio, (int, float))
        or limits.expansion_ratio <= 0
        or isinstance(limits.processing_seconds, bool)
        or not isinstance(limits.processing_seconds, (int, float))
        or limits.processing_seconds <= 0
    ):
        raise ContainerSafetyError("container limits must be positive")


def _check_time(started: float, limits: ContainerLimits) -> None:
    if time.monotonic() - started > limits.processing_seconds:
        raise ContainerSafetyError("container processing time limit exceeded")


def _read_snapshot(
    snapshot: SourceSnapshot, limits: ContainerLimits
) -> bytes:
    if not isinstance(snapshot, SourceSnapshot):
        raise ContainerSafetyError("source snapshot is invalid")
    if snapshot.byte_length > limits.compressed_bytes:
        raise ContainerSafetyError("compressed byte limit exceeded")
    try:
        before = snapshot.staged_path.lstat()
        if stat.S_ISLNK(before.st_mode) or not stat.S_ISREG(before.st_mode):
            raise ContainerSafetyError("staged source is not a regular file")
        descriptor = os.open(
            snapshot.staged_path,
            os.O_RDONLY
            | getattr(os, "O_NOFOLLOW", 0)
            | getattr(os, "O_CLOEXEC", 0),
        )
    except ContainerSafetyError:
        raise
    except OSError as exc:
        raise ContainerSafetyError("staged source could not be opened") from exc
    try:
        opened = os.fstat(descriptor)
        if (
            opened.st_dev,
            opened.st_ino,
            opened.st_size,
            opened.st_mtime_ns,
            opened.st_ctime_ns,
        ) != (
            before.st_dev,
            before.st_ino,
            before.st_size,
            before.st_mtime_ns,
            before.st_ctime_ns,
        ):
            raise ContainerSafetyError("staged source identity changed")
        chunks: list[bytes] = []
        length = 0
        digest = hashlib.sha256()
        while True:
            chunk = os.read(descriptor, 1_048_576)
            if not chunk:
                break
            length += len(chunk)
            if length > limits.compressed_bytes:
                raise ContainerSafetyError("compressed byte limit exceeded")
            digest.update(chunk)
            chunks.append(chunk)
        after = os.fstat(descriptor)
        if (
            opened.st_size != after.st_size
            or opened.st_mtime_ns != after.st_mtime_ns
            or opened.st_ctime_ns != after.st_ctime_ns
            or length != snapshot.byte_length
            or "sha256:" + digest.hexdigest() != snapshot.source_sha256
        ):
            raise ContainerSafetyError("staged source no longer matches snapshot")
        return b"".join(chunks)
    except OSError as exc:
        raise ContainerSafetyError("staged source could not be read") from exc
    finally:
        os.close(descriptor)


def _decode_text(value: bytes) -> str:
    if b"\x00" in value:
        raise ContainerSafetyError("NUL bytes are prohibited")
    try:
        return value.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ContainerSafetyError("invalid text encoding") from exc


def _validate_headers(headers: tuple[str, ...]) -> None:
    if not headers:
        raise ContainerSafetyError("data table must declare headers")
    if len(set(headers)) != len(headers):
        raise ContainerSafetyError("duplicate header is prohibited")


def _inspect_delimited(
    value: bytes,
    *,
    delimiter: str,
    table: str,
    media_type: str,
    limits: ContainerLimits,
    started: float,
) -> ContainerInventory:
    text = _decode_text(value)
    try:
        reader = csv.reader(StringIO(text, newline=""), delimiter=delimiter, strict=True)
        try:
            header_row = next(reader)
        except StopIteration as exc:
            raise ContainerSafetyError("data table must declare headers") from exc
        headers = tuple(header_row)
        _validate_headers(headers)
        cells: list[InventoryCell] = []
        row_count = 0
        for row_number, row in enumerate(reader, start=2):
            _check_time(started, limits)
            if len(row) != len(headers):
                raise ContainerSafetyError("mixed row width is prohibited")
            row_count += 1
            if row_count > limits.row_count:
                raise ContainerSafetyError("row limit exceeded")
            cells.extend(
                InventoryCell(
                    table=table,
                    row_number=row_number,
                    column_name=header,
                    value=cell,
                )
                for header, cell in zip(headers, row, strict=True)
            )
    except ContainerSafetyError:
        raise
    except csv.Error as exc:
        raise ContainerSafetyError("malformed delimited data") from exc
    return ContainerInventory(
        media_type=media_type,
        tables=(table,),
        headers=(headers,),
        cells=tuple(cells),
        row_count=row_count,
    )


def _reject_duplicate_json_keys(
    pairs: list[tuple[str, object]],
) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in pairs:
        if key in result:
            raise ContainerSafetyError("duplicate key is prohibited in JSON")
        result[key] = value
    return result


def _reject_nonfinite_json(value: str) -> object:
    raise ContainerSafetyError("nonfinite JSON constants are prohibited")


def _enforce_json_depth(value: object, depth: int = 0) -> None:
    if depth > _JSON_MAX_NESTING:
        raise ContainerSafetyError("JSON nesting limit exceeded")
    if isinstance(value, dict):
        for item in value.values():
            _enforce_json_depth(item, depth + 1)
    elif isinstance(value, list):
        for item in value:
            _enforce_json_depth(item, depth + 1)


def _json_cell_string(value: object) -> str:
    if isinstance(value, str):
        return value
    if value is None:
        return "null"
    if value is True:
        return "true"
    if value is False:
        return "false"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
    raise ContainerSafetyError("JSON row value is unsupported")


def _json_cell_shape(value: object) -> object:
    if isinstance(value, dict):
        return (
            "object",
            tuple(
                (key, _json_cell_shape(item))
                for key, item in sorted(value.items())
            ),
        )
    if isinstance(value, list):
        shapes = {_json_cell_shape(item) for item in value}
        if len(shapes) > 1:
            raise ContainerSafetyError("mixed nested JSON shapes are prohibited")
        return ("array", next(iter(shapes), None))
    return "scalar"


def _inspect_json(
    value: bytes,
    *,
    table_hint: str,
    limits: ContainerLimits,
    started: float,
) -> ContainerInventory:
    text = _decode_text(value)
    try:
        document = json.loads(
            text,
            object_pairs_hook=_reject_duplicate_json_keys,
            parse_constant=_reject_nonfinite_json,
            parse_int=str,
            parse_float=str,
        )
    except ContainerSafetyError:
        raise
    except (json.JSONDecodeError, RecursionError, UnicodeError) as exc:
        raise ContainerSafetyError("malformed JSON data") from exc
    _enforce_json_depth(document)
    if isinstance(document, list):
        table = table_hint
        rows = document
    elif isinstance(document, dict) and len(document) == 1:
        table, rows = next(iter(document.items()))
        if not isinstance(table, str) or not table or not isinstance(rows, list):
            raise ContainerSafetyError(
                "JSON must contain one declared row array"
            )
    else:
        raise ContainerSafetyError("JSON must contain one declared row array")

    if not rows:
        return ContainerInventory(
            media_type="application/json",
            tables=(table,),
            headers=((),),
            cells=(),
            row_count=0,
        )
    first = rows[0]
    if not isinstance(first, dict):
        raise ContainerSafetyError("JSON nesting or row shape is unsupported")
    headers = tuple(first)
    _validate_headers(headers)
    shapes = {header: _json_cell_shape(first[header]) for header in headers}
    cells: list[InventoryCell] = []
    for index, row in enumerate(rows, start=1):
        _check_time(started, limits)
        if not isinstance(row, dict) or set(row) != set(headers):
            raise ContainerSafetyError("JSON rows must use one closed schema")
        if any(_json_cell_shape(row[header]) != shapes[header] for header in headers):
            raise ContainerSafetyError("mixed nested JSON shapes are prohibited")
        if index > limits.row_count:
            raise ContainerSafetyError("row limit exceeded")
        cells.extend(
            InventoryCell(
                table=table,
                row_number=index,
                column_name=header,
                value=_json_cell_string(row[header]),
            )
            for header in headers
        )
    return ContainerInventory(
        media_type="application/json",
        tables=(table,),
        headers=(headers,),
        cells=tuple(cells),
        row_count=len(rows),
    )


def _normalized_member_name(name: str) -> str:
    if not isinstance(name, str) or "\x00" in name:
        raise ContainerSafetyError("NUL or invalid ZIP member name is prohibited")
    try:
        replaced = unicodedata.normalize("NFC", name.replace("\\", "/"))
    except UnicodeError as exc:
        raise ContainerSafetyError("invalid ZIP member name encoding") from exc
    if (
        not replaced
        or replaced.startswith("/")
        or re.match(r"^[A-Za-z]:", replaced)
        or ".." in PurePosixPath(replaced).parts
    ):
        raise ContainerSafetyError("unsafe member path is prohibited")
    normalized = posixpath.normpath(replaced)
    if normalized in {"", ".", ".."} or normalized.startswith("../"):
        raise ContainerSafetyError("unsafe member path is prohibited")
    return normalized.casefold()


def _validate_archive_inventory(
    archive: zipfile.ZipFile,
    *,
    limits: ContainerLimits,
    budget: _ArchiveBudget,
) -> list[tuple[zipfile.ZipInfo, str]]:
    members = archive.infolist()
    budget.consume(members, limits)
    normalized: set[str] = set()
    result: list[tuple[zipfile.ZipInfo, str]] = []
    for member in members:
        try:
            original_name = member.orig_filename
            decoded_name = member.filename
            if (
                not isinstance(original_name, str)
                or not isinstance(decoded_name, str)
                or original_name != decoded_name
                or "\x00" in original_name
                or "\x00" in decoded_name
            ):
                raise ContainerSafetyError(
                    "NUL or divergent ZIP member name is prohibited"
                )
            name = _normalized_member_name(decoded_name)
        except ContainerSafetyError:
            raise
        except UnicodeError as exc:
            raise ContainerSafetyError(
                "invalid ZIP member name encoding"
            ) from exc
        if name in normalized:
            raise ContainerSafetyError("duplicate normalized member is prohibited")
        normalized.add(name)
        if member.flag_bits & 0x1:
            raise ContainerSafetyError("encrypted archive members are prohibited")
        mode = member.external_attr >> 16
        if stat.S_ISLNK(mode):
            raise ContainerSafetyError("archive symlink members are prohibited")
        result.append((member, name))
    return result


def _validate_zip_envelope(
    value: bytes,
    archive: zipfile.ZipFile,
    members: list[tuple[zipfile.ZipInfo, str]],
) -> None:
    minimum_eocd = max(0, len(value) - (_ZIP_EOCD.size + 65_535))
    candidates: list[tuple[int, tuple[object, ...]]] = []
    cursor = minimum_eocd
    while True:
        offset = value.find(b"PK\x05\x06", cursor)
        if offset < 0:
            break
        if offset + _ZIP_EOCD.size <= len(value):
            fields = _ZIP_EOCD.unpack_from(value, offset)
            comment_length = fields[-1]
            if offset + _ZIP_EOCD.size + comment_length == len(value):
                candidates.append((offset, fields))
        cursor = offset + 1
    if len(candidates) != 1:
        raise ContainerSafetyError(
            "ZIP envelope must contain one full-range EOCD"
        )
    eocd_offset, fields = candidates[0]
    (
        signature,
        disk_number,
        central_disk,
        disk_entries,
        total_entries,
        central_size,
        central_offset,
        comment_length,
    ) = fields
    if (
        signature != b"PK\x05\x06"
        or disk_number != 0
        or central_disk != 0
        or disk_entries != total_entries
        or total_entries != len(members)
        or central_offset + central_size != eocd_offset
        or getattr(archive, "start_dir", None) != central_offset
        or value[eocd_offset + _ZIP_EOCD.size:] != archive.comment
        or len(archive.comment) != comment_length
    ):
        raise ContainerSafetyError("unsupported ZIP envelope is prohibited")

    ordered_members = sorted(
        (member for member, _normalized in members),
        key=lambda member: member.header_offset,
    )
    if not ordered_members or ordered_members[0].header_offset != 0:
        raise ContainerSafetyError("ZIP prefix bytes are prohibited")
    for index, member in enumerate(ordered_members):
        offset = member.header_offset
        if offset + _ZIP_LOCAL_HEADER.size > len(value):
            raise ContainerSafetyError("malformed ZIP local record")
        local = _ZIP_LOCAL_HEADER.unpack_from(value, offset)
        (
            local_signature,
            _extract_version,
            local_flags,
            local_compression,
            _modified_time,
            _modified_date,
            local_crc,
            local_compressed_size,
            local_uncompressed_size,
            filename_length,
            extra_length,
        ) = local
        if (
            local_signature != b"PK\x03\x04"
            or local_flags != member.flag_bits
            or local_compression != member.compress_type
        ):
            raise ContainerSafetyError("divergent ZIP local record")
        data_start = (
            offset + _ZIP_LOCAL_HEADER.size + filename_length + extra_length
        )
        data_end = data_start + member.compress_size
        if data_end > len(value):
            raise ContainerSafetyError("malformed ZIP member extent")
        record_end = data_end
        if local_flags & 0x08:
            if value[data_end:data_end + 4] == b"PK\x07\x08":
                if data_end + _ZIP_DATA_DESCRIPTOR.size > len(value):
                    raise ContainerSafetyError("malformed ZIP data descriptor")
                descriptor = _ZIP_DATA_DESCRIPTOR.unpack_from(value, data_end)
                _descriptor_signature, crc, compressed_size, uncompressed_size = (
                    descriptor
                )
                record_end += _ZIP_DATA_DESCRIPTOR.size
            else:
                if data_end + _ZIP_DATA_DESCRIPTOR_NO_SIGNATURE.size > len(value):
                    raise ContainerSafetyError("malformed ZIP data descriptor")
                crc, compressed_size, uncompressed_size = (
                    _ZIP_DATA_DESCRIPTOR_NO_SIGNATURE.unpack_from(value, data_end)
                )
                record_end += _ZIP_DATA_DESCRIPTOR_NO_SIGNATURE.size
            if (
                crc != member.CRC
                or compressed_size != member.compress_size
                or uncompressed_size != member.file_size
            ):
                raise ContainerSafetyError("divergent ZIP data descriptor")
        elif (
            local_crc != member.CRC
            or local_compressed_size != member.compress_size
            or local_uncompressed_size != member.file_size
        ):
            raise ContainerSafetyError("divergent ZIP local sizes")
        expected_end = (
            ordered_members[index + 1].header_offset
            if index + 1 < len(ordered_members)
            else central_offset
        )
        if record_end != expected_end:
            raise ContainerSafetyError(
                "unsupported ZIP inter-record slack is prohibited"
            )

    central_cursor = central_offset
    members_in_central_order = [member for member, _normalized in members]
    for member in members_in_central_order:
        if central_cursor + _ZIP_CENTRAL_HEADER.size > eocd_offset:
            raise ContainerSafetyError("malformed ZIP central directory")
        central = _ZIP_CENTRAL_HEADER.unpack_from(value, central_cursor)
        (
            central_signature,
            _created_version,
            _extract_version,
            central_flags,
            central_compression,
            _modified_time,
            _modified_date,
            central_crc,
            central_compressed_size,
            central_uncompressed_size,
            filename_length,
            extra_length,
            member_comment_length,
            disk_start,
            _internal_attributes,
            _external_attributes,
            local_offset,
        ) = central
        if (
            central_signature != b"PK\x01\x02"
            or central_flags != member.flag_bits
            or central_compression != member.compress_type
            or central_crc != member.CRC
            or central_compressed_size != member.compress_size
            or central_uncompressed_size != member.file_size
            or disk_start != 0
            or local_offset != member.header_offset
        ):
            raise ContainerSafetyError("divergent ZIP central record")
        central_cursor += (
            _ZIP_CENTRAL_HEADER.size
            + filename_length
            + extra_length
            + member_comment_length
        )
    if central_cursor != eocd_offset:
        raise ContainerSafetyError(
            "unsupported ZIP central-directory slack is prohibited"
        )


def _decode_archive_metadata(value: bytes, *, label: str) -> str:
    try:
        text = value.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ContainerSafetyError(
            f"binary or unsupported {label} is prohibited"
        ) from exc
    if any(
        unicodedata.category(character) == "Cc"
        and character not in {"\t", "\n", "\r"}
        for character in text
    ):
        raise ContainerSafetyError(
            f"binary or unsupported {label} is prohibited"
        )
    return text


def _local_archive_fields(
    value: bytes,
    member: zipfile.ZipInfo,
) -> tuple[bytes, str, bytes]:
    offset = member.header_offset
    if (
        isinstance(offset, bool)
        or not isinstance(offset, int)
        or offset < 0
        or offset + _ZIP_LOCAL_HEADER.size > len(value)
    ):
        raise ContainerSafetyError("malformed ZIP local header is prohibited")
    try:
        (
            signature,
            _extract_version,
            local_flags,
            _compression,
            _modified_time,
            _modified_date,
            _crc,
            _compressed_size,
            _uncompressed_size,
            filename_length,
            extra_length,
        ) = _ZIP_LOCAL_HEADER.unpack_from(value, offset)
    except struct.error as exc:
        raise ContainerSafetyError(
            "malformed ZIP local header is prohibited"
        ) from exc
    if signature != b"PK\x03\x04" or local_flags != member.flag_bits:
        raise ContainerSafetyError("divergent ZIP local header is prohibited")
    filename_start = offset + _ZIP_LOCAL_HEADER.size
    filename_end = filename_start + filename_length
    extra_end = filename_end + extra_length
    if extra_end > len(value):
        raise ContainerSafetyError("malformed ZIP local header is prohibited")
    raw_filename = value[filename_start:filename_end]
    encoding = "utf-8" if member.flag_bits & 0x800 else "cp437"
    try:
        decoded_filename = raw_filename.decode(encoding)
        expected_filename = member.orig_filename.encode(encoding)
    except (AttributeError, UnicodeError) as exc:
        raise ContainerSafetyError(
            "invalid ZIP member name encoding"
        ) from exc
    if (
        decoded_filename != member.orig_filename
        or decoded_filename != member.filename
        or "\x00" in decoded_filename
        or raw_filename != expected_filename
    ):
        raise ContainerSafetyError("divergent ZIP member name is prohibited")
    return raw_filename, decoded_filename, value[filename_end:extra_end]


def _parse_archive_extra_fields(value: bytes) -> tuple[tuple[int, bytes], ...]:
    fields: list[tuple[int, bytes]] = []
    observed: set[int] = set()
    cursor = 0
    while cursor < len(value):
        if len(value) - cursor < 4:
            raise ContainerSafetyError("malformed ZIP extra field is prohibited")
        header_id, data_length = struct.unpack_from("<HH", value, cursor)
        cursor += 4
        data_end = cursor + data_length
        if data_end > len(value):
            raise ContainerSafetyError("malformed ZIP extra field is prohibited")
        if header_id in observed:
            raise ContainerSafetyError("duplicate ZIP extra field is prohibited")
        observed.add(header_id)
        fields.append((header_id, value[cursor:data_end]))
        cursor = data_end
    return tuple(fields)


def _unicode_extra_text(
    *,
    header_id: int,
    payload: bytes,
    raw_filename: bytes,
    raw_comment: bytes,
) -> tuple[str, str]:
    if len(payload) < 5 or payload[0] != 1:
        raise ContainerSafetyError("malformed ZIP Unicode extra field is prohibited")
    expected_raw = (
        raw_filename
        if header_id == _ZIP_UNICODE_PATH_EXTRA
        else raw_comment
    )
    expected_crc = zlib.crc32(expected_raw) & 0xFFFFFFFF
    observed_crc = int.from_bytes(payload[1:5], "little")
    if observed_crc != expected_crc:
        raise ContainerSafetyError("divergent ZIP Unicode extra field is prohibited")
    label = (
        "ZIP Unicode path extra field"
        if header_id == _ZIP_UNICODE_PATH_EXTRA
        else "ZIP Unicode comment extra field"
    )
    text = _decode_archive_metadata(payload[5:], label=label)
    if header_id == _ZIP_UNICODE_PATH_EXTRA:
        _normalized_member_name(text)
    return label, text


def _inventory_archive_metadata(
    *,
    archive: zipfile.ZipFile,
    value: bytes,
    members: list[tuple[zipfile.ZipInfo, str]],
    limits: ContainerLimits,
    budget: _MetadataBudget,
    source: str,
) -> tuple[InventoryMetadata, ...]:
    result: list[InventoryMetadata] = []
    if archive.comment:
        result.append(
            budget.add(
                source=source,
                name="archive_comment",
                value=_decode_archive_metadata(
                    archive.comment,
                    label="ZIP archive comment",
                ),
                limits=limits,
            )
        )
    for index, (member, _normalized) in enumerate(members):
        raw_filename, decoded_filename, local_extra = _local_archive_fields(
            value,
            member,
        )
        result.append(
            budget.add(
                source=source,
                name=f"member[{index}].name",
                value=unicodedata.normalize(
                    "NFC", decoded_filename.replace("\\", "/")
                ),
                limits=limits,
            )
        )
        if member.comment:
            result.append(
                budget.add(
                    source=source,
                    name=f"member[{index}].comment",
                    value=_decode_archive_metadata(
                        member.comment,
                        label="ZIP member comment",
                    ),
                    limits=limits,
                )
            )
        if local_extra != member.extra:
            raise ContainerSafetyError("divergent ZIP extra fields are prohibited")
        for header_id, payload in _parse_archive_extra_fields(member.extra):
            if header_id not in {
                _ZIP_UNICODE_PATH_EXTRA,
                _ZIP_UNICODE_COMMENT_EXTRA,
            }:
                raise ContainerSafetyError(
                    "unsupported ZIP extra field is prohibited"
                )
            label, text = _unicode_extra_text(
                header_id=header_id,
                payload=payload,
                raw_filename=raw_filename,
                raw_comment=member.comment,
            )
            result.append(
                budget.add(
                    source=source,
                    name=f"member[{index}].{label}",
                    value=text,
                    limits=limits,
                )
            )
    return tuple(result)


def _strict_xml_text(value: bytes, purpose: str) -> str:
    try:
        text = value.decode("utf-8-sig")
    except UnicodeError as exc:
        raise ContainerSafetyError(
            f"unsupported XML encoding in {purpose}"
        ) from exc
    if "\x00" in text:
        raise ContainerSafetyError(f"unsupported XML encoding in {purpose}")
    declaration = re.match(
        r"^\s*<\?xml\s+([^?]*)\?>",
        text,
        flags=re.IGNORECASE,
    )
    if declaration:
        fields: list[tuple[str, str]] = []
        cursor = 0
        for matched in re.finditer(
            r"([A-Za-z]+)\s*=\s*(['\"])(.*?)\2",
            declaration.group(1),
        ):
            if declaration.group(1)[cursor:matched.start()].strip():
                raise ContainerSafetyError(
                    f"malformed XML declaration in {purpose}"
                )
            fields.append((matched.group(1).casefold(), matched.group(3)))
            cursor = matched.end()
        if declaration.group(1)[cursor:].strip():
            raise ContainerSafetyError(
                f"malformed XML declaration in {purpose}"
            )
        keys = tuple(key for key, _value in fields)
        if (
            not keys
            or keys[0] != "version"
            or keys not in {
                ("version",),
                ("version", "encoding"),
                ("version", "standalone"),
                ("version", "encoding", "standalone"),
            }
            or fields[0][1] != "1.0"
        ):
            raise ContainerSafetyError(
                f"malformed XML declaration in {purpose}"
            )
        values = dict(fields)
        if (
            "encoding" in values
            and re.sub(r"[-_]", "", values["encoding"]).casefold()
            != "utf8"
        ):
            raise ContainerSafetyError(
                f"unsupported XML encoding in {purpose}"
            )
        if (
            "standalone" in values
            and values["standalone"].casefold() not in {"yes", "no"}
        ):
            raise ContainerSafetyError(
                f"malformed XML declaration in {purpose}"
            )
        without_declaration = text[declaration.end():]
    else:
        without_declaration = text
    lowered = without_declaration.casefold()
    if "<!doctype" in lowered or "<!entity" in lowered:
        raise ContainerSafetyError(f"unsafe XML in {purpose}")
    if "<!--" in without_declaration or "<?" in without_declaration:
        raise ContainerSafetyError(
            f"unsupported XML metadata structure in {purpose}"
        )
    return without_declaration


def _parse_xml(value: bytes, purpose: str) -> ET.Element:
    text = _strict_xml_text(value, purpose)
    try:
        return ET.fromstring(text)
    except (ET.ParseError, UnicodeError, ValueError) as exc:
        raise ContainerSafetyError(f"malformed XML in {purpose}") from exc


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _qualified_namespace(tag: str) -> str:
    if tag.startswith("{") and "}" in tag:
        return tag[1:].split("}", 1)[0]
    return ""


def _cell_coordinates(value: str) -> tuple[int, int]:
    if not isinstance(value, str):
        raise ContainerSafetyError("worksheet cell reference is invalid")
    matched = _CELL_REFERENCE.fullmatch(value)
    if matched is None:
        raise ContainerSafetyError("worksheet cell reference is invalid")
    column = 0
    for character in matched.group(1):
        column = column * 26 + ord(character) - ord("A") + 1
    row = int(matched.group(2))
    if column > _MAX_XLSX_COLUMN or row > _MAX_XLSX_ROW:
        raise ContainerSafetyError("worksheet cell reference is invalid")
    return row, column


def _validate_cell_grammar(cell: ET.Element) -> int | None:
    children = list(cell)
    child_names = [_local_name(child.tag) for child in children]
    if "f" in child_names:
        if child_names not in (["f"], ["f", "v"]):
            raise ContainerSafetyError("worksheet formula cell grammar is invalid")
        return None
    cell_type = cell.attrib.get("t", "n")
    if cell_type == "inlineStr":
        if child_names != ["is"]:
            raise ContainerSafetyError(
                "worksheet inline-string cell grammar is invalid"
            )
        inline_children = list(children[0])
        if [_local_name(child.tag) for child in inline_children] != ["t"]:
            raise ContainerSafetyError(
                "worksheet inline-string cell grammar is invalid"
            )
        return None
    if cell_type not in {"n", "s", "str", "b", "e", "d"}:
        raise ContainerSafetyError("worksheet cell type is unsupported")
    if cell_type == "s":
        if child_names != ["v"]:
            raise ContainerSafetyError(
                "shared string cell must contain one index"
            )
        raw_index = children[0].text or ""
        if re.fullmatch(r"0|[1-9][0-9]*", raw_index) is None:
            raise ContainerSafetyError("shared string cell index is invalid")
        return int(raw_index)
    if child_names not in ([], ["v"]):
        raise ContainerSafetyError("worksheet scalar cell grammar is invalid")
    return None


def _worksheet_bounds(root: ET.Element) -> _WorksheetBounds:
    dimensions = [
        element
        for element in root.iter()
        if isinstance(element.tag, str)
        and _local_name(element.tag) == "dimension"
    ]
    if len(dimensions) != 1:
        raise ContainerSafetyError(
            "worksheet must declare one closed dimension"
        )
    reference = dimensions[0].attrib.get("ref")
    if not isinstance(reference, str) or not reference:
        raise ContainerSafetyError("worksheet dimension is invalid")
    endpoints = reference.split(":")
    if len(endpoints) == 1:
        endpoints.append(endpoints[0])
    if len(endpoints) != 2:
        raise ContainerSafetyError("worksheet dimension is invalid")
    minimum = _cell_coordinates(endpoints[0])
    maximum = _cell_coordinates(endpoints[1])
    if minimum != (1, 1) or maximum[0] < 1 or maximum[1] < 1:
        raise ContainerSafetyError(
            "worksheet dimension must begin at A1"
        )

    sheet_data = [
        element
        for element in root.iter()
        if isinstance(element.tag, str)
        and _local_name(element.tag) == "sheetData"
        and _qualified_namespace(element.tag) == _SPREADSHEET_NAMESPACE
    ]
    if len(sheet_data) != 1:
        raise ContainerSafetyError(
            "worksheet must contain one supported sheetData"
        )
    observed_rows: set[int] = set()
    observed_cells: set[tuple[int, int]] = set()
    shared_string_indices: list[int] = []
    previous_row = 0
    for row_element in sheet_data[0]:
        raw_row = row_element.attrib.get("r")
        if (
            not isinstance(raw_row, str)
            or not raw_row.isascii()
            or not raw_row.isdigit()
            or raw_row.startswith("0")
        ):
            raise ContainerSafetyError("worksheet row reference is invalid")
        row_number = int(raw_row)
        if not 1 <= row_number <= maximum[0]:
            raise ContainerSafetyError(
                "worksheet row falls outside its declared dimension"
            )
        if row_number in observed_rows or row_number <= previous_row:
            raise ContainerSafetyError(
                "worksheet row references must strictly increase"
            )
        observed_rows.add(row_number)
        previous_row = row_number
        previous_column = 0
        for cell_element in row_element:
            if _local_name(cell_element.tag) != "c":
                continue
            coordinates = _cell_coordinates(cell_element.attrib.get("r", ""))
            if (
                coordinates[0] != row_number
                or coordinates[0] > maximum[0]
                or coordinates[1] > maximum[1]
            ):
                raise ContainerSafetyError(
                    "worksheet cell falls outside its declared dimension"
                )
            if (
                coordinates[1] <= previous_column
                or coordinates in observed_cells
            ):
                raise ContainerSafetyError(
                    "worksheet cell references must strictly increase"
                )
            observed_cells.add(coordinates)
            previous_column = coordinates[1]
            shared_string_index = _validate_cell_grammar(cell_element)
            if shared_string_index is not None:
                shared_string_indices.append(shared_string_index)
    return _WorksheetBounds(
        max_row=maximum[0],
        max_column=maximum[1],
        shared_string_indices=tuple(shared_string_indices),
    )


def _shared_string_text(element: ET.Element) -> str:
    if (
        _local_name(element.tag) != "t"
        or _qualified_namespace(element.tag) != _SPREADSHEET_NAMESPACE
        or list(element)
        or set(element.attrib) - {_XML_SPACE_ATTRIBUTE}
        or (
            _XML_SPACE_ATTRIBUTE in element.attrib
            and element.attrib[_XML_SPACE_ATTRIBUTE] not in {"default", "preserve"}
        )
        or (element.tail or "").strip()
    ):
        raise ContainerSafetyError("shared string text grammar is invalid")
    return element.text or ""


def _shared_string_run(element: ET.Element) -> str:
    if (
        _local_name(element.tag) != "r"
        or _qualified_namespace(element.tag) != _SPREADSHEET_NAMESPACE
        or element.attrib
        or (element.text or "").strip()
        or (element.tail or "").strip()
    ):
        raise ContainerSafetyError("shared string rich-run grammar is invalid")
    children = list(element)
    if [_local_name(child.tag) for child in children] == ["t"]:
        return _shared_string_text(children[0])
    if [_local_name(child.tag) for child in children] != ["rPr", "t"]:
        raise ContainerSafetyError("shared string rich-run grammar is invalid")
    properties = children[0]
    if (
        _qualified_namespace(properties.tag) != _SPREADSHEET_NAMESPACE
        or properties.attrib
        or (properties.text or "").strip()
        or (properties.tail or "").strip()
    ):
        raise ContainerSafetyError("shared string rich properties are invalid")
    observed_properties: set[str] = set()
    for prop in properties:
        local = _local_name(prop.tag)
        if (
            _qualified_namespace(prop.tag) != _SPREADSHEET_NAMESPACE
            or local not in _RICH_TEXT_PROPERTY_ATTRIBUTES
            or local in observed_properties
            or set(prop.attrib) - _RICH_TEXT_PROPERTY_ATTRIBUTES[local]
            or list(prop)
            or (prop.text or "").strip()
            or (prop.tail or "").strip()
        ):
            raise ContainerSafetyError(
                "shared string rich properties are invalid"
            )
        observed_properties.add(local)
    return _shared_string_text(children[1])


def _parse_shared_string_table(
    root: ET.Element,
    *,
    limits: ContainerLimits,
    budget: _LogicalValueBudget,
    logical_values: list[InventoryMetadata],
) -> _SharedStringTable:
    if (
        _local_name(root.tag) != "sst"
        or _qualified_namespace(root.tag) != _SPREADSHEET_NAMESPACE
        or set(root.attrib) != {"count", "uniqueCount"}
        or (root.text or "").strip()
        or (root.tail or "").strip()
    ):
        raise ContainerSafetyError("shared string table grammar is invalid")
    raw_count = root.attrib["count"]
    raw_unique_count = root.attrib["uniqueCount"]
    if (
        re.fullmatch(r"0|[1-9][0-9]*", raw_count) is None
        or re.fullmatch(r"0|[1-9][0-9]*", raw_unique_count) is None
    ):
        raise ContainerSafetyError("shared string counts are invalid")
    values: list[str] = []
    for index, item in enumerate(root):
        if (
            _local_name(item.tag) != "si"
            or _qualified_namespace(item.tag) != _SPREADSHEET_NAMESPACE
            or item.attrib
            or (item.text or "").strip()
            or (item.tail or "").strip()
        ):
            raise ContainerSafetyError("shared string item grammar is invalid")
        children = list(item)
        if len(children) == 1 and _local_name(children[0].tag) == "t":
            logical_value = _shared_string_text(children[0])
        elif children and all(_local_name(child.tag) == "r" for child in children):
            logical_value = "".join(
                _shared_string_run(child) for child in children
            )
        else:
            raise ContainerSafetyError("shared string item grammar is invalid")
        values.append(logical_value)
        logical_values.append(
            budget.add(
                source="xlsx_logical_shared_string",
                name=f"xl/sharedstrings.xml:si[{index}]",
                value=logical_value,
                limits=limits,
            )
        )
    count = int(raw_count)
    unique_count = int(raw_unique_count)
    if unique_count != len(values) or len(set(values)) != len(values):
        raise ContainerSafetyError("shared string uniqueCount is invalid")
    return _SharedStringTable(
        count=count,
        unique_count=unique_count,
        values=tuple(values),
    )


def _xml_namespace_declarations(
    value: bytes,
    *,
    purpose: str,
) -> tuple[tuple[str, str], ...]:
    declarations: list[tuple[str, str]] = []
    text = _strict_xml_text(value, purpose)
    try:
        for _event, declaration in ET.iterparse(
            StringIO(text),
            events=("start-ns",),
        ):
            prefix, uri = declaration
            declarations.append((prefix or "", uri))
    except (ET.ParseError, UnicodeError) as exc:
        raise ContainerSafetyError(f"malformed XML in {purpose}") from exc
    return tuple(declarations)


def _inventory_xml_metadata(
    *,
    member_name: str,
    root: ET.Element,
    namespaces: tuple[tuple[str, str], ...],
    limits: ContainerLimits,
    budget: _MetadataBudget,
    raw_value_budget: _RawValueBudget,
    raw_values: list[InventoryMetadata],
    started: float,
) -> tuple[InventoryMetadata, ...]:
    result: list[InventoryMetadata] = []

    def add_name(kind: str, value: str) -> None:
        result.append(
            budget.add(
                source="xlsx_xml",
                name=f"{member_name}:{kind}",
                value=value,
                limits=limits,
            )
        )

    for prefix, uri in namespaces:
        add_name("namespace_prefix", prefix)
        add_name("namespace_uri", uri)

    is_worksheet = member_name.startswith("xl/worksheets/sheet")
    pending: list[tuple[ET.Element, str | None, bool]] = [
        (root, None, False)
    ]
    while pending:
        _check_time(started, limits)
        element, parent_local, parent_in_sheet_data = pending.pop()
        if not isinstance(element.tag, str):
            raise ContainerSafetyError(
                "unsupported XLSX XML node is prohibited"
            )
        local = _local_name(element.tag)
        in_sheet_data = parent_in_sheet_data or (
            is_worksheet and local == "sheetData"
        )
        if is_worksheet and local == "sheetData" and (
            _qualified_namespace(element.tag) != _SPREADSHEET_NAMESPACE
        ):
            raise ContainerSafetyError(
                "unsupported worksheet sheetData namespace is prohibited"
            )
        if parent_in_sheet_data:
            allowed = _SHEET_DATA_CHILDREN.get(parent_local or "")
            if (
                allowed is None
                or local not in allowed
                or _qualified_namespace(element.tag)
                != _SPREADSHEET_NAMESPACE
            ):
                raise ContainerSafetyError(
                    "unsupported worksheet sheetData child is prohibited"
                )
        add_name("qualified_element_name", element.tag)
        for raw_name, raw_value in sorted(element.attrib.items()):
            add_name("qualified_attribute_name", raw_name)
            result.append(
                budget.add(
                    source="xlsx_xml",
                    name=(
                        f"{member_name}:{element.tag}.attribute:{raw_name}"
                    ),
                    value=raw_value,
                    limits=limits,
                )
            )
        text = element.text or ""
        known_cell_value_text = (
            in_sheet_data
            and (
                (local == "v" and parent_local == "c")
                or (local == "t" and parent_local == "is")
            )
        ) or (
            member_name == "xl/sharedstrings.xml"
            and local == "t"
            and parent_local in {"r", "si"}
        )
        if (
            text.strip()
            and in_sheet_data
            and not known_cell_value_text
            and local != "f"
        ):
            raise ContainerSafetyError(
                "unsupported worksheet sheetData text is prohibited"
            )
        if known_cell_value_text:
            raw_values.append(
                raw_value_budget.add(
                    source="xlsx_raw_value",
                    name=f"{member_name}:{element.tag}.text",
                    value=text,
                    limits=limits,
                )
            )
        if text.strip() and not known_cell_value_text:
            result.append(
                budget.add(
                    source="xlsx_xml",
                    name=f"{member_name}:{element.tag}.text",
                    value=text,
                    limits=limits,
                )
            )
        tail = element.tail or ""
        if tail.strip() and in_sheet_data:
            raise ContainerSafetyError(
                "unsupported worksheet sheetData text is prohibited"
            )
        if tail.strip():
            result.append(
                budget.add(
                    source="xlsx_xml",
                    name=f"{member_name}:{element.tag}.tail",
                    value=tail,
                    limits=limits,
                )
            )
        pending.extend(
            (child, local, in_sheet_data)
            for child in reversed(list(element))
        )
    return tuple(result)


def _xlsx_member_is_allowed(name: str) -> bool:
    return any(pattern.fullmatch(name) for pattern in _ALLOWED_XLSX_MEMBERS)


def _inspect_xlsx(
    value: bytes,
    *,
    limits: ContainerLimits,
    started: float,
    table_prefix: str = "",
    budget: _ArchiveBudget | None = None,
    metadata_budget: _MetadataBudget | None = None,
    raw_value_budget: _RawValueBudget | None = None,
    logical_value_budget: _LogicalValueBudget | None = None,
) -> ContainerInventory:
    if value.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        raise ContainerSafetyError("encrypted workbook is prohibited")
    if not value.startswith(b"PK"):
        raise ContainerSafetyError("XLSX is not a recognized OOXML container")
    try:
        archive = zipfile.ZipFile(BytesIO(value))
    except (zipfile.BadZipFile, OSError, UnicodeError) as exc:
        raise ContainerSafetyError("XLSX is not a recognized OOXML container") from exc
    with archive:
        if budget is None:
            budget = _ArchiveBudget()
        if metadata_budget is None:
            metadata_budget = _MetadataBudget()
        if raw_value_budget is None:
            raw_value_budget = _RawValueBudget()
        if logical_value_budget is None:
            logical_value_budget = _LogicalValueBudget()
        raw_values: list[InventoryMetadata] = []
        logical_values: list[InventoryMetadata] = []
        members = _validate_archive_inventory(
            archive,
            limits=limits,
            budget=budget,
        )
        _validate_zip_envelope(value, archive, members)
        metadata = list(
            _inventory_archive_metadata(
                archive=archive,
                value=value,
                members=members,
                limits=limits,
                budget=metadata_budget,
                source="xlsx_archive",
            )
        )
        names = {name for _, name in members if not _.is_dir()}
        if "[content_types].xml" not in names or "xl/workbook.xml" not in names:
            raise ContainerSafetyError("XLSX is not a recognized OOXML container")
        if "_rels/.rels" not in names or "xl/_rels/workbook.xml.rels" not in names:
            raise ContainerSafetyError("XLSX is not a recognized OOXML container")

        raw_members: dict[str, bytes] = {}
        for member, normalized in members:
            _check_time(started, limits)
            if member.is_dir():
                continue
            try:
                member_value = archive.read(member)
            except (RuntimeError, OSError, zipfile.BadZipFile, UnicodeError) as exc:
                raise ContainerSafetyError("XLSX member could not be read") from exc
            if normalized.endswith((".xml", ".rels")):
                metadata_budget.consume_xml(len(member_value), limits)
            raw_members[normalized] = member_value

        content_types = _parse_xml(
            raw_members["[content_types].xml"], "XLSX content types"
        )
        _check_time(started, limits)
        workbook_types = {
            element.attrib.get("ContentType", "")
            for element in content_types
            if _local_name(element.tag) == "Override"
            and element.attrib.get("PartName") == "/xl/workbook.xml"
        }
        all_content_types = {
            element.attrib.get("ContentType", "").lower()
            for element in content_types
        }
        if any(
            "macroenabled" in content_type or "vba" in content_type
            for content_type in all_content_types
        ) or any("vbaproject" in name for name in names):
            raise ContainerSafetyError("VBA content is prohibited")
        if not workbook_types.intersection(_OOXML_WORKBOOK_TYPES):
            raise ContainerSafetyError("XLSX is not a recognized OOXML container")
        package_relationships = _parse_xml(
            raw_members["_rels/.rels"], "XLSX package relationships"
        )
        _check_time(started, limits)
        office_document_targets = {
            element.attrib.get("Target", "").lstrip("/")
            for element in package_relationships
            if _local_name(element.tag) == "Relationship"
            and element.attrib.get("Type", "").lower().endswith(
                "/officedocument"
            )
        }
        if "xl/workbook.xml" not in office_document_targets:
            raise ContainerSafetyError("XLSX is not a recognized OOXML container")

        has_shared_strings = "xl/sharedstrings.xml" in names
        shared_content_types = [
            element.attrib.get("ContentType", "").casefold()
            for element in content_types
            if _local_name(element.tag) == "Override"
            and element.attrib.get("PartName", "").casefold()
            == "/xl/sharedstrings.xml"
        ]
        expected_shared_content_type = (
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sharedstrings+xml"
        )
        workbook_relationships = _parse_xml(
            raw_members["xl/_rels/workbook.xml.rels"],
            "XLSX workbook relationships",
        )
        shared_relationships = [
            element
            for element in workbook_relationships
            if _local_name(element.tag) == "Relationship"
            and element.attrib.get("Type", "").casefold().endswith(
                "/sharedstrings"
            )
        ]
        if has_shared_strings:
            if (
                shared_content_types != [expected_shared_content_type]
                or len(shared_relationships) != 1
            ):
                raise ContainerSafetyError(
                    "shared string package binding is invalid"
                )
            shared_target = shared_relationships[0].attrib.get("Target", "")
            if shared_target.startswith("/"):
                normalized_shared_target = shared_target.lstrip("/").casefold()
            else:
                normalized_shared_target = posixpath.normpath(
                    posixpath.join("xl", shared_target)
                ).casefold()
            if normalized_shared_target != "xl/sharedstrings.xml":
                raise ContainerSafetyError(
                    "shared string package binding is invalid"
                )
        elif shared_content_types or shared_relationships:
            raise ContainerSafetyError(
                "shared string package binding is invalid"
            )

        worksheet_bounds: dict[str, _WorksheetBounds] = {}
        shared_string_table: _SharedStringTable | None = None
        for name, member_value in raw_members.items():
            lower_name = name.lower()
            if lower_name.startswith(
                ("xl/embeddings/", "xl/activex/", "xl/ctrlprops/")
            ) or Path(lower_name).suffix in _EXECUTABLE_SUFFIXES:
                raise ContainerSafetyError("embedded executable content is prohibited")
            if lower_name.endswith((".xml", ".rels")):
                _check_time(started, limits)
                root = _parse_xml(member_value, "XLSX package")
                _check_time(started, limits)
                if _xlsx_member_is_allowed(name):
                    metadata.extend(
                        _inventory_xml_metadata(
                            member_name=name,
                            root=root,
                            namespaces=_xml_namespace_declarations(
                                member_value,
                                purpose="XLSX package",
                            ),
                            limits=limits,
                            budget=metadata_budget,
                            raw_value_budget=raw_value_budget,
                            raw_values=raw_values,
                            started=started,
                        )
                    )
                    if name.startswith("xl/worksheets/sheet"):
                        worksheet_bounds[name] = _worksheet_bounds(root)
                    elif name == "xl/sharedstrings.xml":
                        shared_string_table = _parse_shared_string_table(
                            root,
                            limits=limits,
                            budget=logical_value_budget,
                            logical_values=logical_values,
                        )
                for element in root.iter():
                    local = _local_name(element.tag)
                    if (
                        local == "Relationship"
                        and (
                            element.attrib.get("TargetMode", "").lower()
                            == "external"
                            or "external" in element.attrib.get("Type", "").lower()
                        )
                    ):
                        raise ContainerSafetyError(
                            "external relationship is prohibited"
                        )
                    if local == "f":
                        formula = (element.text or "").strip()
                        if re.search(
                            r"(?i)(?:cmd|powershell|mshta|rundll32|dde)\s*\|",
                            formula,
                        ) or re.search(r"(?i)\|[^!]*!", formula):
                            raise ContainerSafetyError("DDE formula is prohibited")
                        raise ContainerSafetyError("formula content is prohibited")
                    if (
                        local == "row"
                        and element.attrib.get("hidden", "").lower()
                        in _TRUTHY_XML
                    ):
                        raise ContainerSafetyError("hidden row is prohibited")
                    if (
                        local == "col"
                        and element.attrib.get("hidden", "").lower()
                        in _TRUTHY_XML
                    ):
                        raise ContainerSafetyError("hidden column is prohibited")

        shared_string_indices = tuple(
            index
            for bounds in worksheet_bounds.values()
            for index in bounds.shared_string_indices
        )
        if shared_string_table is None:
            if shared_string_indices:
                raise ContainerSafetyError(
                    "shared string cells require a shared string table"
                )
        elif (
            shared_string_table.count != len(shared_string_indices)
            or any(
                index >= len(shared_string_table.values)
                for index in shared_string_indices
            )
        ):
            raise ContainerSafetyError(
                "shared string references or count are invalid"
            )

        workbook_root = _parse_xml(
            raw_members["xl/workbook.xml"], "XLSX workbook"
        )
        _check_time(started, limits)
        for element in workbook_root.iter():
            if (
                _local_name(element.tag) == "sheet"
                and element.attrib.get("state", "visible").lower() != "visible"
            ):
                raise ContainerSafetyError("hidden sheet is prohibited")

        for name in names:
            if not _xlsx_member_is_allowed(name):
                raise ContainerSafetyError("non-data XLSX payload is prohibited")

    try:
        workbook = load_workbook(
            BytesIO(value),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise ContainerSafetyError("XLSX data could not be read safely") from exc
    tables: list[str] = []
    all_headers: list[tuple[str, ...]] = []
    cells: list[InventoryCell] = []
    row_count = 0
    try:
        for sheet in workbook.worksheets:
            _check_time(started, limits)
            worksheet_path = getattr(sheet, "_worksheet_path", None)
            try:
                normalized_worksheet_path = _normalized_member_name(
                    worksheet_path
                )
            except ContainerSafetyError as exc:
                raise ContainerSafetyError(
                    "XLSX worksheet path is invalid"
                ) from exc
            bounds = worksheet_bounds.get(normalized_worksheet_path)
            if (
                bounds is None
                or sheet.max_row != bounds.max_row
                or sheet.max_column != bounds.max_column
            ):
                raise ContainerSafetyError(
                    "XLSX parser view diverges from worksheet dimension"
                )
            table = f"{table_prefix}{sheet.title}"
            rows = sheet.iter_rows(values_only=True)
            try:
                raw_headers = next(rows)
            except StopIteration as exc:
                raise ContainerSafetyError(
                    "XLSX data table must declare headers"
                ) from exc
            headers = tuple("" if value is None else str(value) for value in raw_headers)
            _validate_headers(headers)
            tables.append(table)
            all_headers.append(headers)
            for row_number, row in enumerate(rows, start=2):
                _check_time(started, limits)
                row_count += 1
                if row_count > limits.row_count:
                    raise ContainerSafetyError("row limit exceeded")
                padded = tuple(row) + (None,) * (len(headers) - len(row))
                cells.extend(
                    InventoryCell(
                        table=table,
                        row_number=row_number,
                        column_name=header,
                        value="" if cell is None else str(cell),
                    )
                    for header, cell in zip(headers, padded, strict=True)
                )
    finally:
        workbook.close()
    return ContainerInventory(
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        tables=tuple(tables),
        headers=tuple(all_headers),
        cells=tuple(cells),
        row_count=row_count,
        metadata=tuple(metadata),
        raw_values=tuple(raw_values),
        logical_values=tuple(logical_values),
    )


def _combine_zip_inventories(
    inventories: list[ContainerInventory],
    limits: ContainerLimits,
    outer_metadata: tuple[InventoryMetadata, ...],
) -> ContainerInventory:
    row_count = sum(inventory.row_count for inventory in inventories)
    if row_count > limits.row_count:
        raise ContainerSafetyError("row limit exceeded")
    return ContainerInventory(
        media_type="application/zip",
        tables=tuple(
            table for inventory in inventories for table in inventory.tables
        ),
        headers=tuple(
            headers for inventory in inventories for headers in inventory.headers
        ),
        cells=tuple(cell for inventory in inventories for cell in inventory.cells),
        row_count=row_count,
        metadata=(
            outer_metadata
            + tuple(
                item
                for inventory in inventories
                for item in inventory.metadata
            )
        ),
        raw_values=tuple(
            item
            for inventory in inventories
            for item in inventory.raw_values
        ),
        logical_values=tuple(
            item
            for inventory in inventories
            for item in inventory.logical_values
        ),
    )


def _inspect_zip(
    value: bytes,
    *,
    limits: ContainerLimits,
    started: float,
) -> ContainerInventory:
    if limits.recursion_depth < 1:
        raise ContainerSafetyError("archive recursion depth limit exceeded")
    try:
        archive = zipfile.ZipFile(BytesIO(value))
    except (zipfile.BadZipFile, OSError, UnicodeError) as exc:
        raise ContainerSafetyError("malformed ZIP container") from exc
    inventories: list[ContainerInventory] = []
    budget = _ArchiveBudget()
    metadata_budget = _MetadataBudget()
    raw_value_budget = _RawValueBudget()
    logical_value_budget = _LogicalValueBudget()
    with archive:
        members = _validate_archive_inventory(
            archive,
            limits=limits,
            budget=budget,
        )
        _validate_zip_envelope(value, archive, members)
        outer_metadata = _inventory_archive_metadata(
            archive=archive,
            value=value,
            members=members,
            limits=limits,
            budget=metadata_budget,
            source="zip_archive",
        )
        for member, normalized in members:
            _check_time(started, limits)
            if member.is_dir():
                continue
            suffix = Path(normalized).suffix
            if suffix not in {".csv", ".tsv", ".json", ".xlsx"}:
                raise ContainerSafetyError("unsupported member is prohibited")
            try:
                member_value = archive.read(member)
            except (RuntimeError, OSError, zipfile.BadZipFile, UnicodeError) as exc:
                raise ContainerSafetyError("archive member could not be read") from exc
            if _looks_executable(member_value):
                raise ContainerSafetyError(
                    "unexpected executable content is prohibited"
                )
            if member_value.startswith(b"PK") and suffix != ".xlsx":
                raise ContainerSafetyError("arbitrary nested archive is prohibited")
            if suffix == ".csv":
                inventory = _inspect_delimited(
                    member_value,
                    delimiter=",",
                    table=normalized,
                    media_type="text/csv",
                    limits=limits,
                    started=started,
                )
            elif suffix == ".tsv":
                inventory = _inspect_delimited(
                    member_value,
                    delimiter="\t",
                    table=normalized,
                    media_type="text/tab-separated-values",
                    limits=limits,
                    started=started,
                )
            elif suffix == ".json":
                inventory = _inspect_json(
                    member_value,
                    table_hint=normalized,
                    limits=limits,
                    started=started,
                )
                inventory = ContainerInventory(
                    media_type=inventory.media_type,
                    tables=tuple(
                        f"{normalized}:{table}" for table in inventory.tables
                    ),
                    headers=inventory.headers,
                    cells=tuple(
                        InventoryCell(
                            table=f"{normalized}:{cell.table}",
                            row_number=cell.row_number,
                            column_name=cell.column_name,
                            value=cell.value,
                        )
                        for cell in inventory.cells
                    ),
                    row_count=inventory.row_count,
                    metadata=inventory.metadata,
                    raw_values=inventory.raw_values,
                    logical_values=inventory.logical_values,
                )
            else:
                inventory = _inspect_xlsx(
                    member_value,
                    limits=limits,
                    started=started,
                    table_prefix=f"{normalized}:",
                    budget=budget,
                    metadata_budget=metadata_budget,
                    raw_value_budget=raw_value_budget,
                    logical_value_budget=logical_value_budget,
                )
            inventories.append(inventory)
    if not inventories:
        raise ContainerSafetyError("ZIP must contain supported data members")
    return _combine_zip_inventories(inventories, limits, outer_metadata)


def inspect_container(
    snapshot: SourceSnapshot,
    *,
    limits: ContainerLimits = ContainerLimits(),
) -> ContainerInventory:
    _validate_limits(limits)
    started = time.monotonic()
    value = _read_snapshot(snapshot, limits)
    _check_time(started, limits)
    if len(value) > limits.uncompressed_bytes:
        raise ContainerSafetyError("uncompressed byte limit exceeded")
    if _looks_executable(value):
        raise ContainerSafetyError("unexpected executable content is prohibited")
    suffix = snapshot.original_path.suffix.lower()
    if suffix == ".csv":
        if value.startswith(b"PK"):
            raise ContainerSafetyError("container content does not match CSV")
        return _inspect_delimited(
            value,
            delimiter=",",
            table=snapshot.original_path.name,
            media_type="text/csv",
            limits=limits,
            started=started,
        )
    if suffix == ".tsv":
        if value.startswith(b"PK"):
            raise ContainerSafetyError("container content does not match TSV")
        return _inspect_delimited(
            value,
            delimiter="\t",
            table=snapshot.original_path.name,
            media_type="text/tab-separated-values",
            limits=limits,
            started=started,
        )
    if suffix == ".json":
        if value.startswith(b"PK"):
            raise ContainerSafetyError("container content does not match JSON")
        return _inspect_json(
            value,
            table_hint=snapshot.original_path.name,
            limits=limits,
            started=started,
        )
    if suffix == ".xlsx":
        return _inspect_xlsx(value, limits=limits, started=started)
    if suffix == ".zip":
        return _inspect_zip(value, limits=limits, started=started)
    raise ContainerSafetyError("unsupported source container")
