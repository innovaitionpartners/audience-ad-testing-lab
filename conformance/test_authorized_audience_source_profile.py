from __future__ import annotations

import csv
import io
import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
from unittest import mock
import zipfile


ROOT = Path(__file__).resolve().parents[1]
SKILL = ROOT / "skills" / "audience-data-lab"
SCRIPTS = SKILL / "scripts"
FIXTURES = ROOT / "conformance" / "fixtures" / "authorized-audience" / "source-shapes"
sys.path.insert(0, str(SCRIPTS))

from audience_data_lab.authorized_source import (  # noqa: E402
    AUTHORIZED_SOURCE_PROFILE_VERSION,
    MAX_BUNDLE_BYTES,
    MAX_INPUT_FILES,
    ContractError,
    profile_authorized_bundle,
    validate_source_profile,
)
import audience_data_lab.authorized_source as authorized_source  # noqa: E402


class AuthorizedAudienceSourceProfileTests(unittest.TestCase):
    def profile(self, *paths: Path):
        return profile_authorized_bundle(
            list(paths),
            profile_id="fictional-marketplace-cohort",
            profile_version="1.0.0",
            profiled_at="2026-07-23T12:00:00Z",
        )

    def test_profiles_csv_shape_without_copying_cell_values(self):
        profile = self.profile(FIXTURES / "flat-structural.csv")
        self.assertEqual(AUTHORIZED_SOURCE_PROFILE_VERSION, profile["schema_version"])
        self.assertEqual("ready_for_mapping", profile["decision"]["status"])
        self.assertEqual("aggregate_transform", profile["decision"]["allowed_next_route"])
        self.assertEqual(["flat-structural.csv"], [item["display_name"] for item in profile["inputs"]])
        table = profile["tables"][0]
        self.assertEqual("wide", table["shape"])
        self.assertEqual(2, table["row_count"])
        self.assertEqual(3, table["column_count"])
        self.assertEqual(["segment", "respondent_count", "share"], table["field_names"])
        self.assertEqual(["integer"], table["observed_scalar_types"]["respondent_count"])
        self.assertEqual(["identifier_like"], table["sample_safe_value_classes"]["segment"])
        serialized = json.dumps(profile)
        self.assertNotIn("operations_leaders", serialized)
        self.assertNotIn("finance_leaders", serialized)
        self.assertNotIn("420", serialized)

    def test_profiles_declared_nested_json_record_path(self):
        profile = self.profile(FIXTURES / "nested-export.json")
        self.assertEqual("ready_for_mapping", profile["decision"]["status"])
        self.assertEqual(1, len(profile["tables"]))
        table = profile["tables"][0]
        self.assertEqual("export.cohorts", table["record_path"])
        self.assertEqual("nested", table["shape"])
        self.assertEqual(2, table["row_count"])
        self.assertEqual(["cohort_label", "respondent_count", "share"], table["field_names"])

    def test_detects_linked_file_relationship_without_joining_rows(self):
        profile = self.profile(
            FIXTURES / "linked-distributions.csv", FIXTURES / "linked-cohorts.csv"
        )
        self.assertEqual("ready_for_mapping", profile["decision"]["status"])
        self.assertEqual(1, len(profile["relationships"]))
        relationship = profile["relationships"][0]
        self.assertEqual("candidate_shared_field", relationship["kind"])
        self.assertEqual("cohort_id", relationship["field"])
        self.assertEqual(
            ["linked-cohorts.csv", "linked-distributions.csv"], relationship["files"]
        )

    def test_profiles_generated_wide_xlsx_without_formulas_or_links(self):
        from openpyxl import Workbook

        canonical = json.loads((FIXTURES / "canonical-marketplace-cohort.json").read_text())
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "cohort.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Cohort Summary"
            sheet.append(["cohort_id", "sample_size", "share"])
            sheet.append([canonical["cohort_id"], canonical["metrics"]["sample_size"], canonical["metrics"]["share"]])
            workbook.save(source)
            self.assertFalse(any(cell.data_type == "f" for row in sheet.iter_rows() for cell in row))
            self.assertEqual([], workbook._external_links)
            profile = self.profile(source)
        table = profile["tables"][0]
        self.assertEqual("wide", table["shape"])
        self.assertEqual(1, table["row_count"])
        self.assertEqual("Cohort Summary", table["sheet"])
        self.assertEqual([], profile["unresolved"])
        self.assertEqual("ready_for_mapping", profile["decision"]["status"])

    def test_rejects_input_count_and_total_byte_limits_before_inspection(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "small.csv"
            source.write_text("metric\n1\n", encoding="utf-8")
            with self.assertRaisesRegex(ContractError, "at most 100 input files"):
                self.profile(*([source] * (MAX_INPUT_FILES + 1)))
            oversized = Path(directory) / "oversized.csv"
            oversized.write_bytes(b"x" * (MAX_BUNDLE_BYTES + 1))
            with self.assertRaisesRegex(ContractError, "250 MiB"):
                self.profile(oversized)

    def test_streams_valid_nested_json_larger_than_one_mebibyte_without_raw_values(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "large-nested.json"
            with source.open("w", encoding="utf-8") as handle:
                handle.write('{"export":{"cohorts":[')
                for index in range(30_000):
                    if index:
                        handle.write(",")
                    handle.write(
                        '{"cohort_label":"do-not-copy-this-value",'
                        '"respondent_count":7,"share":'
                        + ("null" if index % 4 == 0 else "0.5")
                        + "}"
                    )
                handle.write("]}}")
            self.assertGreater(source.stat().st_size, 1024 * 1024)
            profile = self.profile(source)
        self.assertEqual("ready_for_mapping", profile["decision"]["status"])
        table = profile["tables"][0]
        self.assertEqual("export.cohorts", table["record_path"])
        self.assertEqual(30_000, table["row_count"])
        self.assertEqual(["integer"], table["observed_scalar_types"]["respondent_count"])
        self.assertEqual(0.25, table["null_rates"]["share"])
        self.assertNotIn("do-not-copy-this-value", json.dumps(profile))

    def test_stable_hashes_and_deterministic_input_order(self):
        first = self.profile(
            FIXTURES / "linked-cohorts.csv", FIXTURES / "linked-distributions.csv"
        )
        second = self.profile(
            FIXTURES / "linked-distributions.csv", FIXTURES / "linked-cohorts.csv"
        )
        self.assertEqual(first["bundle_sha256"], second["bundle_sha256"])
        self.assertEqual(first, second)
        self.assertEqual(
            ["linked-cohorts.csv", "linked-distributions.csv"],
            [item["display_name"] for item in first["inputs"]],
        )

    def test_rejects_malformed_encodings_and_unsupported_or_fake_zip_inputs(self):
        with tempfile.TemporaryDirectory() as directory:
            malformed = Path(directory) / "bad.csv"
            malformed.write_bytes(b"metric\n\xff\n")
            profile = self.profile(malformed)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("invalid_utf8", profile["unresolved"][0]["code"])
            legacy = Path(directory) / "legacy.xls"
            legacy.write_bytes(b"not a workbook")
            profile = self.profile(legacy)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("unsupported_format", profile["unresolved"][0]["code"])
            fake = Path(directory) / "fake.xlsx"
            fake.write_bytes(b"PK\x03\x04not-an-xlsx")
            profile = self.profile(fake)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("invalid_xlsx", profile["unresolved"][0]["code"])

    def test_rejects_formula_cells_and_external_links(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as directory:
            formula = Path(directory) / "formula.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["sample_size"])
            sheet.append(["=2+2"])
            workbook.save(formula)
            profile = self.profile(formula)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("formula_cell", profile["unresolved"][0]["code"])

            formula_header = Path(directory) / "formula-header.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["=CONCAT(\"sample\", \"_size\")"])
            sheet.append([4])
            workbook.save(formula_header)
            profile = self.profile(formula_header)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("formula_cell", profile["unresolved"][0]["code"])

            linked = Path(directory) / "linked.xlsx"
            self._write_external_link_workbook(linked)
            profile = self.profile(linked)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("external_links", profile["unresolved"][0]["code"])

    def test_rejects_display_name_collisions_and_routes_person_risk_to_private_aggregation(self):
        with tempfile.TemporaryDirectory() as directory:
            first_dir = Path(directory) / "one"
            second_dir = Path(directory) / "two"
            first_dir.mkdir()
            second_dir.mkdir()
            first = first_dir / "same.csv"
            second = second_dir / "same.csv"
            first.write_text("metric\n1\n", encoding="utf-8")
            second.write_text("metric\n2\n", encoding="utf-8")
            with self.assertRaisesRegex(ContractError, "normalized display-name collision"):
                self.profile(first, second)

            people = Path(directory) / "people.csv"
            people.write_text(
                "email,event_id,transaction_amount\nperson@example.com,evt-1,19.99\n",
                encoding="utf-8",
            )
            profile = self.profile(people)
        self.assertEqual("requires_private_aggregation", profile["decision"]["status"])
        self.assertEqual("private_aggregation", profile["decision"]["allowed_next_route"])
        codes = [item["code"] for item in profile["privacy_risk"]]
        self.assertIn("email", codes)
        self.assertIn("person_level_event_rows", codes)
        serialized = json.dumps(profile)
        self.assertNotIn("person@example.com", serialized)
        self.assertNotIn("evt-1", serialized)
        self.assertNotIn("19.99", serialized)

    def test_recurses_nested_json_scalars_for_privacy_risk_without_copying_values(self):
        with tempfile.TemporaryDirectory() as directory:
            root_source = Path(directory) / "root.json"
            root_source.write_text('{"person":{"email":"person@example.com"},"metric":1}', encoding="utf-8")
            record_source = Path(directory) / "records.json"
            record_source.write_text('{"records":[{"person":{"email":"record@example.com"},"metric":1}]}', encoding="utf-8")
            root_profile = self.profile(root_source)
            record_profile = self.profile(record_source)
        for profile, raw_value in ((root_profile, "person@example.com"), (record_profile, "record@example.com")):
            self.assertEqual("requires_private_aggregation", profile["decision"]["status"])
            self.assertIn("email", [item["code"] for item in profile["privacy_risk"]])
            self.assertNotIn(raw_value, json.dumps(profile))

    def test_rejects_csv_width_mismatches_and_exact_column_bound(self):
        with tempfile.TemporaryDirectory() as directory:
            mismatch = Path(directory) / "mismatch.csv"
            mismatch.write_text("count\n5,person@example.com\n", encoding="utf-8")
            profile = self.profile(mismatch)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("row_width_mismatch", profile["unresolved"][0]["code"])
            too_wide = Path(directory) / "too-wide.csv"
            too_wide.write_text(",".join(f"field_{index}" for index in range(10_001)) + "\n", encoding="utf-8")
            profile = self.profile(too_wide)
        self.assertEqual("rejected", profile["decision"]["status"])
        self.assertEqual("table_column_limit", profile["unresolved"][0]["code"])

    def test_routes_standalone_transaction_rows_to_private_aggregation(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "transactions.csv"
            source.write_text("transaction_id,transaction_amount\ntxn-100,19.99\n", encoding="utf-8")
            profile = self.profile(source)
        self.assertEqual("requires_private_aggregation", profile["decision"]["status"])
        self.assertIn("person_level_event_rows", [item["code"] for item in profile["privacy_risk"]])

    def test_enforces_row_sheet_and_cell_bounds_at_ingestion(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "rows.csv"
            source.write_text("count\n1\n2\n3\n", encoding="utf-8")
            with mock.patch.object(authorized_source, "MAX_TABLE_ROWS", 2):
                profile = self.profile(source)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("table_row_limit", profile["unresolved"][0]["code"])

            sheets = Path(directory) / "sheets.xlsx"
            workbook = Workbook()
            workbook.active.append(["count"])
            workbook.create_sheet("Second").append(["count"])
            workbook.save(sheets)
            with mock.patch.object(authorized_source, "MAX_WORKBOOK_SHEETS", 1):
                profile = self.profile(sheets)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("workbook_sheet_limit", profile["unresolved"][0]["code"])

            cells = Path(directory) / "cells.xlsx"
            workbook = Workbook()
            workbook.active.append(["count", "share"])
            workbook.active.append([1, 0.5])
            workbook.save(cells)
            with mock.patch.object(authorized_source, "MAX_WORKBOOK_CELLS", 3):
                profile = self.profile(cells)
        self.assertEqual("rejected", profile["decision"]["status"])
        self.assertEqual("workbook_cell_limit", profile["unresolved"][0]["code"])

    def test_profile_validator_rejects_unknown_nested_fields_and_invalid_digests(self):
        profile = self.profile(FIXTURES / "flat-structural.csv")
        profile["inputs"][0]["unexpected"] = True
        with self.assertRaisesRegex(ContractError, "source profile.inputs\\[0\\] has unknown fields"):
            validate_source_profile(profile)
        profile = self.profile(FIXTURES / "flat-structural.csv")
        profile["inputs"][0]["sha256"] = "sha256:not-a-digest"
        with self.assertRaisesRegex(ContractError, "inputs\\[0\\].sha256"):
            validate_source_profile(profile)
        profile = self.profile(FIXTURES / "flat-structural.csv")
        profile["tables"][0]["unknown"] = True
        with self.assertRaisesRegex(ContractError, "source profile.tables\\[0\\] has unknown fields"):
            validate_source_profile(profile)

    def test_profile_validator_rejects_unknown_top_level_and_invalid_decision_routes(self):
        profile = self.profile(FIXTURES / "flat-structural.csv")
        profile["unexpected"] = True
        with self.assertRaisesRegex(ContractError, "unknown fields"):
            validate_source_profile(profile)
        profile = self.profile(FIXTURES / "flat-structural.csv")
        profile["decision"]["allowed_next_route"] = "raw_rows"
        with self.assertRaisesRegex(ContractError, "allowed_next_route"):
            validate_source_profile(profile)

    def test_rejects_unsafe_workbook_containers_and_does_not_misclassify_dates_as_phones(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            for name, part, expected in (
                ("ole.xlsx", "xl/embeddings/oleObject1.bin", "embedded_executable"),
                ("activex.xlsx", "xl/activeX/activeX1.bin", "embedded_executable"),
                ("macro-content.xlsx", None, "macro_enabled_workbook"),
            ):
                hostile = directory_path / name
                self._write_hostile_workbook(hostile, part=part, macro_content_type=part is None)
                profile = self.profile(hostile)
                self.assertEqual("rejected", profile["decision"]["status"])
                self.assertEqual(expected, profile["unresolved"][0]["code"])

            encrypted = directory_path / "encrypted.xlsx"
            self._write_hostile_workbook(encrypted)
            self._set_zip_encryption_flag(encrypted)
            profile = self.profile(encrypted)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("encrypted_workbook", profile["unresolved"][0]["code"])

            macro_extension = directory_path / "legacy.xlsm"
            macro_extension.write_bytes(b"not a workbook")
            profile = self.profile(macro_extension)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("unsupported_format", profile["unresolved"][0]["code"])

            dates = directory_path / "dates.csv"
            dates.write_text("purchase_date\n2026-01-01\n", encoding="utf-8")
            profile = self.profile(dates)
        self.assertEqual("ready_for_mapping", profile["decision"]["status"])
        self.assertNotIn("phone", [item["code"] for item in profile["privacy_risk"]])

    def test_common_aggregate_dates_are_not_phones_but_phone_values_still_route_private(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            dates = directory_path / "common-dates.csv"
            dates.write_text(
                "report_date\n07/24/2026\n24-07-2026\n2026/07/24\n",
                encoding="utf-8",
            )
            date_profile = self.profile(dates)
            phones = directory_path / "phones.csv"
            phones.write_text("metric\n+1 212 555 0199\n", encoding="utf-8")
            phone_profile = self.profile(phones)
        self.assertEqual("ready_for_mapping", date_profile["decision"]["status"])
        self.assertNotIn("phone", [item["code"] for item in date_profile["privacy_risk"]])
        self.assertEqual("requires_private_aggregation", phone_profile["decision"]["status"])
        self.assertIn("phone", [item["code"] for item in phone_profile["privacy_risk"]])

    def test_date_safety_requires_the_complete_datetime_value_to_parse(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            mixed = directory_path / "date-prefix-phone-suffix.csv"
            raw_mixed_value = "2026-07-24 2125550199"
            mixed.write_text(f"metric\n{raw_mixed_value}\n", encoding="utf-8")
            mixed_profile = self.profile(mixed)

            valid = directory_path / "valid-datetimes.csv"
            valid.write_text(
                "report_datetime\n"
                "2026-07-24T16:45:30Z\n"
                "2026-07-24 12:45:30-04:00\n",
                encoding="utf-8",
            )
            valid_profile = self.profile(valid)

        self.assertEqual("requires_private_aggregation", mixed_profile["decision"]["status"])
        self.assertIn("phone", [item["code"] for item in mixed_profile["privacy_risk"]])
        self.assertNotIn(raw_mixed_value, json.dumps(mixed_profile))
        self.assertEqual("ready_for_mapping", valid_profile["decision"]["status"])
        self.assertEqual(
            ["date_or_datetime_like"],
            valid_profile["tables"][0]["sample_safe_value_classes"]["report_datetime"],
        )
        self.assertNotIn("phone", [item["code"] for item in valid_profile["privacy_risk"]])

    def test_invalid_or_ambiguous_date_shapes_fall_through_to_phone_privacy_risk(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            invalid = directory_path / "invalid-date.csv"
            invalid.write_text("metric\n12-34-5678\n", encoding="utf-8")
            ambiguous = directory_path / "ambiguous-date.csv"
            ambiguous.write_text("metric\n02-03-2026\n", encoding="utf-8")
            invalid_profile = self.profile(invalid)
            ambiguous_profile = self.profile(ambiguous)
        for profile, raw_value in ((invalid_profile, "12-34-5678"), (ambiguous_profile, "02-03-2026")):
            self.assertEqual("requires_private_aggregation", profile["decision"]["status"])
            self.assertIn("phone", [item["code"] for item in profile["privacy_risk"]])
            self.assertNotIn(raw_value, json.dumps(profile))

    def test_rejects_nested_json_scalar_paths_that_collide_with_literal_dotted_keys(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "collision.json"
            source.write_text(
                '{"network":{"source":{"value":"192.168.1.1"}},'
                '"network.source.value":"safe"}',
                encoding="utf-8",
            )
            profile = self.profile(source)
        self.assertEqual("rejected", profile["decision"]["status"])
        self.assertEqual("nested_path_collision", profile["unresolved"][0]["code"])
        self.assertNotIn("192.168.1.1", json.dumps(profile))

    def test_generic_zip_and_unsupported_extension_write_rejected_profiles_and_exit_five(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            generic_zip = directory_path / "generic.zip"
            with zipfile.ZipFile(generic_zip, "w") as archive:
                archive.writestr("payload.txt", "not a workbook")
            profile = self.profile(generic_zip)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("unknown", profile["inputs"][0]["format"])

            unsupported = directory_path / "unsupported.tsv"
            unsupported.write_text("count\n5\n", encoding="utf-8")
            profile = self.profile(unsupported)
            self.assertEqual("rejected", profile["decision"]["status"])
            self.assertEqual("unknown", profile["inputs"][0]["format"])

            output = directory_path / "generic-profile.json"
            result = subprocess.run([
                sys.executable, str(SCRIPTS / "profile-authorized-audience.py"), str(generic_zip),
                "--profile-id", "fictional-marketplace-cohort", "--profile-version", "1.0.0",
                "--profiled-at", "2026-07-23T12:00:00Z", "--output", str(output),
            ], text=True, capture_output=True, check=False)
            self.assertEqual(5, result.returncode, result.stderr)
            self.assertTrue(output.exists())
            self.assertEqual("rejected", json.loads(output.read_text())["decision"]["status"])

    def test_cli_preserves_existing_outputs_and_returns_intentional_privacy_route(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            source = directory_path / "people.csv"
            source.write_text("email\nperson@example.com\n", encoding="utf-8")
            output = directory_path / "profile.json"
            command = [
                sys.executable,
                str(SCRIPTS / "profile-authorized-audience.py"),
                str(source),
                "--profile-id", "fictional-marketplace-cohort",
                "--profile-version", "1.0.0",
                "--profiled-at", "2026-07-23T12:00:00Z",
                "--output", str(output),
            ]
            result = subprocess.run(command, text=True, capture_output=True, check=False)
            self.assertEqual(4, result.returncode, result.stderr)
            self.assertTrue(output.exists())
            self.assertIn(str(output), result.stdout)
            self.assertIn("requires_private_aggregation", result.stdout)
            second = subprocess.run(command, text=True, capture_output=True, check=False)
            self.assertEqual(2, second.returncode)
            self.assertIn("already exists", second.stderr)

    def test_cli_returns_zero_for_safe_profile_and_five_for_malformed_source_content(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            safe = directory_path / "safe.csv"
            safe.write_text("count\n5\n", encoding="utf-8")
            malformed = directory_path / "blank-header.csv"
            malformed.write_text(",count\nsegment,5\n", encoding="utf-8")
            base = [
                sys.executable, str(SCRIPTS / "profile-authorized-audience.py"),
                "--profile-id", "fictional-marketplace-cohort", "--profile-version", "1.0.0",
                "--profiled-at", "2026-07-23T12:00:00Z",
            ]
            safe_output = directory_path / "safe-profile.json"
            safe_result = subprocess.run([*base, str(safe), "--output", str(safe_output)], text=True, capture_output=True, check=False)
            self.assertEqual(0, safe_result.returncode, safe_result.stderr)
            self.assertTrue(safe_output.exists())
            rejected_output = directory_path / "rejected-profile.json"
            rejected_result = subprocess.run([*base, str(malformed), "--output", str(rejected_output)], text=True, capture_output=True, check=False)
            self.assertEqual(5, rejected_result.returncode, rejected_result.stderr)
            self.assertTrue(rejected_output.exists())
            self.assertEqual("rejected", json.loads(rejected_output.read_text())["decision"]["status"])

    @staticmethod
    def _write_external_link_workbook(path: Path) -> None:
        with zipfile.ZipFile(path, "w") as archive:
            archive.writestr("[Content_Types].xml", """<?xml version=\"1.0\"?><Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\"><Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/><Default Extension=\"xml\" ContentType=\"application/xml\"/><Override PartName=\"/xl/workbook.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml\"/><Override PartName=\"/xl/worksheets/sheet1.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml\"/></Types>""")
            archive.writestr("_rels/.rels", """<?xml version=\"1.0\"?><Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\"><Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" Target=\"xl/workbook.xml\"/></Relationships>""")
            archive.writestr("xl/workbook.xml", """<?xml version=\"1.0\"?><workbook xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\" xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\"><sheets><sheet name=\"Sheet1\" sheetId=\"1\" r:id=\"rId1\"/></sheets></workbook>""")
            archive.writestr("xl/_rels/workbook.xml.rels", """<?xml version=\"1.0\"?><Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\"><Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet\" Target=\"worksheets/sheet1.xml\"/><Relationship Id=\"rId2\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink\" Target=\"externalLinks/externalLink1.xml\"/></Relationships>""")
            archive.writestr("xl/worksheets/sheet1.xml", """<?xml version=\"1.0\"?><worksheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\"><sheetData><row r=\"1\"><c r=\"A1\" t=\"inlineStr\"><is><t>metric</t></is></c></row><row r=\"2\"><c r=\"A2\"><v>1</v></c></row></sheetData></worksheet>""")
            archive.writestr("xl/externalLinks/externalLink1.xml", """<?xml version=\"1.0\"?><externalLink xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\"/>""")

    @staticmethod
    def _write_hostile_workbook(path: Path, *, part: str | None = None, macro_content_type: bool = False) -> None:
        content_type = "application/vnd.ms-excel.sheet.macroEnabled.main+xml" if macro_content_type else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
        with zipfile.ZipFile(path, "w") as archive:
            archive.writestr("[Content_Types].xml", f"""<?xml version=\"1.0\"?><Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\"><Override PartName=\"/xl/workbook.xml\" ContentType=\"{content_type}\"/></Types>""")
            archive.writestr("xl/workbook.xml", "<workbook/>")
            if part:
                archive.writestr(part, b"hostile")

    @staticmethod
    def _set_zip_encryption_flag(path: Path) -> None:
        payload = bytearray(path.read_bytes())
        for signature, offset in ((b"PK\x03\x04", 6), (b"PK\x01\x02", 8)):
            start = 0
            while True:
                index = payload.find(signature, start)
                if index < 0:
                    break
                payload[index + offset] |= 0x01
                start = index + 4
        path.write_bytes(payload)


if __name__ == "__main__":
    unittest.main()
