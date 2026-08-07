from __future__ import annotations

from copy import copy
from dataclasses import fields
from io import BytesIO
import json
from pathlib import Path
import stat
import struct
import sys
import tempfile
import unittest
import zipfile
import zlib

from openpyxl import Workbook
from openpyxl.packaging.custom import StringProperty
from openpyxl.worksheet.table import Table


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "skills" / "real-world-outcome-data-prep" / "scripts"
sys.path.insert(0, str(SCRIPTS))

from outcome_data_prep.container_safety import (  # noqa: E402
    ContainerInventory,
    ContainerLimits,
    ContainerSafetyError,
    InventoryCell,
    InventoryMetadata,
    inspect_container,
)
import outcome_data_prep.privacy as privacy_module  # noqa: E402
from outcome_data_prep.privacy import (  # noqa: E402
    AdmittedSource,
    PrivacyAdmissionError,
    PrivacyDecision,
    admit_source,
    pre_scan_obvious_privacy,
)
from outcome_data_prep.source_snapshot import snapshot_source  # noqa: E402


class SourceSafetyTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporary_directory.cleanup)
        self.base = Path(self.temporary_directory.name)
        self.stage = self.base / "stage"
        self.durable = self.base / "study" / "source.bin"

    def write_bytes(self, name: str, value: bytes) -> Path:
        path = self.base / name
        path.write_bytes(value)
        return path

    def snapshot(self, path: Path):
        return snapshot_source(path, staging_root=self.stage)

    @staticmethod
    def adapter_validation(snapshot, inventory):
        return privacy_module.AdapterAdmissionValidation(
            adapter_id="registered-adapter",
            adapter_version="1.0.0",
            source_sha256=snapshot.source_sha256,
            inventory_sha256=privacy_module.container_inventory_sha256(inventory),
            profile_sha256="sha256:" + ("0" * 64),
            adapter_validation_sha256="sha256:" + ("1" * 64),
            governance_sha256="sha256:" + ("2" * 64),
            accepted=True,
            observed_minimum_group_size=10,
            errors=(),
        )

    def csv_inventory(
        self, *, headers: list[str], rows: list[list[str]]
    ) -> ContainerInventory:
        cells = tuple(
            InventoryCell(
                table="source.csv",
                row_number=row_number,
                column_name=header,
                value=value,
            )
            for row_number, row in enumerate(rows, start=2)
            for header, value in zip(headers, row, strict=True)
        )
        return ContainerInventory(
            media_type="text/csv",
            tables=("source.csv",),
            headers=(tuple(headers),),
            cells=cells,
            row_count=len(rows),
        )

    @staticmethod
    def workbook_bytes(
        *,
        hidden_sheet: bool = False,
        hidden_row: bool = False,
        hidden_column: bool = False,
        formula: str | None = None,
        creator: str | None = None,
        last_modified_by: str | None = None,
        custom_property: tuple[str, str] | None = None,
        sheet_title: str = "data",
        table_name: str | None = None,
    ) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title
        sheet.append(["campaign_id", "impressions"])
        sheet.append(["campaign-1", 100])
        if creator is not None:
            workbook.properties.creator = creator
        if last_modified_by is not None:
            workbook.properties.lastModifiedBy = last_modified_by
        if custom_property is not None:
            workbook.custom_doc_props.append(
                StringProperty(
                    name=custom_property[0],
                    value=custom_property[1],
                )
            )
        if table_name is not None:
            sheet.add_table(Table(displayName=table_name, ref="A1:B2"))
        if hidden_sheet:
            hidden = workbook.create_sheet("private")
            hidden.append(["email"])
            hidden.append(["person@example.com"])
            hidden.sheet_state = "hidden"
        if hidden_row:
            sheet.row_dimensions[2].hidden = True
        if hidden_column:
            sheet.column_dimensions["B"].hidden = True
        if formula is not None:
            sheet["B2"] = formula
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def zip_bytes(members: list[tuple[str, bytes]]) -> bytes:
        output = BytesIO()
        with zipfile.ZipFile(
            output, "w", compression=zipfile.ZIP_DEFLATED
        ) as archive:
            for name, value in members:
                archive.writestr(name, value)
        return output.getvalue()

    @staticmethod
    def replace_zip_member(value: bytes, name: str, replacement: bytes) -> bytes:
        output = BytesIO()
        with zipfile.ZipFile(BytesIO(value)) as source_archive:
            with zipfile.ZipFile(output, "w") as target_archive:
                for member in source_archive.infolist():
                    target_archive.writestr(
                        member,
                        replacement
                        if member.filename == name
                        else source_archive.read(member),
                    )
        return output.getvalue()

    @staticmethod
    def rename_zip_member(value: bytes, old_name: str, new_name: str) -> bytes:
        output = BytesIO()
        found = False
        with zipfile.ZipFile(BytesIO(value)) as source_archive:
            with zipfile.ZipFile(output, "w") as target_archive:
                for source_info in source_archive.infolist():
                    target_info = copy(source_info)
                    if source_info.filename == old_name:
                        target_info.filename = new_name
                        target_info.orig_filename = new_name
                        found = True
                    target_archive.writestr(
                        target_info,
                        source_archive.read(source_info),
                    )
        if not found:
            raise AssertionError(f"ZIP member was not found: {old_name}")
        return output.getvalue()

    @staticmethod
    def replace_archive_name_bytes(
        value: bytes,
        old_name: bytes,
        new_name: bytes,
    ) -> bytes:
        if len(old_name) != len(new_name):
            raise AssertionError("archive-name replacement must preserve length")
        if value.count(old_name) != 2:
            raise AssertionError("expected one local and one central ZIP name")
        return value.replace(old_name, new_name)

    @staticmethod
    def insert_zip_central_slack(value: bytes, slack: bytes) -> bytes:
        if not slack:
            raise AssertionError("ZIP slack must not be empty")
        eocd = value.rfind(b"PK\x05\x06")
        if eocd < 0:
            raise AssertionError("ZIP EOCD was not found")
        central_offset = int.from_bytes(value[eocd + 16:eocd + 20], "little")
        result = bytearray(
            value[:central_offset] + slack + value[central_offset:]
        )
        shifted_eocd = eocd + len(slack)
        result[shifted_eocd + 16:shifted_eocd + 20] = (
            central_offset + len(slack)
        ).to_bytes(4, "little")
        return bytes(result)

    @staticmethod
    def zip_bytes_with_metadata(
        *,
        member_name: str,
        member_value: bytes,
        archive_comment: bytes = b"",
        member_comment: bytes = b"",
        member_extra: bytes = b"",
    ) -> bytes:
        output = BytesIO()
        with zipfile.ZipFile(output, "w") as archive:
            info = zipfile.ZipInfo(member_name)
            info.comment = member_comment
            info.extra = member_extra
            archive.writestr(info, member_value)
            archive.comment = archive_comment
        return output.getvalue()

    @staticmethod
    def workbook_with_archive_metadata(
        value: bytes,
        *,
        archive_comment: bytes = b"",
        member_name: str = "docProps/core.xml",
        member_comment: bytes = b"",
        member_extra: bytes = b"",
    ) -> bytes:
        output = BytesIO()
        with zipfile.ZipFile(BytesIO(value)) as source_archive:
            with zipfile.ZipFile(output, "w") as target_archive:
                target_archive.comment = archive_comment
                for source_info in source_archive.infolist():
                    target_info = copy(source_info)
                    if source_info.filename == member_name:
                        target_info.comment = member_comment
                        target_info.extra = member_extra
                    target_archive.writestr(
                        target_info,
                        source_archive.read(source_info),
                    )
        return output.getvalue()

    @staticmethod
    def workbook_with_shared_strings(
        value: bytes,
        *,
        shared_strings_xml: bytes,
        worksheet_xml: bytes | None = None,
    ) -> bytes:
        output = BytesIO()
        with zipfile.ZipFile(BytesIO(value)) as source_archive:
            content_types = source_archive.read("[Content_Types].xml").replace(
                b"</Types>",
                (
                    b'<Override PartName="/xl/sharedStrings.xml" '
                    b'ContentType="application/vnd.openxmlformats-'
                    b'officedocument.spreadsheetml.sharedStrings+xml" />'
                    b"</Types>"
                ),
                1,
            )
            workbook_relationships = source_archive.read(
                "xl/_rels/workbook.xml.rels"
            ).replace(
                b"</Relationships>",
                (
                    b'<Relationship Type="http://schemas.openxmlformats.org/'
                    b'officeDocument/2006/relationships/sharedStrings" '
                    b'Target="sharedStrings.xml" Id="rId4" />'
                    b"</Relationships>"
                ),
                1,
            )
            with zipfile.ZipFile(output, "w") as target_archive:
                for member in source_archive.infolist():
                    replacement = None
                    if member.filename == "[Content_Types].xml":
                        replacement = content_types
                    elif member.filename == "xl/_rels/workbook.xml.rels":
                        replacement = workbook_relationships
                    elif (
                        member.filename == "xl/worksheets/sheet1.xml"
                        and worksheet_xml is not None
                    ):
                        replacement = worksheet_xml
                    target_archive.writestr(
                        member,
                        replacement
                        if replacement is not None
                        else source_archive.read(member),
                    )
                target_archive.writestr(
                    "xl/sharedStrings.xml",
                    shared_strings_xml,
                )
        return output.getvalue()

    @staticmethod
    def shared_strings_xml(
        items: bytes,
        *,
        count: int,
        unique_count: int,
    ) -> bytes:
        return (
            b'<sst xmlns="http://schemas.openxmlformats.org/'
            b'spreadsheetml/2006/main" count="'
            + str(count).encode()
            + b'" uniqueCount="'
            + str(unique_count).encode()
            + b'">'
            + items
            + b"</sst>"
        )

    @staticmethod
    def unicode_extra_field(header_id: int, raw: bytes, text: bytes) -> bytes:
        payload = (
            b"\x01"
            + (zlib.crc32(raw) & 0xFFFFFFFF).to_bytes(4, "little")
            + text
        )
        return struct.pack("<HH", header_id, len(payload)) + payload

    @staticmethod
    def mutate_first_local_extra_byte(value: bytes) -> bytes:
        result = bytearray(value)
        with zipfile.ZipFile(BytesIO(value)) as archive:
            member = archive.infolist()[0]
        header = struct.Struct("<4s5H3L2H")
        fields = header.unpack_from(value, member.header_offset)
        filename_length = fields[-2]
        extra_length = fields[-1]
        if extra_length == 0:
            raise AssertionError("fixture must contain a local extra field")
        extra_start = member.header_offset + header.size + filename_length
        result[extra_start + extra_length - 1] ^= 0x01
        return bytes(result)

    def assert_rejected_before_metadata_admission(
        self,
        *,
        name: str,
        source_bytes: bytes,
        expected_category: str | None = None,
    ) -> None:
        source = self.write_bytes(name, source_bytes)
        snapshot = self.snapshot(source)
        try:
            inventory = inspect_container(snapshot)
        except ContainerSafetyError:
            return
        decision = pre_scan_obvious_privacy(inventory)
        destination = self.base / "unsafe-metadata-admission" / name
        if decision.status == "pre_scan_clear":
            admitted = admit_source(
                snapshot,
                inventory,
                decision,
                self.adapter_validation(snapshot, inventory),
                destination,
            )
            self.assertEqual(source_bytes, admitted.source_path.read_bytes())
            self.fail("unscanned container metadata reached durable admission")
        self.assertEqual("blocked_person_level", decision.status)
        if expected_category is not None:
            self.assertIn(expected_category, decision.blocked_categories)
        with self.assertRaisesRegex(PrivacyAdmissionError, "privacy"):
            admit_source(
                snapshot,
                inventory,
                decision,
                self.adapter_validation(snapshot, inventory),
                destination,
            )
        self.assertFalse(destination.exists())

    def test_public_dataclasses_are_closed_and_frozen(self):
        self.assertEqual(
            ContainerLimits(
                compressed_bytes=50_000_000,
                uncompressed_bytes=250_000_000,
                member_count=128,
                recursion_depth=1,
                expansion_ratio=20.0,
                row_count=2_000_000,
                metadata_count=100_000,
                metadata_value_chars=4_096,
                metadata_total_chars=2_000_000,
                raw_value_count=2_000_000,
                raw_value_chars=1_000_000,
                raw_value_total_chars=250_000_000,
                logical_value_count=1_000_000,
                logical_value_chars=1_000_000,
                logical_value_total_chars=250_000_000,
                xml_bytes=250_000_000,
                processing_seconds=30.0,
            ),
            ContainerLimits(),
        )
        self.assertEqual(
            ["table", "row_number", "column_name", "value"],
            [field.name for field in fields(InventoryCell)],
        )
        self.assertEqual(
            ["source", "name", "value"],
            [field.name for field in fields(InventoryMetadata)],
        )
        self.assertEqual(
            [
                "media_type",
                "tables",
                "headers",
                "cells",
                "row_count",
                "metadata",
                "raw_values",
                "logical_values",
            ],
            [field.name for field in fields(ContainerInventory)],
        )
        self.assertEqual(
            [
                "status",
                "observed_minimum_group_size",
                "blocked_categories",
            ],
            [field.name for field in fields(PrivacyDecision)],
        )
        self.assertEqual(
            [
                "source_path",
                "source_sha256",
                "byte_length",
                "source_name",
                "snapshot_sha256",
                "inventory_sha256",
                "pre_scan_sha256",
                "adapter_validation_sha256",
                "admission_sha256",
            ],
            [field.name for field in fields(AdmittedSource)],
        )
        self.assertEqual(
            [
                "adapter_id",
                "adapter_version",
                "source_sha256",
                "inventory_sha256",
                "profile_sha256",
                "adapter_validation_sha256",
                "governance_sha256",
                "accepted",
                "observed_minimum_group_size",
                "errors",
            ],
            [
                field.name
                for field in fields(
                    privacy_module.AdapterAdmissionValidation
                )
            ],
        )
        self.assertTrue(InventoryCell.__dataclass_params__.frozen)
        self.assertTrue(InventoryMetadata.__dataclass_params__.frozen)
        self.assertTrue(ContainerInventory.__dataclass_params__.frozen)
        self.assertTrue(PrivacyDecision.__dataclass_params__.frozen)
        self.assertTrue(AdmittedSource.__dataclass_params__.frozen)
        self.assertTrue(
            privacy_module.AdapterAdmissionValidation.__dataclass_params__.frozen
        )

    def test_csv_preserves_raw_headers_strings_and_last_row(self):
        source = self.write_bytes(
            "report.csv",
            b" Campaign ID ,impressions\n00123,500\n00999,700\n",
        )
        inventory = inspect_container(self.snapshot(source))
        self.assertEqual("text/csv", inventory.media_type)
        self.assertEqual(((" Campaign ID ", "impressions"),), inventory.headers)
        self.assertEqual(2, inventory.row_count)
        self.assertEqual("00999", inventory.cells[-2].value)

    def test_csv_rejects_duplicate_headers_nul_mixed_width_and_invalid_utf8(self):
        cases = {
            "duplicate.csv": (b"a,a\n1,2\n", "duplicate header"),
            "nul.csv": (b"a,b\n1,\x00\n", "NUL"),
            "mixed.csv": (b"a,b\n1\n", "row width"),
            "invalid.csv": (b"a,b\n1,\xff\n", "encoding"),
        }
        for name, (value, message) in cases.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(ContainerSafetyError, message):
                    inspect_container(self.snapshot(self.write_bytes(name, value)))

    def test_csv_row_limit_is_enforced(self):
        source = self.write_bytes("rows.csv", b"a\n1\n2\n")
        with self.assertRaisesRegex(ContainerSafetyError, "row limit"):
            inspect_container(
                self.snapshot(source),
                limits=ContainerLimits(row_count=1),
            )

    def test_tsv_uses_a_closed_tabular_schema(self):
        source = self.write_bytes(
            "report.tsv",
            b"campaign_id\timpressions\ncampaign-1\t500\n",
        )
        inventory = inspect_container(self.snapshot(source))
        self.assertEqual("text/tab-separated-values", inventory.media_type)
        self.assertEqual(("campaign_id", "impressions"), inventory.headers[0])
        self.assertEqual(("campaign-1", "500"), tuple(
            cell.value for cell in inventory.cells
        ))

    def test_json_accepts_only_one_closed_row_array_and_preserves_strings(self):
        source = self.write_bytes(
            "rows.json",
            b'{"rows":[{"campaign_id":"00123","impressions":500}]}',
        )
        inventory = inspect_container(self.snapshot(source))
        self.assertEqual("application/json", inventory.media_type)
        self.assertEqual(("rows",), inventory.tables)
        self.assertEqual("00123", inventory.cells[0].value)
        self.assertEqual("500", inventory.cells[1].value)

    def test_json_preserves_bounded_nested_cells_and_scans_nested_privacy(self):
        safe = self.write_bytes(
            "nested-safe.json",
            (
                b'{"rows":[{"campaign":{"id":"campaign-1"},'
                b'"actions":[{"action_type":"purchase","value":"2"}]}]}'
            ),
        )
        inventory = inspect_container(self.snapshot(safe))
        self.assertEqual(
            '{"id":"campaign-1"}',
            inventory.cells[0].value,
        )
        self.assertEqual("pre_scan_clear", pre_scan_obvious_privacy(inventory).status)

        cases = {
            "nested-person.json": (
                b'{"rows":[{"campaign":{"email":"person@example.com"}}]}',
                "email",
            ),
            "nested-secret.json": (
                b'{"rows":[{"campaign":{"api_key":"configured"}}]}',
                "secret_header",
            ),
        }
        for name, (value, category) in cases.items():
            with self.subTest(name=name):
                decision = pre_scan_obvious_privacy(
                    inspect_container(self.snapshot(self.write_bytes(name, value)))
                )
                self.assertEqual("blocked_person_level", decision.status)
                self.assertIn(category, decision.blocked_categories)

    def test_json_rejects_mixed_nested_cell_shapes(self):
        source = self.write_bytes(
            "mixed-nested.json",
            b'{"rows":[{"campaign":{"id":"one"}},{"campaign":"two"}]}',
        )
        with self.assertRaisesRegex(ContainerSafetyError, "mixed nested"):
            inspect_container(self.snapshot(source))

    def test_json_rejects_duplicate_keys_nonfinite_nesting_and_open_shapes(self):
        cases = {
            "duplicate.json": (
                b'{"rows":[{"campaign":"a","campaign":"b"}]}',
                "duplicate key",
            ),
            "nonfinite.json": (b'[{"value":NaN}]', "nonfinite"),
            "nested.json": (
                (
                    '{"rows":['
                    + "[" * 70
                    + '"value"'
                    + "]" * 70
                    + "]}"
                ).encode(),
                "nesting",
            ),
            "open.json": (
                b'{"rows":[],"metadata":{}}',
                "one declared row array",
            ),
        }
        for name, (value, message) in cases.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(ContainerSafetyError, message):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, value)),
                        limits=ContainerLimits(),
                    )

    def test_zip_expansion_ratio_is_bounded(self):
        source = self.write_bytes(
            "bomb.zip",
            self.zip_bytes([("report.csv", b"header\n" + b"A" * 50_000)]),
        )
        with self.assertRaisesRegex(ContainerSafetyError, "expansion ratio"):
            inspect_container(self.snapshot(source))

    def test_zip_rejects_traversal_duplicate_nested_executable_and_encrypted(self):
        nested = self.zip_bytes([("inner.csv", b"a\n1\n")])
        cases = {
            "traversal.zip": (
                [("../private.csv", b"a\n1\n")],
                "unsafe member path",
            ),
            "absolute.zip": (
                [("/private.csv", b"a\n1\n")],
                "unsafe member path",
            ),
            "duplicate.zip": (
                [("reports/a.csv", b"a\n1\n"), ("reports/./a.csv", b"a\n2\n")],
                "duplicate normalized member",
            ),
            "nested.zip": ([("nested.zip", nested)], "unsupported member"),
            "executable.zip": (
                [("payload.exe", b"MZ executable")],
                "unsupported member",
            ),
            "disguised-executable.zip": (
                [("report.csv", b"MZ executable")],
                "executable content",
            ),
        }
        for name, (members, message) in cases.items():
            with self.subTest(name=name):
                source = self.write_bytes(name, self.zip_bytes(members))
                with self.assertRaisesRegex(ContainerSafetyError, message):
                    inspect_container(self.snapshot(source))

        output = BytesIO()
        with zipfile.ZipFile(output, "w") as archive:
            info = zipfile.ZipInfo("report.csv")
            info.flag_bits |= 0x1
            archive.writestr(info, b"a\n1\n")
        encrypted = bytearray(output.getvalue())
        encrypted[6] |= 0x1
        central = encrypted.find(b"PK\x01\x02")
        encrypted[central + 8] |= 0x1
        with self.assertRaisesRegex(ContainerSafetyError, "encrypted"):
            inspect_container(
                self.snapshot(self.write_bytes("encrypted.zip", bytes(encrypted)))
            )

    def test_zip_rejects_symlinks_and_member_limit(self):
        output = BytesIO()
        with zipfile.ZipFile(output, "w") as archive:
            info = zipfile.ZipInfo("report.csv")
            info.create_system = 3
            info.external_attr = (stat.S_IFLNK | 0o777) << 16
            archive.writestr(info, b"target")
        with self.assertRaisesRegex(ContainerSafetyError, "symlink"):
            inspect_container(
                self.snapshot(self.write_bytes("symlink.zip", output.getvalue()))
            )

        source = self.write_bytes(
            "members.zip",
            self.zip_bytes(
                [("one.csv", b"a\n1\n"), ("two.csv", b"a\n2\n")]
            ),
        )
        with self.assertRaisesRegex(ContainerSafetyError, "member count"):
            inspect_container(
                self.snapshot(source), limits=ContainerLimits(member_count=1)
            )
        with self.assertRaisesRegex(ContainerSafetyError, "recursion depth"):
            inspect_container(
                self.snapshot(source), limits=ContainerLimits(recursion_depth=0)
            )

    def test_zip_can_inventory_supported_data_members_including_xlsx(self):
        source = self.write_bytes(
            "reports.zip",
            self.zip_bytes(
                [
                    ("one.csv", b"campaign_id,impressions\none,10\n"),
                    ("two.xlsx", self.workbook_bytes()),
                ]
            ),
        )
        inventory = inspect_container(self.snapshot(source))
        self.assertEqual("application/zip", inventory.media_type)
        self.assertEqual(2, inventory.row_count)
        self.assertEqual(
            ("one.csv", "two.xlsx:data"),
            inventory.tables,
        )

    def test_zip_xlsx_descendants_share_one_cumulative_member_budget(self):
        workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            workbook_members = len(archive.infolist())
        source = self.write_bytes(
            "workbooks.zip",
            self.zip_bytes(
                [("one.xlsx", workbook), ("two.xlsx", workbook)]
            ),
        )
        limit = 2 + workbook_members
        with self.assertRaisesRegex(ContainerSafetyError, "member count"):
            inspect_container(
                self.snapshot(source),
                limits=ContainerLimits(member_count=limit),
            )

    def test_zip_xlsx_descendants_share_one_cumulative_uncompressed_budget(self):
        workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            workbook_uncompressed = sum(
                member.file_size for member in archive.infolist()
            )
        outer_uncompressed = len(workbook) * 2
        source = self.write_bytes(
            "workbooks.zip",
            self.zip_bytes(
                [("one.xlsx", workbook), ("two.xlsx", workbook)]
            ),
        )
        limit = max(outer_uncompressed, workbook_uncompressed) + 1
        self.assertLess(
            limit,
            outer_uncompressed + (2 * workbook_uncompressed),
        )
        with self.assertRaisesRegex(ContainerSafetyError, "uncompressed byte"):
            inspect_container(
                self.snapshot(source),
                limits=ContainerLimits(uncompressed_bytes=limit),
            )

    def test_hidden_xlsx_sheet_row_and_column_are_rejected(self):
        cases = {
            "hidden-sheet.xlsx": (
                self.workbook_bytes(hidden_sheet=True),
                "hidden sheet",
            ),
            "hidden-row.xlsx": (
                self.workbook_bytes(hidden_row=True),
                "hidden row",
            ),
            "hidden-column.xlsx": (
                self.workbook_bytes(hidden_column=True),
                "hidden column",
            ),
        }
        for name, (value, message) in cases.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(ContainerSafetyError, message):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, value))
                    )

    def test_xlsx_rejects_encryption_and_duplicate_headers(self):
        encrypted = self.write_bytes(
            "encrypted.xlsx",
            b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 128,
        )
        with self.assertRaisesRegex(ContainerSafetyError, "encrypted workbook"):
            inspect_container(self.snapshot(encrypted))

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["campaign_id", "campaign_id"])
        sheet.append(["one", "two"])
        output = BytesIO()
        workbook.save(output)
        duplicate = self.write_bytes("duplicate.xlsx", output.getvalue())
        with self.assertRaisesRegex(ContainerSafetyError, "duplicate header"):
            inspect_container(self.snapshot(duplicate))

    def test_xlsx_metadata_inventory_has_independent_resource_limits(self):
        source = self.write_bytes("bounded.xlsx", self.workbook_bytes())
        cases = {
            "count": (
                ContainerLimits(metadata_count=0),
                "metadata count",
            ),
            "per-value": (
                ContainerLimits(metadata_value_chars=1),
                "metadata value length",
            ),
            "total-characters": (
                ContainerLimits(metadata_total_chars=1),
                "metadata character",
            ),
            "xml-bytes": (
                ContainerLimits(xml_bytes=1),
                "XML byte",
            ),
        }
        for name, (limits, message) in cases.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(ContainerSafetyError, message):
                    inspect_container(self.snapshot(source), limits=limits)

    def test_xlsx_rejects_unbounded_comment_and_instruction_metadata(self):
        base = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(base)) as archive:
            core = archive.read("docProps/core.xml")
        cases = {
            "comment.xlsx": core.replace(
                b"</cp:coreProperties>",
                b"<!-- owner@example.com --></cp:coreProperties>",
            ),
            "instruction.xlsx": core.replace(
                b"</cp:coreProperties>",
                b"<?owner person@example.com?></cp:coreProperties>",
            ),
        }
        for name, replacement in cases.items():
            with self.subTest(name=name):
                workbook_bytes = self.replace_zip_member(
                    base,
                    "docProps/core.xml",
                    replacement,
                )
                with self.assertRaisesRegex(
                    ContainerSafetyError, "unsupported XML metadata"
                ):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, workbook_bytes))
                    )

    def test_xlsx_rejects_formulas_dde_macros_external_links_and_payloads(self):
        formula = self.write_bytes(
            "formula.xlsx", self.workbook_bytes(formula="=SUM(1,2)")
        )
        with self.assertRaisesRegex(ContainerSafetyError, "formula"):
            inspect_container(self.snapshot(formula))

        dde = self.write_bytes(
            "dde.xlsx",
            self.workbook_bytes(formula="=cmd|' /C calc'!A0"),
        )
        with self.assertRaisesRegex(ContainerSafetyError, "DDE"):
            inspect_container(self.snapshot(dde))

        base = self.workbook_bytes()
        mutation_cases = {
            "macro.xlsx": (
                ("xl/vbaProject.bin", b"MZ"),
                "VBA",
            ),
            "payload.xlsx": (
                ("xl/embeddings/oleObject1.bin", b"MZ"),
                "embedded",
            ),
            "external.xlsx": (
                (
                    "xl/externalLinks/_rels/externalLink1.xml.rels",
                    (
                        b'<?xml version="1.0" encoding="UTF-8"?>'
                        b'<Relationships xmlns="http://schemas.openxmlformats.org/'
                        b'package/2006/relationships"><Relationship Id="rId1" '
                        b'Type="http://schemas.openxmlformats.org/officeDocument/'
                        b'2006/relationships/externalLinkPath" '
                        b'Target="file:///private.csv" TargetMode="External"/>'
                        b"</Relationships>"
                    ),
                ),
                "external relationship",
            ),
        }
        for name, ((member_name, member_value), message) in mutation_cases.items():
            with self.subTest(name=name):
                output = BytesIO()
                with zipfile.ZipFile(BytesIO(base)) as source_archive:
                    with zipfile.ZipFile(output, "w") as target_archive:
                        for item in source_archive.infolist():
                            target_archive.writestr(item, source_archive.read(item))
                        target_archive.writestr(member_name, member_value)
                with self.assertRaisesRegex(ContainerSafetyError, message):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, output.getvalue()))
                    )

    def test_xlsx_metadata_privacy_blocks_before_raw_byte_admission(self):
        relationship_workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(relationship_workbook)) as archive:
            relationship_xml = archive.read("_rels/.rels")
        relationship_workbook = self.replace_zip_member(
            relationship_workbook,
            "_rels/.rels",
            relationship_xml.replace(
                b'Id="rId1"', b'Id="owner@example.com"', 1
            ),
        )
        cases = {
            "creator.xlsx": (
                self.workbook_bytes(creator="creator@example.com"),
                "email",
            ),
            "last-editor.xlsx": (
                self.workbook_bytes(
                    last_modified_by="last.editor@example.com"
                ),
                "email",
            ),
            "custom-property.xlsx": (
                self.workbook_bytes(
                    custom_property=(
                        "campaign_note",
                        "Bearer abcdefghijklmnopqrstuvwxyz012345",
                    )
                ),
                "access_token",
            ),
            "worksheet-name.xlsx": (
                self.workbook_bytes(sheet_title="owner@example.com"),
                "email",
            ),
            "relationship-name.xlsx": (
                relationship_workbook,
                "email",
            ),
        }
        for name, (workbook_bytes, expected_category) in cases.items():
            with self.subTest(name=name):
                source = self.write_bytes(name, workbook_bytes)
                snapshot = self.snapshot(source)
                inventory = inspect_container(snapshot)
                decision = pre_scan_obvious_privacy(inventory)
                destination = self.base / "unsafe-admission" / name
                if decision.status == "pre_scan_clear":
                    admitted = admit_source(
                        snapshot,
                        inventory,
                        decision,
                        self.adapter_validation(snapshot, inventory),
                        destination,
                    )
                    self.assertEqual(workbook_bytes, admitted.source_path.read_bytes())
                    admitted.source_path.unlink()
                self.assertEqual("blocked_person_level", decision.status)
                self.assertIn(expected_category, decision.blocked_categories)
                with self.assertRaisesRegex(PrivacyAdmissionError, "privacy"):
                    admit_source(
                        snapshot,
                        inventory,
                        decision,
                        self.adapter_validation(snapshot, inventory),
                        destination,
                    )
                self.assertFalse(destination.exists())

    def test_xlsx_qualified_and_sheetdata_strings_never_bypass_privacy(self):
        base = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(base)) as archive:
            workbook_xml = archive.read("xl/workbook.xml")
            worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
        namespace_prefix = b"AKIAABCDEFGHIJKLMNOP"
        cases = {
            "unused-namespace-uri.xlsx": (
                "xl/workbook.xml",
                workbook_xml.replace(
                    b"<workbook ",
                    (
                        b'<workbook xmlns:unused="https://owner@example.com/ns" '
                    ),
                    1,
                ),
                "email",
            ),
            "unused-namespace-prefix.xlsx": (
                "xl/workbook.xml",
                workbook_xml.replace(
                    b"<workbook ",
                    b'<workbook xmlns:' + namespace_prefix + b'="urn:acme" ',
                    1,
                ),
                "cloud_credential",
            ),
            "unused-secret-header-namespace-prefix.xlsx": (
                "xl/workbook.xml",
                workbook_xml.replace(
                    b"<workbook ",
                    b'<workbook xmlns:api_key="urn:acme" ',
                    1,
                ),
                "secret_header",
            ),
            "qualified-element-name.xlsx": (
                "xl/workbook.xml",
                workbook_xml.replace(
                    b"<workbook ",
                    b'<workbook xmlns:unsafe="urn:acme" ',
                    1,
                ).replace(
                    b"</workbook>",
                    b"<unsafe:AKIAABCDEFGHIJKLMNOP /></workbook>",
                    1,
                ),
                "cloud_credential",
            ),
            "qualified-person-level-element-name.xlsx": (
                "xl/workbook.xml",
                workbook_xml.replace(
                    b"<workbook ",
                    b'<workbook xmlns:unsafe="urn:acme" ',
                    1,
                ).replace(
                    b"</workbook>",
                    b"<unsafe:email>aggregate</unsafe:email></workbook>",
                    1,
                ),
                "person_level_identifier",
            ),
            "qualified-attribute-name.xlsx": (
                "xl/workbook.xml",
                workbook_xml.replace(
                    b"<workbook ",
                    (
                        b'<workbook xmlns:unsafe="urn:acme" '
                        b'unsafe:AKIAABCDEFGHIJKLMNOP="aggregate" '
                    ),
                    1,
                ),
                "cloud_credential",
            ),
            "worksheet-row-attribute.xlsx": (
                "xl/worksheets/sheet1.xml",
                worksheet_xml.replace(
                    b'<row r="1"',
                    b'<row r="1" data-note="owner@example.com"',
                    1,
                ),
                "email",
            ),
            "worksheet-cell-attribute.xlsx": (
                "xl/worksheets/sheet1.xml",
                worksheet_xml.replace(
                    b'<c r="A1"',
                    (
                        b'<c r="A1" data-note="Bearer '
                        b'abcdefghijklmnopqrstuvwxyz012345"'
                    ),
                    1,
                ),
                "access_token",
            ),
            "unknown-sheetdata-child.xlsx": (
                "xl/worksheets/sheet1.xml",
                worksheet_xml.replace(
                    b"<sheetData>",
                    b'<sheetData><unknown note="owner@example.com" />',
                    1,
                ),
                None,
            ),
        }
        for name, (member_name, replacement, category) in cases.items():
            with self.subTest(name=name):
                self.assert_rejected_before_metadata_admission(
                    name=name,
                    source_bytes=self.replace_zip_member(
                        base,
                        member_name,
                        replacement,
                    ),
                    expected_category=category,
                )

    def test_outer_and_xlsx_archive_metadata_never_bypasses_privacy(self):
        csv_bytes = b"campaign_id,impressions\ncampaign-1,100\n"
        pii_extra = struct.pack("<HH", 0xCAFE, 17) + b"owner@example.com"
        workbook = self.workbook_bytes()
        cases = {
            "outer-archive-comment.zip": (
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    archive_comment=b"owner@example.com",
                ),
                "email",
            ),
            "outer-member-comment.zip": (
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    member_comment=b"owner@example.com",
                ),
                "email",
            ),
            "outer-member-extra.zip": (
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    member_extra=pii_extra,
                ),
                None,
            ),
            "xlsx-archive-comment.xlsx": (
                self.workbook_with_archive_metadata(
                    workbook,
                    archive_comment=b"owner@example.com",
                ),
                "email",
            ),
            "xlsx-member-comment.xlsx": (
                self.workbook_with_archive_metadata(
                    workbook,
                    member_comment=b"owner@example.com",
                ),
                "email",
            ),
            "xlsx-member-extra.xlsx": (
                self.workbook_with_archive_metadata(
                    workbook,
                    member_extra=pii_extra,
                ),
                None,
            ),
        }
        for name, (source_bytes, category) in cases.items():
            with self.subTest(name=name):
                self.assert_rejected_before_metadata_admission(
                    name=name,
                    source_bytes=source_bytes,
                    expected_category=category,
                )

    def test_archive_comments_and_qualified_names_bind_inventory_digest(self):
        csv_bytes = b"campaign_id,impressions\ncampaign-1,100\n"
        outer_hashes = []
        for comment in (b"Acme export one", b"Acme export two"):
            source = self.write_bytes(
                f"outer-{len(outer_hashes)}.zip",
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    archive_comment=comment,
                ),
            )
            inventory = inspect_container(self.snapshot(source))
            self.assertEqual(
                "pre_scan_clear",
                pre_scan_obvious_privacy(inventory).status,
            )
            outer_hashes.append(
                privacy_module.container_inventory_sha256(inventory)
            )
        self.assertNotEqual(*outer_hashes)

        workbook = self.workbook_bytes()
        xlsx_hashes = []
        for comment in (b"Acme workbook one", b"Acme workbook two"):
            source = self.write_bytes(
                f"xlsx-{len(xlsx_hashes)}.xlsx",
                self.workbook_with_archive_metadata(
                    workbook,
                    archive_comment=comment,
                    member_comment=b"Acme core properties",
                ),
            )
            inventory = inspect_container(self.snapshot(source))
            self.assertEqual(
                "pre_scan_clear",
                pre_scan_obvious_privacy(inventory).status,
            )
            xlsx_hashes.append(
                privacy_module.container_inventory_sha256(inventory)
            )
        self.assertNotEqual(*xlsx_hashes)

        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            workbook_xml = archive.read("xl/workbook.xml")
        qualified = self.replace_zip_member(
            workbook,
            "xl/workbook.xml",
            workbook_xml.replace(
                b"<workbook ",
                (
                    b'<workbook xmlns:acme="urn:acme:aggregate" '
                    b'acme:note="aggregate export" '
                ),
                1,
            ),
        )
        inventory = inspect_container(
            self.snapshot(self.write_bytes("safe-qualified.xlsx", qualified))
        )
        self.assertEqual(
            "pre_scan_clear",
            pre_scan_obvious_privacy(inventory).status,
        )
        metadata_values = {item.value for item in inventory.metadata}
        self.assertIn("acme", metadata_values)
        self.assertIn("urn:acme:aggregate", metadata_values)
        self.assertIn(
            "{urn:acme:aggregate}note",
            metadata_values,
        )

    def test_archive_metadata_limits_and_extra_fields_fail_closed(self):
        csv_bytes = b"campaign_id,impressions\ncampaign-1,100\n"
        too_long = self.zip_bytes_with_metadata(
            member_name="report.csv",
            member_value=csv_bytes,
            archive_comment=b"a" * 4_097,
        )
        with self.assertRaisesRegex(ContainerSafetyError, "metadata value length"):
            inspect_container(
                self.snapshot(self.write_bytes("long-comment.zip", too_long))
            )

        malformed_extra = b"\x01\x00\x05\x00abc"
        duplicate_extra = (
            struct.pack("<HH", 0xCAFE, 1)
            + b"a"
            + struct.pack("<HH", 0xCAFE, 1)
            + b"b"
        )
        unsupported_extra = struct.pack("<HH", 0xCAFE, 4) + b"Acme"
        binary_unicode_extra = self.unicode_extra_field(
            0x7075,
            b"report.csv",
            b"\xff",
        )
        cases = {
            "malformed-extra.zip": (
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    member_extra=malformed_extra,
                ),
                "malformed|container",
            ),
            "duplicate-extra.zip": (
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    member_extra=duplicate_extra,
                ),
                "duplicate ZIP extra",
            ),
            "unsupported-extra.zip": (
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    member_extra=unsupported_extra,
                ),
                "unsupported ZIP extra",
            ),
            "binary-extra.zip": (
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    member_extra=binary_unicode_extra,
                ),
                "binary or unsupported|malformed ZIP container",
            ),
            "binary-archive-comment.zip": (
                self.zip_bytes_with_metadata(
                    member_name="report.csv",
                    member_value=csv_bytes,
                    archive_comment=b"\xff",
                ),
                "binary or unsupported",
            ),
        }
        for name, (source_bytes, message) in cases.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(ContainerSafetyError, message):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, source_bytes))
                    )

        safe_extra = (
            self.unicode_extra_field(0x7075, b"report.csv", b"report.csv")
            + self.unicode_extra_field(0x6375, b"Acme report", b"Acme report")
        )
        safe_bytes = self.zip_bytes_with_metadata(
            member_name="report.csv",
            member_value=csv_bytes,
            member_comment=b"Acme report",
            member_extra=safe_extra,
        )
        safe_source = self.write_bytes("safe-unicode-extra.zip", safe_bytes)
        safe_snapshot = self.snapshot(safe_source)
        safe_inventory = inspect_container(safe_snapshot)
        safe_decision = pre_scan_obvious_privacy(safe_inventory)
        self.assertEqual("pre_scan_clear", safe_decision.status)
        safe_destination = self.base / "safe-extra" / "source.zip"
        admitted = admit_source(
            safe_snapshot,
            safe_inventory,
            safe_decision,
            self.adapter_validation(safe_snapshot, safe_inventory),
            safe_destination,
        )
        self.assertEqual(safe_bytes, admitted.source_path.read_bytes())

        xlsx_comment = b"Acme core properties"
        xlsx_extra = self.unicode_extra_field(
            0x7075,
            b"docProps/core.xml",
            b"docProps/core.xml",
        ) + self.unicode_extra_field(
            0x6375,
            xlsx_comment,
            xlsx_comment,
        )
        safe_xlsx_bytes = self.workbook_with_archive_metadata(
            self.workbook_bytes(),
            member_comment=xlsx_comment,
            member_extra=xlsx_extra,
        )
        safe_xlsx_source = self.write_bytes(
            "safe-unicode-extra.xlsx",
            safe_xlsx_bytes,
        )
        safe_xlsx_snapshot = self.snapshot(safe_xlsx_source)
        safe_xlsx_inventory = inspect_container(safe_xlsx_snapshot)
        safe_xlsx_decision = pre_scan_obvious_privacy(safe_xlsx_inventory)
        self.assertEqual("pre_scan_clear", safe_xlsx_decision.status)
        safe_xlsx_destination = self.base / "safe-extra" / "source.xlsx"
        admitted_xlsx = admit_source(
            safe_xlsx_snapshot,
            safe_xlsx_inventory,
            safe_xlsx_decision,
            self.adapter_validation(safe_xlsx_snapshot, safe_xlsx_inventory),
            safe_xlsx_destination,
        )
        self.assertEqual(
            safe_xlsx_bytes,
            admitted_xlsx.source_path.read_bytes(),
        )

        divergent = self.mutate_first_local_extra_byte(safe_bytes)
        with self.assertRaisesRegex(ContainerSafetyError, "divergent ZIP extra"):
            inspect_container(
                self.snapshot(
                    self.write_bytes("divergent-local-extra.zip", divergent)
                )
            )

    def test_allowed_xlsx_xml_requires_safe_utf8_structure(self):
        workbook = self.workbook_bytes()
        namespace = (
            'xmlns:cp="http://schemas.openxmlformats.org/package/'
            '2006/metadata/core-properties"'
        )
        cases = {
            "utf16le-comment.xlsx": (
                "utf-16-le",
                b"\xff\xfe",
                f'<cp:coreProperties {namespace}><!-- Acme --></cp:coreProperties>',
            ),
            "utf16be-pi.xlsx": (
                "utf-16-be",
                b"\xfe\xff",
                f'<cp:coreProperties {namespace}><?acme safe?></cp:coreProperties>',
            ),
            "utf32le-doctype.xlsx": (
                "utf-32-le",
                b"\xff\xfe\x00\x00",
                (
                    '<!DOCTYPE cp:coreProperties>'
                    f'<cp:coreProperties {namespace}/>'
                ),
            ),
            "utf32be-entity.xlsx": (
                "utf-32-be",
                b"\x00\x00\xfe\xff",
                (
                    '<!DOCTYPE cp:coreProperties [<!ENTITY safe "Acme">]>'
                    f'<cp:coreProperties {namespace}>&safe;</cp:coreProperties>'
                ),
            ),
        }
        for name, (encoding, bom, body) in cases.items():
            with self.subTest(name=name):
                declaration = (
                    '<?xml version="1.0" encoding="'
                    + ("UTF-16" if "16" in encoding else "UTF-32")
                    + '"?>'
                )
                source_bytes = self.replace_zip_member(
                    workbook,
                    "docProps/core.xml",
                    bom + (declaration + body).encode(encoding),
                )
                self.assert_rejected_before_metadata_admission(
                    name=name,
                    source_bytes=source_bytes,
                )

    def test_raw_cells_outside_declared_dimension_never_bypass_privacy(self):
        workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
        self.assertIn(b'<dimension ref="A1:B2"', worksheet_xml)
        cases = {
            "raw-a3.xlsx": (
                b'<row r="3"><c r="A3" t="str">'
                b"<v>owner@example.com</v></c></row>"
            ),
            "raw-a100.xlsx": (
                b'<row r="100"><c r="A100" t="str">'
                b"<v>owner@example.com</v></c></row>"
            ),
            "raw-z2.xlsx": (
                b'<row r="2"><c r="Z2" t="inlineStr"><is>'
                b"<t>owner@example.com</t></is></c></row>"
            ),
        }
        for name, injected_row in cases.items():
            with self.subTest(name=name):
                if name == "raw-z2.xlsx":
                    replacement = worksheet_xml.replace(
                        b"</row></sheetData>",
                        injected_row.removeprefix(b'<row r="2">')
                        .removesuffix(b"</row>")
                        + b"</row></sheetData>",
                        1,
                    )
                else:
                    replacement = worksheet_xml.replace(
                        b"</sheetData>",
                        injected_row + b"</sheetData>",
                        1,
                    )
                source_bytes = self.replace_zip_member(
                    workbook,
                    "xl/worksheets/sheet1.xml",
                    replacement,
                )
                with self.assertRaisesRegex(
                    ContainerSafetyError,
                    r"worksheet (?:row|cell) falls outside.*dimension",
                ):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, source_bytes))
                    )

    def test_outer_and_xlsx_member_names_reject_nul_and_invalid_encoding(self):
        outer_nul = self.replace_archive_name_bytes(
            self.zip_bytes_with_metadata(
                member_name="report.csvX",
                member_value=b"campaign_id,impressions\ncampaign-1,100\n",
            ),
            b"report.csvX",
            b"report.csv\x00",
        )
        renamed_workbook = self.rename_zip_member(
            self.workbook_bytes(),
            "xl/workbook.xml",
            "xl/workbook.xmlX",
        )
        xlsx_nul = self.replace_archive_name_bytes(
            renamed_workbook,
            b"xl/workbook.xmlX",
            b"xl/workbook.xml\x00",
        )
        for name, source_bytes in {
            "outer-nul-name.zip": outer_nul,
            "xlsx-nul-name.xlsx": xlsx_nul,
        }.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(ContainerSafetyError, "NUL|member name"):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, source_bytes))
                    )

        outer_unicode = self.zip_bytes_with_metadata(
            member_name="réport.csv",
            member_value=b"campaign_id,impressions\ncampaign-1,100\n",
        )
        outer_invalid = self.replace_archive_name_bytes(
            outer_unicode,
            "réport.csv".encode(),
            b"r\xff\xa9port.csv",
        )
        xlsx_unicode = self.rename_zip_member(
            self.workbook_bytes(),
            "docProps/core.xml",
            "docProps/côre.xml",
        )
        xlsx_invalid = self.replace_archive_name_bytes(
            xlsx_unicode,
            "docProps/côre.xml".encode(),
            b"docProps/c\xff\xb4re.xml",
        )
        for name, source_bytes in {
            "outer-invalid-name.zip": outer_invalid,
            "xlsx-invalid-name.xlsx": xlsx_invalid,
        }.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(
                    ContainerSafetyError,
                    "encoding|member name|container",
                ):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, source_bytes))
                    )

    def test_repeated_xml_name_occurrences_consume_metadata_budget(self):
        workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            workbook_xml = archive.read("xl/workbook.xml")
        repetitions = {
            "repeated-names.xlsx": b"<safe/>" * 100_001,
            "repeated-declarations.xlsx": (
                b'<safe xmlns:acme="urn:acme"/>' * 100_001
            ),
        }
        for name, repeated in repetitions.items():
            with self.subTest(name=name):
                source_bytes = self.replace_zip_member(
                    workbook,
                    "xl/workbook.xml",
                    workbook_xml.replace(
                        b"</workbook>",
                        repeated + b"</workbook>",
                        1,
                    ),
                )
                with self.assertRaisesRegex(
                    ContainerSafetyError,
                    "metadata count",
                ):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, source_bytes)),
                        limits=ContainerLimits(
                            expansion_ratio=10_000.0,
                            metadata_total_chars=100_000_000,
                            processing_seconds=60.0,
                        ),
                    )

    def test_xml_occurrence_and_name_boundaries_preserve_safe_bytes(self):
        workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            core_xml = archive.read("docProps/core.xml")
            workbook_xml = archive.read("xl/workbook.xml")
        utf8_bom_workbook = self.replace_zip_member(
            workbook,
            "docProps/core.xml",
            b"\xef\xbb\xbf" + core_xml,
        )
        safe_source = self.write_bytes("safe-utf8-bom.xlsx", utf8_bom_workbook)
        safe_snapshot = self.snapshot(safe_source)
        safe_inventory = inspect_container(safe_snapshot)
        safe_decision = pre_scan_obvious_privacy(safe_inventory)
        self.assertEqual("pre_scan_clear", safe_decision.status)
        self.assertIn(
            "100",
            {cell.value for cell in safe_inventory.cells},
        )
        admitted = admit_source(
            safe_snapshot,
            safe_inventory,
            safe_decision,
            self.adapter_validation(safe_snapshot, safe_inventory),
            self.base / "safe-utf8-bom" / "source.xlsx",
        )
        self.assertEqual(utf8_bom_workbook, admitted.source_path.read_bytes())

        exact_count = len(safe_inventory.metadata)
        exact_inventory = inspect_container(
            safe_snapshot,
            limits=ContainerLimits(metadata_count=exact_count),
        )
        self.assertEqual(exact_count, len(exact_inventory.metadata))
        with self.assertRaisesRegex(ContainerSafetyError, "metadata count"):
            inspect_container(
                safe_snapshot,
                limits=ContainerLimits(metadata_count=exact_count - 1),
            )

        occurrence_hashes = []
        for count in (1, 2):
            occurrence_bytes = self.replace_zip_member(
                workbook,
                "xl/workbook.xml",
                workbook_xml.replace(
                    b"</workbook>",
                    (b"<safe/>" * count) + b"</workbook>",
                    1,
                ),
            )
            occurrence_inventory = inspect_container(
                self.snapshot(
                    self.write_bytes(
                        f"safe-occurrence-{count}.xlsx",
                        occurrence_bytes,
                    )
                )
            )
            occurrence_hashes.append(
                privacy_module.container_inventory_sha256(occurrence_inventory)
            )
        self.assertNotEqual(*occurrence_hashes)

        unicode_zip = self.zip_bytes_with_metadata(
            member_name="réport.csv",
            member_value=b"campaign_id,impressions\ncampaign-1,100\n",
        )
        unicode_source = self.write_bytes("safe-unicode-name.zip", unicode_zip)
        unicode_snapshot = self.snapshot(unicode_source)
        unicode_inventory = inspect_container(unicode_snapshot)
        unicode_decision = pre_scan_obvious_privacy(unicode_inventory)
        self.assertEqual("pre_scan_clear", unicode_decision.status)
        self.assertIn(
            "réport.csv",
            {item.value for item in unicode_inventory.metadata},
        )
        unicode_admitted = admit_source(
            unicode_snapshot,
            unicode_inventory,
            unicode_decision,
            self.adapter_validation(unicode_snapshot, unicode_inventory),
            self.base / "safe-unicode-name" / "source.zip",
        )
        self.assertEqual(unicode_zip, unicode_admitted.source_path.read_bytes())

    def test_parser_ignored_worksheet_values_never_reach_admission(self):
        workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
        attacks = {
            "duplicate-v.xlsx": worksheet_xml.replace(
                b"<v>100</v>",
                b"<v>100</v><v>owner@example.com</v>",
                1,
            ),
            "duplicate-inline-t.xlsx": worksheet_xml.replace(
                b"<t>campaign-1</t>",
                b"<t>owner@example.com</t><t>campaign-1</t>",
                1,
            ),
            "inline-plus-v.xlsx": worksheet_xml.replace(
                b"</is></c><c r=\"B2\"",
                b"</is><v>owner@example.com</v></c><c r=\"B2\"",
                1,
            ),
            "numeric-plus-inline.xlsx": worksheet_xml.replace(
                b"<v>100</v>",
                b"<v>100</v><is><t>owner@example.com</t></is>",
                1,
            ),
            "out-of-order-row.xlsx": worksheet_xml.replace(
                b'<dimension ref="A1:B2"',
                b'<dimension ref="A1:B3"',
                1,
            ).replace(
                (
                    b'<row r="2"><c r="A2" t="inlineStr"><is>'
                    b"<t>campaign-1</t></is></c><c r=\"B2\" t=\"n\">"
                    b"<v>100</v></c></row>"
                ),
                (
                    b'<row r="3"><c r="A3" t="inlineStr"><is>'
                    b"<t>campaign-1</t></is></c><c r=\"B3\" t=\"n\">"
                    b"<v>100</v></c></row>"
                    b'<row r="2"><c r="A2" t="inlineStr"><is>'
                    b"<t>owner@example.com</t></is></c></row>"
                ),
                1,
            ),
        }
        for name, replacement in attacks.items():
            with self.subTest(name=name):
                self.assert_rejected_before_metadata_admission(
                    name=name,
                    source_bytes=self.replace_zip_member(
                        workbook,
                        "xl/worksheets/sheet1.xml",
                        replacement,
                    ),
                    expected_category="email",
                )

    def test_zip_envelope_rejects_prefix_trailer_and_inter_record_slack(self):
        outer = self.zip_bytes_with_metadata(
            member_name="report.csv",
            member_value=b"campaign_id,impressions\ncampaign-1,100\n",
        )
        workbook = self.workbook_bytes()
        cases = {
            "outer-prefix.zip": b"prefix" + outer,
            "outer-trailer.zip": outer + b"trailer",
            "outer-slack.zip": self.insert_zip_central_slack(
                outer,
                b"slack",
            ),
            "xlsx-prefix.xlsx": b"prefix" + workbook,
            "xlsx-trailer.xlsx": workbook + b"trailer",
            "xlsx-slack.xlsx": self.insert_zip_central_slack(
                workbook,
                b"slack",
            ),
        }
        for name, source_bytes in cases.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(
                    ContainerSafetyError,
                    "ZIP|XLSX|container|envelope|slack",
                ):
                    inspect_container(
                        self.snapshot(self.write_bytes(name, source_bytes))
                    )

    def test_sparse_workbook_and_data_descriptor_zip_remain_safe(self):
        sparse = Workbook()
        sheet = sparse.active
        sheet.title = "data"
        sheet["A1"] = "campaign_id"
        sheet["C1"] = "impressions"
        sheet["A3"] = "campaign-1"
        sheet["C3"] = 100
        sparse_output = BytesIO()
        sparse.save(sparse_output)
        sparse_bytes = sparse_output.getvalue()
        sparse_source = self.write_bytes("safe-sparse.xlsx", sparse_bytes)
        sparse_snapshot = self.snapshot(sparse_source)
        sparse_inventory = inspect_container(sparse_snapshot)
        self.assertEqual("pre_scan_clear", pre_scan_obvious_privacy(sparse_inventory).status)
        self.assertEqual(
            {"campaign_id", "impressions", "campaign-1", "100"},
            {item.value for item in sparse_inventory.raw_values},
        )
        raw_value_count = len(sparse_inventory.raw_values)
        exact_inventory = inspect_container(
            sparse_snapshot,
            limits=ContainerLimits(raw_value_count=raw_value_count),
        )
        self.assertEqual(raw_value_count, len(exact_inventory.raw_values))
        with self.assertRaisesRegex(
            ContainerSafetyError,
            "raw worksheet value count",
        ):
            inspect_container(
                sparse_snapshot,
                limits=ContainerLimits(raw_value_count=raw_value_count - 1),
            )
        longest_raw_value = max(
            len(item.value) for item in sparse_inventory.raw_values
        )
        with self.assertRaisesRegex(
            ContainerSafetyError,
            "raw worksheet value length",
        ):
            inspect_container(
                sparse_snapshot,
                limits=ContainerLimits(
                    raw_value_chars=longest_raw_value - 1,
                ),
            )
        raw_value_chars = sum(
            len(item.value) for item in sparse_inventory.raw_values
        )
        with self.assertRaisesRegex(
            ContainerSafetyError,
            "raw worksheet value character",
        ):
            inspect_container(
                sparse_snapshot,
                limits=ContainerLimits(
                    raw_value_total_chars=raw_value_chars - 1,
                ),
            )
        raw_privacy_attack = ContainerInventory(
            media_type=sparse_inventory.media_type,
            tables=sparse_inventory.tables,
            headers=sparse_inventory.headers,
            cells=sparse_inventory.cells,
            row_count=sparse_inventory.row_count,
            metadata=sparse_inventory.metadata,
            raw_values=sparse_inventory.raw_values
            + (
                InventoryMetadata(
                    source="xlsx_raw_value",
                    name="xl/worksheets/sheet1.xml:A2:v",
                    value="owner@example.com",
                ),
            ),
        )
        self.assertEqual(
            "blocked_person_level",
            pre_scan_obvious_privacy(raw_privacy_attack).status,
        )
        self.assertNotEqual(
            privacy_module.container_inventory_sha256(sparse_inventory),
            privacy_module.container_inventory_sha256(raw_privacy_attack),
        )
        sparse_admitted = admit_source(
            sparse_snapshot,
            sparse_inventory,
            pre_scan_obvious_privacy(sparse_inventory),
            self.adapter_validation(sparse_snapshot, sparse_inventory),
            self.base / "safe-sparse" / "source.xlsx",
        )
        self.assertEqual(sparse_bytes, sparse_admitted.source_path.read_bytes())

        class NonSeekableBytesIO(BytesIO):
            def seekable(self):
                return False

            def seek(self, *args, **kwargs):
                raise OSError("fixture is intentionally non-seekable")

        descriptor_output = NonSeekableBytesIO()
        with zipfile.ZipFile(
            descriptor_output,
            "w",
            compression=zipfile.ZIP_DEFLATED,
        ) as archive:
            archive.writestr(
                "report.csv",
                b"campaign_id,impressions\ncampaign-1,100\n",
            )
            archive.comment = b"Acme descriptor export"
        descriptor_bytes = descriptor_output.getvalue()
        self.assertIn(b"PK\x07\x08", descriptor_bytes)
        descriptor_source = self.write_bytes(
            "safe-data-descriptor.zip",
            descriptor_bytes,
        )
        descriptor_snapshot = self.snapshot(descriptor_source)
        descriptor_inventory = inspect_container(descriptor_snapshot)
        descriptor_decision = pre_scan_obvious_privacy(descriptor_inventory)
        self.assertEqual("pre_scan_clear", descriptor_decision.status)
        descriptor_admitted = admit_source(
            descriptor_snapshot,
            descriptor_inventory,
            descriptor_decision,
            self.adapter_validation(descriptor_snapshot, descriptor_inventory),
            self.base / "safe-descriptor" / "source.zip",
        )
        self.assertEqual(
            descriptor_bytes,
            descriptor_admitted.source_path.read_bytes(),
        )

    def test_shared_strings_plain_rich_referenced_and_unreferenced_are_safe(self):
        workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
        referenced_worksheet = worksheet_xml.replace(
            (
                b'<c r="A2" t="inlineStr"><is>'
                b"<t>campaign-1</t></is></c>"
            ),
            b'<c r="A2" t="s"><v>0</v></c>',
            1,
        )
        cases = {
            "plain-shared.xlsx": (
                b"<si><t>campaign-1</t></si>",
                referenced_worksheet,
                1,
                "campaign-1",
            ),
            "rich-shared.xlsx": (
                (
                    b"<si><r><rPr><b /><color rgb=\"FF112233\" />"
                    b"</rPr><t>campaign</t></r><r><t>-1</t></r></si>"
                ),
                referenced_worksheet,
                1,
                "campaign-1",
            ),
            "unreferenced-shared.xlsx": (
                b"<si><t>Acme unused label</t></si>",
                worksheet_xml,
                0,
                "Acme unused label",
            ),
        }
        for name, (items, sheet_xml, count, logical_value) in cases.items():
            with self.subTest(name=name):
                source_bytes = self.workbook_with_shared_strings(
                    workbook,
                    shared_strings_xml=self.shared_strings_xml(
                        items,
                        count=count,
                        unique_count=1,
                    ),
                    worksheet_xml=sheet_xml,
                )
                source = self.write_bytes(name, source_bytes)
                snapshot = self.snapshot(source)
                inventory = inspect_container(snapshot)
                decision = pre_scan_obvious_privacy(inventory)
                self.assertEqual("pre_scan_clear", decision.status)
                self.assertEqual(
                    (logical_value,),
                    tuple(item.value for item in inventory.logical_values),
                )
                if name == "rich-shared.xlsx":
                    self.assertTrue(
                        {"campaign", "-1"}.issubset(
                            {item.value for item in inventory.raw_values}
                        )
                    )
                admitted = admit_source(
                    snapshot,
                    inventory,
                    decision,
                    self.adapter_validation(snapshot, inventory),
                    self.base / "admitted-shared" / name,
                )
                self.assertEqual(source_bytes, admitted.source_path.read_bytes())

    def test_shared_strings_logical_privacy_grammar_counts_and_indexes(self):
        workbook = self.workbook_bytes()
        with zipfile.ZipFile(BytesIO(workbook)) as archive:
            worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
        referenced_worksheet = worksheet_xml.replace(
            (
                b'<c r="A2" t="inlineStr"><is>'
                b"<t>campaign-1</t></is></c>"
            ),
            b'<c r="A2" t="s"><v>0</v></c>',
            1,
        )
        split_attacks = {
            "split-email.xlsx": (
                b"<si><r><t>owner@</t></r><r><t>example.com</t></r></si>",
                "email",
            ),
            "split-secret.xlsx": (
                (
                    b"<si><r><t>AKIA</t></r>"
                    b"<r><t>ABCDEFGHIJKLMNOP</t></r></si>"
                ),
                "cloud_credential",
            ),
        }
        for name, (items, category) in split_attacks.items():
            with self.subTest(name=name):
                self.assert_rejected_before_metadata_admission(
                    name=name,
                    source_bytes=self.workbook_with_shared_strings(
                        workbook,
                        shared_strings_xml=self.shared_strings_xml(
                            items,
                            count=0,
                            unique_count=1,
                        ),
                    ),
                    expected_category=category,
                )

        invalid_items = {
            "phonetic-shared.xlsx": b"<si><t>Acme</t><rPh><t>hidden</t></rPh></si>",
            "extension-shared.xlsx": b"<si><t>Acme</t><extLst /></si>",
            "unknown-shared.xlsx": b"<si><unknown>hidden</unknown></si>",
            "mixed-shared.xlsx": b"<si><t>Acme</t><r><t>hidden</t></r></si>",
        }
        for name, items in invalid_items.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(
                    ContainerSafetyError,
                    "shared string|non-data XLSX|payload",
                ):
                    inspect_container(
                        self.snapshot(
                            self.write_bytes(
                                name,
                                self.workbook_with_shared_strings(
                                    workbook,
                                    shared_strings_xml=self.shared_strings_xml(
                                        items,
                                        count=0,
                                        unique_count=1,
                                    ),
                                ),
                            )
                        )
                    )

        invalid_tables = {
            "count-mismatch.xlsx": (
                self.shared_strings_xml(
                    b"<si><t>campaign-1</t></si>",
                    count=0,
                    unique_count=1,
                ),
                referenced_worksheet,
            ),
            "unique-mismatch.xlsx": (
                self.shared_strings_xml(
                    b"<si><t>campaign-1</t></si>",
                    count=1,
                    unique_count=2,
                ),
                referenced_worksheet,
            ),
            "index-out-of-range.xlsx": (
                self.shared_strings_xml(
                    b"<si><t>campaign-1</t></si>",
                    count=1,
                    unique_count=1,
                ),
                referenced_worksheet.replace(b"<v>0</v>", b"<v>1</v>", 1),
            ),
            "index-negative.xlsx": (
                self.shared_strings_xml(
                    b"<si><t>campaign-1</t></si>",
                    count=1,
                    unique_count=1,
                ),
                referenced_worksheet.replace(b"<v>0</v>", b"<v>-1</v>", 1),
            ),
            "index-noninteger.xlsx": (
                self.shared_strings_xml(
                    b"<si><t>campaign-1</t></si>",
                    count=1,
                    unique_count=1,
                ),
                referenced_worksheet.replace(b"<v>0</v>", b"<v>x</v>", 1),
            ),
            "duplicate-logical-value.xlsx": (
                self.shared_strings_xml(
                    (
                        b"<si><t>campaign-1</t></si>"
                        b"<si><r><t>campaign</t></r><r><t>-1</t></r></si>"
                    ),
                    count=1,
                    unique_count=2,
                ),
                referenced_worksheet,
            ),
        }
        for name, (shared_xml, sheet_xml) in invalid_tables.items():
            with self.subTest(name=name):
                with self.assertRaisesRegex(
                    ContainerSafetyError,
                    "shared string|non-data XLSX|payload",
                ):
                    inspect_container(
                        self.snapshot(
                            self.write_bytes(
                                name,
                                self.workbook_with_shared_strings(
                                    workbook,
                                    shared_strings_xml=shared_xml,
                                    worksheet_xml=sheet_xml,
                                ),
                            )
                        )
                    )

        bounded_bytes = self.workbook_with_shared_strings(
            workbook,
            shared_strings_xml=self.shared_strings_xml(
                b"<si><r><t>Acme </t></r><r><t>logical</t></r></si>",
                count=0,
                unique_count=1,
            ),
        )
        bounded_source = self.write_bytes("bounded-shared.xlsx", bounded_bytes)
        bounded_snapshot = self.snapshot(bounded_source)
        bounded_inventory = inspect_container(bounded_snapshot)
        self.assertEqual(
            ("Acme logical",),
            tuple(item.value for item in bounded_inventory.logical_values),
        )
        logical_chars = len("Acme logical")
        exact_inventory = inspect_container(
            bounded_snapshot,
            limits=ContainerLimits(
                logical_value_count=1,
                logical_value_chars=logical_chars,
                logical_value_total_chars=logical_chars,
            ),
        )
        self.assertEqual(1, len(exact_inventory.logical_values))
        for field, message in {
            "logical_value_count": "logical shared string count",
            "logical_value_chars": "logical shared string length",
            "logical_value_total_chars": "logical shared string character",
        }.items():
            with self.subTest(field=field):
                limits = {
                    "logical_value_count": 1,
                    "logical_value_chars": logical_chars,
                    "logical_value_total_chars": logical_chars,
                }
                limits[field] -= 1
                with self.assertRaisesRegex(ContainerSafetyError, message):
                    inspect_container(
                        bounded_snapshot,
                        limits=ContainerLimits(**limits),
                    )
        changed_logical = ContainerInventory(
            media_type=bounded_inventory.media_type,
            tables=bounded_inventory.tables,
            headers=bounded_inventory.headers,
            cells=bounded_inventory.cells,
            row_count=bounded_inventory.row_count,
            metadata=bounded_inventory.metadata,
            raw_values=bounded_inventory.raw_values,
            logical_values=(
                InventoryMetadata(
                    source="xlsx_logical_shared_string",
                    name="xl/sharedstrings.xml:si[0]",
                    value="Acme changed",
                ),
            ),
        )
        self.assertNotEqual(
            privacy_module.container_inventory_sha256(bounded_inventory),
            privacy_module.container_inventory_sha256(changed_logical),
        )

    def test_zip_member_name_privacy_blocks_before_raw_byte_admission(self):
        source_bytes = self.zip_bytes(
            [
                (
                    "owner@example.com.csv",
                    b"campaign_id,impressions\ncampaign-1,100\n",
                )
            ]
        )
        source = self.write_bytes("member-name.zip", source_bytes)
        snapshot = self.snapshot(source)
        inventory = inspect_container(snapshot)
        decision = pre_scan_obvious_privacy(inventory)
        destination = self.base / "unsafe-admission" / "member-name.zip"
        if decision.status == "pre_scan_clear":
            admitted = admit_source(
                snapshot,
                inventory,
                decision,
                self.adapter_validation(snapshot, inventory),
                destination,
            )
            self.assertEqual(source_bytes, admitted.source_path.read_bytes())
            admitted.source_path.unlink()
        self.assertEqual("blocked_person_level", decision.status)
        self.assertIn("email", decision.blocked_categories)
        with self.assertRaisesRegex(PrivacyAdmissionError, "privacy"):
            admit_source(
                snapshot,
                inventory,
                decision,
                self.adapter_validation(snapshot, inventory),
                destination,
            )
        self.assertFalse(destination.exists())

    def test_source_basename_privacy_blocks_every_supported_container(self):
        safe_payloads = {
            ".csv": b"campaign_id,impressions\ncampaign-1,100\n",
            ".json": (
                b'[{"campaign_id":"campaign-1","impressions":"100"}]'
            ),
            ".xlsx": self.workbook_bytes(),
            ".zip": self.zip_bytes([
                (
                    "reports/aggregate.csv",
                    b"campaign_id,impressions\ncampaign-1,100\n",
                )
            ]),
        }
        private_names = {
            "person@example.com": "email",
            "AKIAABCDEFGHIJKLMNOP": "cloud_credential",
        }
        for private_name, category in private_names.items():
            for suffix, source_bytes in safe_payloads.items():
                with self.subTest(private_name=private_name, suffix=suffix):
                    source = self.write_bytes(
                        f"{private_name}{suffix}", source_bytes
                    )
                    snapshot = self.snapshot(source)
                    inventory = inspect_container(snapshot)
                    decision = pre_scan_obvious_privacy(
                        inventory, source_name=snapshot.original_path.name
                    )
                    self.assertEqual("blocked_person_level", decision.status)
                    self.assertIn(category, decision.blocked_categories)
                    destination = (
                        self.base
                        / "unsafe-basename"
                        / f"{category}-source{suffix}"
                    )
                    with self.assertRaisesRegex(
                        PrivacyAdmissionError, "privacy"
                    ):
                        admit_source(
                            snapshot,
                            inventory,
                            decision,
                            self.adapter_validation(snapshot, inventory),
                            destination,
                        )
                    self.assertFalse(destination.exists())
                    if suffix in {".xlsx", ".zip"}:
                        inventory_only = pre_scan_obvious_privacy(inventory)
                        self.assertEqual(
                            "pre_scan_clear", inventory_only.status
                        )
                        with self.assertRaisesRegex(
                            PrivacyAdmissionError, "privacy"
                        ):
                            admit_source(
                                snapshot,
                                inventory,
                                inventory_only,
                                self.adapter_validation(snapshot, inventory),
                                destination,
                            )
                        self.assertFalse(destination.exists())

    def test_safe_source_basenames_preserve_exact_raw_bytes(self):
        safe_payloads = {
            "aggregate-export.csv": (
                b"campaign_id,impressions\ncampaign-1,100\n"
            ),
            "aggregate-export.json": (
                b'[{"campaign_id":"campaign-1","impressions":"100"}]'
            ),
            "aggregate-export.xlsx": self.workbook_bytes(),
            "aggregate-export.zip": self.zip_bytes([
                (
                    "reports/aggregate.csv",
                    b"campaign_id,impressions\ncampaign-1,100\n",
                )
            ]),
        }
        for source_name, source_bytes in safe_payloads.items():
            with self.subTest(source_name=source_name):
                source = self.write_bytes(source_name, source_bytes)
                snapshot = self.snapshot(source)
                inventory = inspect_container(snapshot)
                decision = pre_scan_obvious_privacy(
                    inventory, source_name=snapshot.original_path.name
                )
                self.assertEqual("pre_scan_clear", decision.status)
                destination = self.base / "safe-basename" / source_name
                admitted = admit_source(
                    snapshot,
                    inventory,
                    decision,
                    self.adapter_validation(snapshot, inventory),
                    destination,
                )
                self.assertEqual(source_bytes, admitted.source_path.read_bytes())

    def test_json_xlsx_and_nested_zip_table_names_are_privacy_scanned(self):
        sources = {
            "table-name.json": (
                b'{"person@example.com":[{"campaign_id":"campaign-1",'
                b'"impressions":"100"}]}'
            ),
            "table-name.xlsx": self.workbook_bytes(
                sheet_title="person@example.com"
            ),
            "nested-table.zip": self.zip_bytes([
                (
                    "reports/aggregate.json",
                    b'{"person@example.com":[{"campaign_id":"campaign-1",'
                    b'"impressions":"100"}]}',
                )
            ]),
            "secret-table.json": (
                b'{"AKIAABCDEFGHIJKLMNOP":[{"campaign_id":"campaign-1",'
                b'"impressions":"100"}]}'
            ),
        }
        for source_name, source_bytes in sources.items():
            with self.subTest(source_name=source_name):
                inventory = inspect_container(
                    self.snapshot(self.write_bytes(source_name, source_bytes))
                )
                private_value, expected = (
                    ("AKIAABCDEFGHIJKLMNOP", "cloud_credential")
                    if source_name == "secret-table.json"
                    else ("person@example.com", "email")
                )
                self.assertTrue(
                    any(private_value in table for table in inventory.tables)
                )
                decision = pre_scan_obvious_privacy(inventory)
                self.assertEqual("blocked_person_level", decision.status)
                self.assertIn(expected, decision.blocked_categories)

    def test_safe_xlsx_metadata_and_member_names_preserve_raw_bytes(self):
        workbook_bytes = self.workbook_bytes(
            creator="Acme Analytics",
            last_modified_by="Acme Operator",
            custom_property=("campaign_note", "aggregate export"),
            sheet_title="Campaign Data",
        )
        source_bytes = self.zip_bytes(
            [("reports/acme-campaign.xlsx", workbook_bytes)]
        )
        source = self.write_bytes("safe-metadata.zip", source_bytes)
        snapshot = self.snapshot(source)
        inventory = inspect_container(snapshot)
        decision = pre_scan_obvious_privacy(inventory)
        self.assertEqual("pre_scan_clear", decision.status)
        self.assertTrue(inventory.metadata)
        destination = self.base / "safe-admission" / "safe-metadata.zip"
        admitted = admit_source(
            snapshot,
            inventory,
            decision,
            self.adapter_validation(snapshot, inventory),
            destination,
        )
        self.assertEqual(source_bytes, admitted.source_path.read_bytes())

    def test_unsupported_xlsx_table_name_fails_closed_before_admission(self):
        source = self.write_bytes(
            "unsafe-table.xlsx",
            self.workbook_bytes(table_name="AKIAABCDEFGHIJKLMNOP"),
        )
        with self.assertRaisesRegex(ContainerSafetyError, "non-data XLSX"):
            inspect_container(self.snapshot(source))
        self.assertFalse(self.durable.exists())

    def test_plain_long_numeric_value_is_not_assumed_to_be_a_phone(self):
        inventory = self.csv_inventory(
            headers=["customer_id", "ad_id", "impressions"],
            rows=[["1234567890123456", "9876543210987654", "500"]],
        )
        decision = pre_scan_obvious_privacy(inventory)
        self.assertEqual("pre_scan_clear", decision.status)

    def test_phone_and_email_values_block_pre_scan(self):
        inventory = self.csv_inventory(
            headers=["contact_phone", "campaign_label"],
            rows=[["+1 212 555 0100", "owner@example.com"]],
        )
        decision = pre_scan_obvious_privacy(inventory)
        self.assertEqual("blocked_person_level", decision.status)
        self.assertEqual(
            ("email", "phone"),
            decision.blocked_categories,
        )

    def test_device_token_blocks_only_in_device_header_context(self):
        token = "a" * 32
        ambiguous = self.csv_inventory(
            headers=["campaign_token"],
            rows=[[token]],
        )
        self.assertEqual(
            "pre_scan_clear",
            pre_scan_obvious_privacy(ambiguous).status,
        )
        device = self.csv_inventory(
            headers=["advertising_id"],
            rows=[[token]],
        )
        self.assertEqual(
            ("device_token",),
            pre_scan_obvious_privacy(device).blocked_categories,
        )

    def test_secret_headers_and_high_confidence_secret_values_block(self):
        inventory = self.csv_inventory(
            headers=[
                "api_key",
                "campaign_note",
                "credential_material",
                "cloud_reference",
            ],
            rows=[
                [
                    "configured",
                    "Bearer abcdefghijklmnopqrstuvwxyz012345",
                    "-----BEGIN PRIVATE KEY-----",
                    "AKIAABCDEFGHIJKLMNOP",
                ]
            ],
        )
        decision = pre_scan_obvious_privacy(inventory)
        self.assertEqual("blocked_person_level", decision.status)
        self.assertEqual(
            (
                "access_token",
                "cloud_credential",
                "private_key",
                "secret_header",
            ),
            decision.blocked_categories,
        )
        self.assertNotIn("AKIA", repr(decision))

    def test_every_header_and_value_is_scanned_without_recording_values(self):
        header_only = ContainerInventory(
            media_type="text/csv",
            tables=("empty.csv",),
            headers=(("email", "impressions"),),
            cells=(),
            row_count=0,
        )
        self.assertEqual(
            "blocked_person_level",
            pre_scan_obvious_privacy(header_only).status,
        )
        values = self.csv_inventory(
            headers=["campaign", "landing_page", "metric"],
            rows=[
                ["first", "https://example.com/path?user=abc", "10"],
                ["second", "192.0.2.1", "20"],
            ],
        )
        decision = pre_scan_obvious_privacy(values)
        self.assertEqual(("ipv4", "url_query"), decision.blocked_categories)
        self.assertNotIn("example.com", repr(decision))

    def test_email_in_last_csv_row_blocks_durable_admission_without_leakage(self):
        source = self.write_bytes(
            "privacy.csv",
            b"campaign,impressions\nsafe,10\nperson@example.com,20\n",
        )
        snapshot = self.snapshot(source)
        inventory = inspect_container(snapshot)
        decision = pre_scan_obvious_privacy(inventory)
        self.assertEqual("blocked_person_level", decision.status)
        with self.assertRaisesRegex(PrivacyAdmissionError, "privacy"):
            admit_source(
                snapshot,
                inventory,
                decision,
                self.adapter_validation(snapshot, inventory),
                self.durable,
            )
        self.assertFalse(self.durable.exists())

    def test_adapter_validation_must_pass_before_durable_admission(self):
        source = self.write_bytes(
            "safe.csv", b"campaign_id,impressions\ncampaign-1,100\n"
        )
        snapshot = self.snapshot(source)
        inventory = inspect_container(snapshot)
        pre_scan = pre_scan_obvious_privacy(inventory)
        with self.assertRaisesRegex(PrivacyAdmissionError, "adapter"):
            admit_source(
                snapshot,
                inventory,
                pre_scan,
                privacy_module.AdapterAdmissionValidation(
                    adapter_id="registered-adapter",
                    adapter_version="1.0.0",
                    source_sha256=snapshot.source_sha256,
                    inventory_sha256=(
                        privacy_module.container_inventory_sha256(inventory)
                    ),
                    profile_sha256="sha256:" + ("0" * 64),
                    adapter_validation_sha256="sha256:" + ("1" * 64),
                    governance_sha256="sha256:" + ("2" * 64),
                    accepted=False,
                    observed_minimum_group_size=None,
                    errors=("missing_registered_group_size",),
                ),
                self.durable,
            )
        self.assertFalse(self.durable.exists())

        admitted = admit_source(
            snapshot,
            inventory,
            pre_scan,
            self.adapter_validation(snapshot, inventory),
            self.durable,
        )
        self.assertEqual(snapshot.source_sha256, admitted.source_sha256)
        self.assertEqual(snapshot.byte_length, admitted.byte_length)
        self.assertEqual(source.read_bytes(), admitted.source_path.read_bytes())

    def test_admission_never_overwrites_an_existing_durable_source(self):
        source = self.write_bytes(
            "safe.csv", b"campaign_id,impressions\ncampaign-1,100\n"
        )
        snapshot = self.snapshot(source)
        inventory = inspect_container(snapshot)
        pre_scan = pre_scan_obvious_privacy(inventory)
        self.durable.parent.mkdir()
        self.durable.write_bytes(b"already admitted")
        with self.assertRaisesRegex(PrivacyAdmissionError, "could not be created"):
            admit_source(
                snapshot,
                inventory,
                pre_scan,
                self.adapter_validation(snapshot, inventory),
                self.durable,
            )
        self.assertEqual(b"already admitted", self.durable.read_bytes())

    def test_mapping_or_unrecognized_adapter_result_cannot_admit(self):
        source = self.write_bytes(
            "safe.csv", b"campaign_id,impressions\ncampaign-1,100\n"
        )
        snapshot = self.snapshot(source)
        inventory = inspect_container(snapshot)
        pre_scan = pre_scan_obvious_privacy(inventory)
        for index, validation in enumerate(
            ({"accepted": True}, object()),
            start=1,
        ):
            with self.subTest(validation=type(validation).__name__):
                destination = self.base / f"untrusted-{index}.bin"
                with self.assertRaisesRegex(
                    PrivacyAdmissionError, "adapter privacy validation"
                ):
                    admit_source(
                        snapshot,
                        inventory,
                        pre_scan,
                        validation,
                        destination,
                    )
                self.assertFalse(destination.exists())

    def test_adapter_validation_is_bound_to_source_and_inventory(self):
        source = self.write_bytes(
            "safe.csv", b"campaign_id,impressions\ncampaign-1,100\n"
        )
        snapshot = self.snapshot(source)
        inventory = inspect_container(snapshot)
        pre_scan = pre_scan_obvious_privacy(inventory)
        valid = self.adapter_validation(snapshot, inventory)
        mismatches = (
            privacy_module.AdapterAdmissionValidation(
                adapter_id=valid.adapter_id,
                adapter_version=valid.adapter_version,
                source_sha256="sha256:" + ("0" * 64),
                inventory_sha256=valid.inventory_sha256,
                profile_sha256=valid.profile_sha256,
                adapter_validation_sha256=(
                    valid.adapter_validation_sha256
                ),
                governance_sha256=valid.governance_sha256,
                accepted=True,
                observed_minimum_group_size=10,
                errors=(),
            ),
            privacy_module.AdapterAdmissionValidation(
                adapter_id=valid.adapter_id,
                adapter_version=valid.adapter_version,
                source_sha256=valid.source_sha256,
                inventory_sha256="sha256:" + ("0" * 64),
                profile_sha256=valid.profile_sha256,
                adapter_validation_sha256=(
                    valid.adapter_validation_sha256
                ),
                governance_sha256=valid.governance_sha256,
                accepted=True,
                observed_minimum_group_size=10,
                errors=(),
            ),
        )
        for index, validation in enumerate(mismatches, start=1):
            with self.subTest(index=index):
                destination = self.base / f"mismatch-{index}.bin"
                with self.assertRaisesRegex(
                    PrivacyAdmissionError, "adapter privacy validation"
                ):
                    admit_source(
                        snapshot,
                        inventory,
                        pre_scan,
                        validation,
                        destination,
                    )
                self.assertFalse(destination.exists())


if __name__ == "__main__":
    unittest.main()
